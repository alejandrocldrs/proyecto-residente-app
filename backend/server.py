from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, Form, Header, Query, Request
from fastapi.responses import StreamingResponse, Response
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from starlette.middleware.gzip import GZipMiddleware
from motor.motor_asyncio import AsyncIOMotorClient, AsyncIOMotorGridFSBucket
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timedelta, timezone
import jwt
import hashlib
from fastapi import UploadFile, File
import pandas as pd
import io
import json
import asyncio
from pywebpush import webpush, WebPushException
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.interval import IntervalTrigger
from apscheduler.triggers.cron import CronTrigger
import pytz
from PIL import Image
import base64
from functools import lru_cache
import time

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]
fs_bucket = AsyncIOMotorGridFSBucket(db, bucket_name="presentations_fs")

# Create the main app
app = FastAPI(title="Puerto ENARM Mazatlán API")

# Add Gzip compression middleware for all responses > 500 bytes
app.add_middleware(GZipMiddleware, minimum_size=500)

# =============================================================================
# SIMPLE IN-MEMORY CACHE FOR STATIC DATA
# =============================================================================
class SimpleCache:
    """Simple TTL-based cache for frequently accessed static data"""
    def __init__(self):
        self._cache: Dict[str, Any] = {}
        self._timestamps: Dict[str, float] = {}
    
    def get(self, key: str, ttl_seconds: int = 300) -> Optional[Any]:
        """Get cached value if not expired"""
        if key in self._cache:
            if time.time() - self._timestamps[key] < ttl_seconds:
                return self._cache[key]
            else:
                # Expired, remove it
                del self._cache[key]
                del self._timestamps[key]
        return None
    
    def set(self, key: str, value: Any):
        """Set cache value with current timestamp"""
        self._cache[key] = value
        self._timestamps[key] = time.time()
    
    def invalidate(self, key: str):
        """Remove specific key from cache"""
        self._cache.pop(key, None)
        self._timestamps.pop(key, None)
    
    def invalidate_prefix(self, prefix: str):
        """Remove all keys starting with prefix"""
        keys_to_remove = [k for k in self._cache.keys() if k.startswith(prefix)]
        for key in keys_to_remove:
            self.invalidate(key)

# Global cache instance
cache = SimpleCache()

# =============================================================================

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")
# Add your routes to the router instead of directly to app
@api_router.get("/")
async def root():
    return {"message": "Puerto ENARM Mazatlán API", "status": "running"}

# JWT and Password settings
SECRET_KEY = os.environ.get("SECRET_KEY")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 30 * 24 * 60  # 30 days

security = HTTPBearer()

# SSE Connection Manager for real-time notifications
class SSEManager:
    def __init__(self):
        # Store active SSE connections for each user
        self.active_connections: Dict[str, asyncio.Queue] = {}
        self.connection_tasks: Dict[str, asyncio.Task] = {}
    
    def add_connection(self, user_id: str) -> asyncio.Queue:
        """Add a new SSE connection queue for user"""
        if user_id in self.active_connections:
            # Remove old connection
            self.remove_connection(user_id)
        
        queue = asyncio.Queue()
        self.active_connections[user_id] = queue
        print(f"SSE connection added for user: {user_id}")
        return queue
    
    def remove_connection(self, user_id: str):
        """Remove SSE connection for user"""
        if user_id in self.active_connections:
            del self.active_connections[user_id]
        if user_id in self.connection_tasks:
            task = self.connection_tasks[user_id]
            if not task.done():
                task.cancel()
            del self.connection_tasks[user_id]
        print(f"SSE connection removed for user: {user_id}")
    
    async def send_message(self, user_id: str, message: dict) -> bool:
        """Send message to specific user via SSE"""
        if user_id in self.active_connections:
            try:
                queue = self.active_connections[user_id]
                await queue.put(message)
                print(f"SSE message queued for user {user_id}: {message['type']}")
                return True
            except Exception as e:
                print(f"Error queuing SSE message for user {user_id}: {e}")
                self.remove_connection(user_id)
                return False
        return False

sse_manager = SSEManager()

# Helper function to compress profile images
def compress_image(content: bytes, max_size: int = 150, quality: int = 70) -> str:
    """Compress image to reduce size. Returns base64 data URL."""
    try:
        img = Image.open(io.BytesIO(content))
        
        # Convert to RGB if necessary (for PNG with transparency)
        if img.mode in ('RGBA', 'P'):
            img = img.convert('RGB')
        
        # Resize if larger than max_size
        if img.width > max_size or img.height > max_size:
            img.thumbnail((max_size, max_size), Image.Resampling.LANCZOS)
        
        # Save to bytes with compression
        output = io.BytesIO()
        img.save(output, format='JPEG', quality=quality, optimize=True)
        output.seek(0)
        
        # Convert to base64
        base64_image = base64.b64encode(output.getvalue()).decode('utf-8')
        return f"data:image/jpeg;base64,{base64_image}"
    except Exception as e:
        logging.error(f"Error compressing image: {e}")
        # Fallback: return original as base64 if compression fails
        base64_image = base64.b64encode(content).decode('utf-8')
        return f"data:image/jpeg;base64,{base64_image}"

# Simple hash functions (for MVP - use better hashing in production)
def get_password_hash(password: str) -> str:
    return hashlib.sha256(password.encode()).hexdigest()

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return get_password_hash(plain_password) == hashed_password

# Pydantic Models
class User(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    full_name: str
    email: EmailStr
    hashed_password: str
    is_admin: bool = False
    is_approved: bool = False
    profile_image: Optional[str] = None  # Base64 encoded image or URL
    gender: Optional[str] = None  # "male" or "female"
    universidad: Optional[str] = None  # University name
    # Escape Room Progress Fields
    cases_completed: int = 0
    cases_successful: int = 0
    current_rank: str = "Estudiante Universitario"
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    subscription_expires: Optional[str] = None

class UserCreate(BaseModel):
    full_name: str
    email: EmailStr
    password: str
    gender: Optional[str] = None
    universidad: Optional[str] = None

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class UserResponse(BaseModel):
    id: str
    full_name: str
    email: str
    is_admin: bool
    is_approved: bool
    profile_image: Optional[str] = None
    gender: Optional[str] = None
    universidad: Optional[str] = None
    subscription_expires: Optional[str] = None
    account_type: Optional[str] = None

class Question(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    question_text: str
    option_a: str
    option_b: str
    option_c: str
    option_d: str
    correct_answer: str  # A, B, C, or D
    explanation: str
    reference: str
    specialty: str  # Ginecología, Cirugía, Pediatría, Medicina Interna, Otros
    topic: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class QuestionCreate(BaseModel):
    question_text: str
    option_a: str
    option_b: str
    option_c: str
    option_d: str
    correct_answer: str
    explanation: str
    reference: str
    specialty: str
    topic: str

class Quiz(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    title: str
    specialty: str
    topic: str
    questions: List[str]  # Question IDs
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class QuizCreate(BaseModel):
    title: str
    specialty: str
    topic: str
    questions: List[str]

class QuizAttempt(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    quiz_id: str
    answers: dict  # {question_id: selected_answer}
    score: float
    total_questions: int
    correct_answers: int
    completed_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class QuizAttemptCreate(BaseModel):
    quiz_id: str
    answers: dict

class QuizProgress(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    quiz_id: str
    current_question_index: int
    answers: dict  # Partial answers
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class QuizProgressCreate(BaseModel):
    quiz_id: str
    current_question_index: int
    answers: dict

class Duel(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    player1_id: str
    player2_id: str
    
    # Pre-generated game data (same for both players)
    round_specialties: List[str] = []  # 5 specialties from wheel spins
    round_questions: List[str] = []    # 5 question IDs (same for both players)
    
    # Player 1 results
    player1_answers: List[str] = []     # Player 1's 5 answers
    player1_completed: bool = False
    player1_score: int = 0
    player1_completed_at: Optional[datetime] = None
    
    # Player 2 results  
    player2_answers: List[str] = []     # Player 2's 5 answers
    player2_completed: bool = False
    player2_score: int = 0
    player2_completed_at: Optional[datetime] = None
    
    # Duel question source and topic
    question_source: str = "duel"  # "gpc" (legacy) or "duel" (new duel_questions collection)
    duel_topic: Optional[str] = None  # None = General, or specific topic
    
    # Messages
    challenger_message: Optional[str] = None  # Message from player1 to player2
    winner_message: Optional[str] = None      # Message from winner (if player2 wins)
    
    # Final results (only revealed when both complete)
    status: str = "waiting_player1"  # waiting_player1, waiting_player2, completed
    winner_id: Optional[str] = None
    forfeit: bool = False  # True if someone abandoned the duel
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    completed_at: Optional[datetime] = None

class DuelQuestion(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    specialty: str  # One of 5 base topics
    materia: str    # Sub-category from CSV
    tema: str       # Specific topic from CSV
    question_text: str
    option_a: str
    option_b: str
    option_c: str
    correct_answer: str  # A, B, or C
    global_usage_count: int = 0
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class DuelCreate(BaseModel):
    player2_email: str
    challenger_message: Optional[str] = None
    duel_topic: Optional[str] = None  # None = "General", or specific topic name

class DuelAnswer(BaseModel):
    duel_id: str
    question_index: int
    selected_answer: str

class DuelInvite(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    from_user_id: str
    to_user_email: str
    status: str = "pending"  # pending, accepted, declined
    created_at: datetime = Field(default_factory=lambda: str(uuid.uuid4()))

# Escape Room / Clinical Cases Models
class ClinicalCase(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    title: str
    theme: str  # Cirugía, Medicina Interna, Pediatría, Ginecología y Obstetricia, Otros
    description: str
    steps: List[Dict]  # 7 steps with questions and alternatives
    global_messages: Dict[str, str]  # MensajeCuracion, MensajeMuerte
    module: str = "Otros"  # Módulo (igual que Cuestionarios)
    submodule: str = "General"  # Submódulo (igual que Cuestionarios)
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: str

class GameSession(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    case_id: str
    current_step: int = 1
    steps_taken: int = 0
    errors_count: int = 0
    alternative_branches_used: List[int] = []
    is_completed: bool = False
    is_successful: bool = False
    final_message: Optional[str] = None
    started_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    completed_at: Optional[datetime] = None

class CaseAnswer(BaseModel):
    session_id: str
    step_number: int
    selected_option: str  # A, B, C, D
    is_alternative: bool = False

# Imagen DX Models
class ImagenDXCase(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    system: str  # Tórax, Abdomen, Neuro, etc. (subtema)
    modality: str  # radiografía, TC, RM, etc.
    finding_or_sign: str  # Respuesta correcta
    interpretation_clave: str
    diagnostico_sugerido: str
    caso_clinico: str  # Campo "CASO CLINICO"
    option_a: str
    option_b: str
    option_c: str
    option_d: str
    image_url: Optional[str] = None  # URL o path de la imagen subida (principal)
    image_urls: Optional[List[str]] = None  # URLs adicionales para Dermatología/Patología (hasta 3)
    link_imagen_referencia: Optional[str] = None
    is_published: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: str
    updated_by: Optional[str] = None

class ImagenDXAnswer(BaseModel):
    case_id: str
    selected_option: str  # A, B, C, D

# Flashcard Models
class Flashcard(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    question_id: Optional[str] = None  # nullable for manual cards
    quiz_id: Optional[str] = None  # nullable for manual cards
    quiz_title: str  # Name of the quiz/deck
    specialty: str  # Tema base: Ginecología y Obstetricia, Cirugía, Pediatría, Medicina Interna, Otros
    topic: str  # Subtema
    question_text: str  # Front of card
    answer_text: str  # Back of card (correct answer)
    explanation: Optional[str] = None
    personal_notes: Optional[str] = None
    is_manual: bool = False
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class FlashcardCreate(BaseModel):
    question_id: str
    quiz_id: str
    quiz_title: str
    specialty: str
    topic: str
    question_text: str
    answer_text: str
    explanation: Optional[str] = None

class FlashcardManualCreate(BaseModel):
    specialty: str
    topic: str
    quiz_title: str  # Can be existing or new deck name
    question_text: str
    answer_text: str
    personal_notes: Optional[str] = None

class FlashcardUpdate(BaseModel):
    question_text: Optional[str] = None
    answer_text: Optional[str] = None
    personal_notes: Optional[str] = None

# ==================== PERLAS DIARIAS MODELS ====================

class PearlModule(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str  # Ginecología y Obstetricia, Cirugía, Pediatría, Medicina Interna, Otros
    description: Optional[str] = None
    is_active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PearlSubtopic(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    module_id: str
    name: str
    description: Optional[str] = None
    is_active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class Pearl(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    subtopic_id: str
    title: str  # Max 40 chars
    message: str  # Max 200 chars - shown in notification
    long_body: Optional[str] = None  # Full explanation when opened
    tags: List[str] = []
    difficulty: Optional[str] = None  # Básico, Intermedio, Avanzado
    is_active: bool = True
    priority: int = 0  # Higher = sent first
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PearlCreate(BaseModel):
    subtopic_id: str
    title: str
    message: str
    long_body: Optional[str] = None
    tags: List[str] = []
    difficulty: Optional[str] = None
    is_active: bool = True
    priority: int = 0

class PushSubscription(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    endpoint: str
    keys: Dict[str, str]  # p256dh, auth
    is_active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    last_success: Optional[datetime] = None
    last_error: Optional[str] = None
    error_count: int = 0

class UserPearlPreference(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    module_id: str
    subtopic_id: str
    interval_minutes: int = 60  # 15, 30, 60, 120 minutes between perlas
    start_hour: str = "08:00"
    end_hour: str = "22:00"
    timezone: str = "America/Mexico_City"
    is_active: bool = True
    is_paused: bool = False
    last_sent_at: Optional[str] = None  # Track when last pearl was sent
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PearlDeliveryLog(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    pearl_id: str
    preference_id: str
    sent_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    status: str = "pending"  # pending, delivered, failed, clicked
    error_message: Optional[str] = None

class PearlPreferenceCreate(BaseModel):
    module_id: str
    subtopic_id: str
    interval_minutes: int = 60  # 15, 30, 60, 120
    start_hour: str = "08:00"
    end_hour: str = "22:00"
    timezone: str = "America/Mexico_City"

class PushSubscriptionCreate(BaseModel):
    endpoint: str
    keys: Dict[str, str]

class Token(BaseModel):
    access_token: str
    token_type: str

# Helper functions already defined above

# Escape Room Rank Calculation
def get_rank_from_successful_cases(successful_cases: int) -> str:
    """Calculate user rank based on successful escape room cases completed"""
    if successful_cases >= 91:
        return "Estas muy viejo"
    elif successful_cases >= 81:
        return "Médico con Alta Especialidad"
    elif successful_cases >= 71:
        return "Médico Subespecialista"
    elif successful_cases >= 61:
        return "Médico Especialista"
    elif successful_cases >= 51:
        return "Residente de Último Año"
    elif successful_cases >= 41:
        return "Residente de 1er Año"
    elif successful_cases >= 31:
        return "Médico General"
    elif successful_cases >= 21:
        return "Médico Pasante"
    elif successful_cases >= 11:
        return "Médico Interno de Pregrado"
    else:
        return "Estudiante Universitario"

def get_rank_progress(successful_cases: int) -> dict:
    """Get progress information for current rank"""
    ranks = [
        (0, "Estudiante Universitario"),
        (11, "Médico Interno de Pregrado"),
        (21, "Médico Pasante"),
        (31, "Médico General"),
        (41, "Residente de 1er Año"),
        (51, "Residente de Último Año"),
        (61, "Médico Especialista"),
        (71, "Médico Subespecialista"),
        (81, "Médico con Alta Especialidad"),
        (91, "Estas muy viejo")
    ]
    
    current_rank = get_rank_from_successful_cases(successful_cases)
    
    # Find current and next rank
    for i, (threshold, rank_name) in enumerate(ranks):
        if successful_cases < threshold:
            if i > 0:
                prev_threshold, prev_rank = ranks[i-1]
                return {
                    "current_rank": prev_rank,
                    "next_rank": rank_name,
                    "current_cases": successful_cases,
                    "cases_for_next_rank": threshold,
                    "progress_percentage": round(((successful_cases - prev_threshold) / (threshold - prev_threshold)) * 100, 1)
                }
    
    # Max rank reached
    return {
        "current_rank": current_rank,
        "next_rank": None,
        "current_cases": successful_cases,
        "cases_for_next_rank": None,
        "progress_percentage": 100
    }

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.now(timezone.utc) + expires_delta
    else:
        expire = datetime.now(timezone.utc) + timedelta(minutes=15)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        email: str = payload.get("sub")
        if email is None:
            raise credentials_exception
    except jwt.PyJWTError:
        raise credentials_exception
    
    user = await db.users.find_one({"email": email})
    if user is None:
        raise credentials_exception
    
    # Check subscription expiry for non-admin users (blocks expired trials/subs mid-session)
    if not user.get("is_admin", False) and user.get("subscription_expires"):
        expires_dt = datetime.fromisoformat(user["subscription_expires"])
        if datetime.now(timezone.utc) > expires_dt:
            raise HTTPException(
                status_code=401,
                detail="subscription_expired"
            )
    
    return User(**user)

async def get_admin_user(current_user: User = Depends(get_current_user)):
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="Not enough permissions")
    return current_user

# Auth Routes
@api_router.post("/auth/register", response_model=dict)
async def register(user: UserCreate):
    # Check if user exists
    existing_user = await db.users.find_one({"email": user.email})
    if existing_user:
        # Allow re-registration for trial users who haven't paid
        if existing_user.get("account_type") == "trial" and existing_user.get("payment_status") != "completed":
            await db.users.delete_one({"email": user.email})
        else:
            raise HTTPException(status_code=400, detail="Email already registered")
    
    hashed_password = get_password_hash(user.password)
    trial_expires = datetime.now(timezone.utc) + timedelta(hours=72)
    new_user = User(
        full_name=user.full_name,
        email=user.email,
        hashed_password=hashed_password,
        is_admin=False,
        is_approved=True,
        gender=user.gender,
        universidad=user.universidad,
        subscription_expires=trial_expires.isoformat()
    )
    
    user_dict = new_user.dict()
    user_dict["temp_password"] = user.password
    user_dict["account_type"] = "trial"
    user_dict["registered_at"] = datetime.now(timezone.utc).isoformat()
    await db.users.insert_one(user_dict)
    return {
        "message": "Registro exitoso. Tienes 72 horas de acceso gratuito.",
        "user_id": new_user.id
    }

@api_router.post("/auth/upload-profile-image-registration/{user_id}")
async def upload_profile_image_during_registration(
    user_id: str,
    file: UploadFile = File(...)
):
    """Allow users to upload profile image during registration (before approval)"""
    if not file.content_type.startswith('image/'):
        raise HTTPException(status_code=400, detail="File must be an image")
    
    # Find user by ID (don't require authentication for registration)
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Read file and compress
    content = await file.read()
    if len(content) > 5 * 1024 * 1024:  # 5MB limit
        raise HTTPException(status_code=400, detail="Image too large. Maximum size is 5MB")
    
    # Compress image to ~150x150 for profile pictures
    image_data_url = compress_image(content, max_size=150, quality=70)
    
    # Update user profile image
    result = await db.users.update_one(
        {"id": user_id},
        {"$set": {"profile_image": image_data_url}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {"message": "Profile image uploaded successfully", "profile_image": image_data_url}

@api_router.post("/auth/login", response_model=dict)
async def login(user: UserLogin):
    db_user = await db.users.find_one({"email": user.email})
    if not db_user or not verify_password(user.password, db_user["hashed_password"]):
        raise HTTPException(status_code=400, detail="Incorrect email or password")
    
    if not db_user.get("is_approved", False):
        raise HTTPException(status_code=400, detail="Account not approved by admin")
    
    # Block login if user paid but hasn't activated from email yet
    if not db_user.get("is_admin", False):
        if db_user.get("payment_status") == "completed" and db_user.get("account_type") == "trial":
            raise HTTPException(
                status_code=403,
                detail={
                    "code": "pending_activation",
                    "user_id": db_user["id"],
                    "email": db_user["email"],
                    "message": "Tu pago fue recibido. Para entrar, debes activar tu cuenta desde el correo que te enviamos."
                }
            )
    
    # Check subscription expiry for non-admin users
    if not db_user.get("is_admin", False):
        subscription_expires = db_user.get("subscription_expires")
        if subscription_expires:
            expires_dt = datetime.fromisoformat(subscription_expires)
            if datetime.now(timezone.utc) > expires_dt:
                account_type = db_user.get("account_type", "trial")
                if account_type == "trial":
                    raise HTTPException(
                        status_code=403,
                        detail={
                            "code": "trial_expired",
                            "user_id": db_user["id"],
                            "email": db_user["email"],
                            "message": "Tu version de prueba de 72 horas ha caducado. Debes pagar para seguir usando la app."
                        }
                    )
                else:
                    raise HTTPException(
                        status_code=403,
                        detail={
                            "code": "subscription_expired",
                            "user_id": db_user["id"],
                            "email": db_user["email"],
                            "message": "Tu suscripcion ha expirado. Debes renovar para seguir usando la app."
                        }
                    )
    
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": user.email}, expires_delta=access_token_expires
    )
    return {"access_token": access_token, "token_type": "bearer"}

@api_router.get("/auth/me", response_model=UserResponse)
async def get_me(current_user: User = Depends(get_current_user)):
    # Fetch account_type from DB (not in User model)
    db_user = await db.users.find_one({"id": current_user.id}, {"account_type": 1, "_id": 0})
    return UserResponse(
        id=current_user.id,
        full_name=current_user.full_name,
        email=current_user.email,
        is_admin=current_user.is_admin,
        is_approved=current_user.is_approved,
        profile_image=current_user.profile_image,
        gender=current_user.gender,
        universidad=current_user.universidad,
        subscription_expires=current_user.subscription_expires,
        account_type=db_user.get("account_type") if db_user else None
    )

@api_router.post("/auth/upload-profile-image")
async def upload_profile_image(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    if not file.content_type.startswith('image/'):
        raise HTTPException(status_code=400, detail="File must be an image")
    
    # Read file and compress
    content = await file.read()
    if len(content) > 5 * 1024 * 1024:  # 5MB limit
        raise HTTPException(status_code=400, detail="Image too large. Maximum size is 5MB")
    
    # Compress image to ~150x150 for profile pictures
    image_data_url = compress_image(content, max_size=150, quality=70)
    
    # Update user profile image
    result = await db.users.update_one(
        {"id": current_user.id},
        {"$set": {"profile_image": image_data_url}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {"message": "Profile image updated successfully", "profile_image": image_data_url}


@api_router.put("/auth/profile")
async def update_profile(data: dict, current_user: User = Depends(get_current_user)):
    """Update user profile fields (name, universidad, gender)."""
    update_fields = {}
    if "full_name" in data and data["full_name"].strip():
        update_fields["full_name"] = data["full_name"].strip()
    if "universidad" in data and data["universidad"].strip():
        update_fields["universidad"] = data["universidad"].strip()
    if "gender" in data and data["gender"] in ("male", "female"):
        update_fields["gender"] = data["gender"]
    if not update_fields:
        raise HTTPException(status_code=400, detail="No hay campos válidos para actualizar")
    await db.users.update_one({"id": current_user.id}, {"$set": update_fields})
    updated = await db.users.find_one({"id": current_user.id}, {"_id": 0, "password": 0})
    return {
        "message": "Perfil actualizado",
        "user": {
            "id": updated["id"],
            "full_name": updated["full_name"],
            "email": updated["email"],
            "is_admin": updated.get("is_admin", False),
            "is_approved": updated.get("is_approved", False),
            "profile_image": updated.get("profile_image"),
            "gender": updated.get("gender"),
            "universidad": updated.get("universidad", ""),
        }
    }


# ============================================================
# POINTS & RANK SYSTEM
# ============================================================

RANKS = [
    {"key": "estudiante_de_medicina", "name": "Estudiante de Medicina", "points": 0},
    {"key": "interno_de_pregrado", "name": "Interno de Pregrado", "points": 200},
    {"key": "medico_pasante", "name": "Médico Pasante de Servicio Social", "points": 600},
    {"key": "medico_general", "name": "Médico General", "points": 1200},
    {"key": "residente_primer_ano", "name": "Residente de Primer Año", "points": 2000},
    {"key": "residente_ultimo_ano", "name": "Residente de Último Año", "points": 3200},
    {"key": "jefe_de_residentes", "name": "Jefe de Residentes", "points": 5000},
    {"key": "medico_especialista", "name": "Médico Especialista", "points": 8000},
    {"key": "subespecialista", "name": "Subespecialista", "points": 14000},
    {"key": "alta_especialidad", "name": "Alta Especialidad", "points": 30000},
    {"key": "maestria_ciencias", "name": "Maestría en Ciencias Médicas", "points": 50000},
    {"key": "doctorado_ciencias", "name": "Doctorado en Ciencias Médicas", "points": 80000},
    {"key": "jefe_de_servicio", "name": "Jefe de Servicio", "points": 120000},
    {"key": "director_hospital", "name": "Director de Hospital", "points": 180000},
    {"key": "secretario_salud", "name": "Secretario de Salud", "points": 260000},
    {"key": "director_oms", "name": "Director General de la OMS", "points": 400000},
    {"key": "premio_nobel", "name": "Premio Nobel de Medicina", "points": 700000},
]

POINT_VALUES = {
    "imagendx": 5,
    "escape_room": 20,
    "quiz": 40,
    "simulacro": 300,
    "duel_win": 15,
    "duel_loss": 5,
}

def get_rank_for_points(total_points: int):
    """Return current rank and next rank info based on total points."""
    current_rank = RANKS[0]
    next_rank = RANKS[1] if len(RANKS) > 1 else None
    
    for i, rank in enumerate(RANKS):
        if total_points >= rank["points"]:
            current_rank = rank
            next_rank = RANKS[i + 1] if i + 1 < len(RANKS) else None
        else:
            break
    
    return current_rank, next_rank

async def award_points(user_id: str, activity_type: str, activity_id: str, base_points: int):
    """Award points with anti-farming logic. Returns points awarded and rank info."""
    now = datetime.now(timezone.utc)
    actual_points = base_points
    is_repeat = False
    
    # Anti-farming: check if activity was done before
    if activity_type in ("quiz", "escape_room", "imagendx"):
        existing = await db.point_transactions.find_one({
            "user_id": user_id,
            "activity_type": activity_type,
            "activity_id": activity_id
        })
        if existing:
            is_repeat = True
            actual_points = base_points // 2  # 50% for repeats
    
    elif activity_type == "simulacro":
        # Only award points first time per day
        start_of_day = now.replace(hour=0, minute=0, second=0, microsecond=0)
        existing_today = await db.point_transactions.find_one({
            "user_id": user_id,
            "activity_type": "simulacro",
            "activity_id": activity_id,
            "created_at": {"$gte": start_of_day}
        })
        if existing_today:
            actual_points = 0  # No points for same simulacro same day
    
    # Duels have no limit (duel_win, duel_loss) - no anti-farming check needed
    
    if actual_points <= 0:
        # Still return current rank info even if no points awarded
        total = await db.point_transactions.aggregate([
            {"$match": {"user_id": user_id}},
            {"$group": {"_id": None, "total": {"$sum": "$points"}}}
        ]).to_list(1)
        total_points = total[0]["total"] if total else 0
        current_rank, next_rank = get_rank_for_points(total_points)
        return {
            "points_awarded": 0,
            "total_points": total_points,
            "current_rank": current_rank,
            "next_rank": next_rank,
            "rank_up": False
        }
    
    # Get old total to detect rank-up
    old_total_result = await db.point_transactions.aggregate([
        {"$match": {"user_id": user_id}},
        {"$group": {"_id": None, "total": {"$sum": "$points"}}}
    ]).to_list(1)
    old_total = old_total_result[0]["total"] if old_total_result else 0
    old_rank, _ = get_rank_for_points(old_total)
    
    # Insert transaction
    await db.point_transactions.insert_one({
        "id": str(uuid.uuid4()),
        "user_id": user_id,
        "activity_type": activity_type,
        "activity_id": activity_id,
        "points": actual_points,
        "base_points": base_points,
        "is_repeat": is_repeat,
        "created_at": now
    })
    
    new_total = old_total + actual_points
    new_rank, next_rank = get_rank_for_points(new_total)
    rank_up = new_rank["key"] != old_rank["key"]
    
    return {
        "points_awarded": actual_points,
        "total_points": new_total,
        "current_rank": new_rank,
        "next_rank": next_rank,
        "rank_up": rank_up,
        "new_rank_name": new_rank["name"] if rank_up else None
    }

@api_router.get("/points/me")
async def get_my_points(current_user: User = Depends(get_current_user)):
    """Get current user's total points, rank, and progress."""
    total_result = await db.point_transactions.aggregate([
        {"$match": {"user_id": current_user.id}},
        {"$group": {"_id": None, "total": {"$sum": "$points"}}}
    ]).to_list(1)
    total_points = total_result[0]["total"] if total_result else 0
    
    current_rank, next_rank = get_rank_for_points(total_points)
    
    return {
        "total_points": total_points,
        "current_rank": current_rank,
        "next_rank": next_rank,
        "all_ranks": RANKS
    }

@api_router.get("/points/history")
async def get_points_history(
    limit: int = 20,
    current_user: User = Depends(get_current_user)
):
    """Get recent point transactions."""
    transactions = await db.point_transactions.find(
        {"user_id": current_user.id},
        {"_id": 0}
    ).sort("created_at", -1).limit(limit).to_list(limit)
    return transactions

@api_router.post("/auth/update-gender")
async def update_gender(data: dict, current_user: User = Depends(get_current_user)):
    """Allow users to update their gender."""
    gender = data.get("gender")
    if gender not in ("male", "female"):
        raise HTTPException(status_code=400, detail="Gender must be 'male' or 'female'")
    await db.users.update_one(
        {"id": current_user.id},
        {"$set": {"gender": gender}}
    )
    return {"message": "Gender updated", "gender": gender}


@api_router.post("/admin/set-points")
async def admin_set_points(data: dict, admin_user: User = Depends(get_admin_user)):
    """Admin-only: Set exact point total for the admin account (for testing ranks)."""
    target_points = data.get("points", 0)
    if not isinstance(target_points, (int, float)) or target_points < 0:
        raise HTTPException(status_code=400, detail="Points must be a non-negative number")
    target_points = int(target_points)
    
    # Get current total
    total_result = await db.point_transactions.aggregate([
        {"$match": {"user_id": admin_user.id}},
        {"$group": {"_id": None, "total": {"$sum": "$points"}}}
    ]).to_list(1)
    current_total = total_result[0]["total"] if total_result else 0
    
    # Delete all existing transactions for admin
    await db.point_transactions.delete_many({"user_id": admin_user.id})
    
    # Insert a single transaction with the exact target
    if target_points > 0:
        await db.point_transactions.insert_one({
            "id": str(uuid.uuid4()),
            "user_id": admin_user.id,
            "activity_type": "admin_set",
            "activity_id": "manual",
            "points": target_points,
            "base_points": target_points,
            "is_repeat": False,
            "created_at": datetime.now(timezone.utc)
        })
    
    current_rank, next_rank = get_rank_for_points(target_points)
    
    return {
        "message": f"Points set to {target_points}",
        "previous_points": current_total,
        "total_points": target_points,
        "current_rank": current_rank,
        "next_rank": next_rank
    }



@api_router.get("/admin/pending-users", response_model=List[UserResponse])
async def get_pending_users(admin_user: User = Depends(get_admin_user)):
    users = await db.users.find({
        "is_admin": {"$ne": True},
        "$or": [
            {"account_type": "trial"},
            {"account_type": {"$exists": False}, "is_approved": False}
        ]
    }).to_list(1000)
    return [UserResponse(**user) for user in users]

@api_router.get("/admin/approved-users", response_model=List[UserResponse])
async def get_approved_users(admin_user: User = Depends(get_admin_user)):
    users = await db.users.find({
        "is_admin": {"$ne": True},
        "$or": [
            {"account_type": "paid"},
            {"account_type": {"$exists": False}, "is_approved": True}
        ]
    }).to_list(1000)
    return [UserResponse(**user) for user in users]

@api_router.patch("/admin/approve-user/{user_id}")
async def approve_user(user_id: str, admin_user: User = Depends(get_admin_user)):
    subscription_expires = datetime.now(timezone.utc) + timedelta(days=180)
    result = await db.users.update_one(
        {"id": user_id},
        {"$set": {
            "is_approved": True,
            "subscription_expires": subscription_expires.isoformat(),
            "activated_at": datetime.now(timezone.utc).isoformat(),
            "account_type": "paid"
        }}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    return {"message": "User approved successfully"}

@api_router.delete("/admin/reject-user/{user_id}")
async def reject_user(user_id: str, admin_user: User = Depends(get_admin_user)):
    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    return {"message": "User rejected and deleted"}

# Question Routes
@api_router.post("/admin/questions", response_model=Question)
async def create_question(question: QuestionCreate, admin_user: User = Depends(get_admin_user)):
    new_question = Question(**question.dict(), created_at=datetime.now(timezone.utc))
    await db.questions.insert_one(new_question.dict())
    return new_question

@api_router.get("/questions", response_model=List[Question])
async def get_questions(
    specialty: Optional[str] = None,
    topic: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    filter_dict = {}
    if specialty:
        filter_dict["specialty"] = specialty
    if topic:
        filter_dict["topic"] = topic
    
    questions = await db.questions.find(filter_dict).to_list(30000)
    return [Question(**question) for question in questions]

@api_router.get("/questions/{question_id}", response_model=Question)
async def get_question(question_id: str, current_user: User = Depends(get_current_user)):
    question = await db.questions.find_one({"id": question_id})
    if not question:
        raise HTTPException(status_code=404, detail="Question not found")
    return Question(**question)

@api_router.put("/admin/questions/{question_id}", response_model=Question)
async def update_question(
    question_id: str, 
    question_update: QuestionCreate, 
    admin_user: User = Depends(get_admin_user)
):
    result = await db.questions.update_one(
        {"id": question_id},
        {"$set": question_update.dict()}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Question not found")
    
    updated_question = await db.questions.find_one({"id": question_id})
    return Question(**updated_question)

@api_router.delete("/admin/questions/{question_id}")
async def delete_question(question_id: str, admin_user: User = Depends(get_admin_user)):
    result = await db.questions.delete_one({"id": question_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Question not found")
    return {"message": "Question deleted successfully"}

@api_router.post("/admin/questions/delete-multiple")
async def delete_multiple_questions(
    question_ids: List[str],
    current_user: User = Depends(get_admin_user)
):
    """Delete multiple questions at once (admin only)"""
    if not question_ids:
        raise HTTPException(status_code=400, detail="No question IDs provided")
    
    result = await db.questions.delete_many({"id": {"$in": question_ids}})
    
    return {
        "message": f"Successfully deleted {result.deleted_count} questions",
        "deleted_count": result.deleted_count
    }

@api_router.post("/admin/import-questions")
async def import_questions_from_file(
    file: UploadFile = File(...),
    specialty: str = None,
    admin_user: User = Depends(get_admin_user)
):
    if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
        raise HTTPException(status_code=400, detail="File must be Excel (.xlsx, .xls) or CSV (.csv)")
    
    try:
        # Read file content
        content = await file.read()
        
        if file.filename.endswith('.csv'):
            # Try multiple encodings for CSV files
            encodings = ['utf-8', 'utf-8-sig', 'latin1', 'cp1252', 'iso-8859-1']
            df = None
            
            for encoding in encodings:
                try:
                    df = pd.read_csv(io.StringIO(content.decode(encoding)))
                    break
                except (UnicodeDecodeError, UnicodeError):
                    continue
            
            if df is None:
                raise HTTPException(
                    status_code=400, 
                    detail="No se pudo leer el archivo CSV. Asegúrate de que esté guardado en formato UTF-8."
                )
        else:
            df = pd.read_excel(io.BytesIO(content))
        
        # Validate required columns
        required_columns = [
            'question_text', 'option_a', 'option_b', 'option_c', 'option_d',
            'correct_answer', 'explanation', 'reference', 'specialty', 'topic'
        ]
        
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise HTTPException(
                status_code=400, 
                detail=f"Missing required columns: {', '.join(missing_columns)}"
            )
        
        # Process and import questions
        imported_count = 0
        errors = []
        
        print(f"[IMPORT] Starting import with specialty override: {specialty}")
        print(f"[IMPORT] Total rows in CSV: {len(df)}")
        
        for index, row in df.iterrows():
            try:
                # Validate row data
                if pd.isna(row['question_text']) or not row['question_text'].strip():
                    errors.append(f"Row {index + 2}: Question text is empty")
                    continue
                
                if row['correct_answer'] not in ['A', 'B', 'C', 'D']:
                    errors.append(f"Row {index + 2}: Correct answer must be A, B, C, or D")
                    continue
                
                # Use specialty from query parameter if provided, otherwise use from file
                specialty_value = specialty if specialty else str(row['specialty']).strip()
                
                if index < 2:  # Debug first 2 rows
                    print(f"[IMPORT] Row {index}: specialty={specialty_value}, topic={row['topic']}")
                
                # Create question object
                question = Question(
                    question_text=str(row['question_text']).strip(),
                    option_a=str(row['option_a']).strip(),
                    option_b=str(row['option_b']).strip(),
                    option_c=str(row['option_c']).strip(),
                    option_d=str(row['option_d']).strip(),
                    correct_answer=str(row['correct_answer']).strip().upper(),
                    explanation=str(row['explanation']).strip(),
                    reference=str(row['reference']).strip(),
                    specialty=specialty_value,
                    topic=str(row['topic']).strip()
                )
                
                # Insert into database
                question_dict = question.dict()
                result = await db.questions.insert_one(question_dict)
                imported_count += 1
                
                if index < 2:  # Debug first 2 inserts
                    print(f"[IMPORT] Successfully inserted question {index} with ID: {result.inserted_id}")
                
            except Exception as e:
                print(f"[IMPORT ERROR] Row {index + 2}: {str(e)}")
                errors.append(f"Row {index + 2}: {str(e)}")
        
        return {
            "message": f"Import completed. {imported_count} questions imported successfully.",
            "imported_count": imported_count,
            "errors": errors[:10] if errors else []  # Show first 10 errors
        }
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error processing file: {str(e)}")

@api_router.get("/admin/import-template")
async def download_import_template(admin_user: User = Depends(get_admin_user)):
    """Download Excel template for importing questions"""
    template_data = {
        'question_text': ['¿Ejemplo de pregunta médica?'],
        'option_a': ['Primera opción'],
        'option_b': ['Segunda opción'],
        'option_c': ['Tercera opción'],
        'option_d': ['Cuarta opción'],
        'correct_answer': ['A'],
        'explanation': ['Explicación detallada de por qué la respuesta A es correcta'],
        'reference': ['Libro de referencia, páginas, etc.'],
        'specialty': ['Medicina Interna'],
        'topic': ['Tema específico']
    }
    
    return {
        "template_columns": list(template_data.keys()),
        "example_row": template_data,
        "instructions": {
            "1": "Descarga este template y llénalo con tus preguntas",
            "2": "Asegúrate que correct_answer sea A, B, C o D",
            "3": "Specialty debe coincidir con: Ginecología y Obstetricia, Cirugía, Pediatría, Medicina Interna, Otros",
            "4": "Guarda como Excel (.xlsx) o CSV y súbelo en el panel de administración"
        }
    }

# Quiz Routes - Auto-create quizzes by topic
@api_router.get("/quizzes", response_model=List[Quiz])
async def get_quizzes(
    specialty: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    # Get questions and group by specialty and topic to create dynamic quizzes
    filter_dict = {}
    if specialty:
        filter_dict["specialty"] = specialty
    
    questions = await db.questions.find(filter_dict).to_list(30000)
    
    # Group questions by specialty and topic
    quizzes_by_topic = {}
    for question in questions:
        key = f"{question['specialty']}|{question['topic']}"
        if key not in quizzes_by_topic:
            quizzes_by_topic[key] = {
                "specialty": question["specialty"],
                "topic": question["topic"],
                "questions": []
            }
        quizzes_by_topic[key]["questions"].append(question["id"])
    
    # Create quiz objects
    quizzes = []
    for key, data in quizzes_by_topic.items():
        quiz_id = hashlib.sha256(key.encode()).hexdigest()[:16]
        quiz = Quiz(
            id=quiz_id,
            title=f"{data['topic']} - {data['specialty']}",
            specialty=data["specialty"],
            topic=data["topic"],
            questions=data["questions"],
            created_by="system"
        )
        quizzes.append(quiz)
    
    return quizzes

@api_router.get("/quizzes/{quiz_id}")
async def get_quiz_with_questions(quiz_id: str, current_user: User = Depends(get_current_user)):
    # First try to find an existing quiz
    quiz = await db.quizzes.find_one({"id": quiz_id})
    
    if not quiz:
        # Dynamic quiz creation: get ALL distinct specialty|topic combinations
        # and match against quiz_id hash
        # IMPORTANT: Set limit high to get ALL questions
        all_questions = await db.questions.find().to_list(30000)
        
        # Build a map of hash -> (specialty, topic)
        hash_map = {}
        for question in all_questions:
            key = f"{question['specialty']}|{question['topic']}"
            test_id = hashlib.sha256(key.encode()).hexdigest()[:16]
            if test_id not in hash_map:
                hash_map[test_id] = {
                    'specialty': question['specialty'],
                    'topic': question['topic']
                }
        
        # Check if our quiz_id exists in the map
        if quiz_id in hash_map:
            match_data = hash_map[quiz_id]
            # Get all questions for this specialty/topic
            topic_questions = await db.questions.find({
                "specialty": match_data['specialty'],
                "topic": match_data['topic']
            }).to_list(30000)
            
            if topic_questions:
                quiz_data = {
                    "id": quiz_id,
                    "title": f"{match_data['topic']} - {match_data['specialty']}",
                    "specialty": match_data['specialty'],
                    "topic": match_data['topic'],
                    "questions": [q["id"] for q in topic_questions],
                    "created_by": "system"
                }
                
                return {
                    "quiz": Quiz(**quiz_data),
                    "questions": [Question(**q) for q in topic_questions]
                }
        
        raise HTTPException(status_code=404, detail="Quiz not found")
    
    # Get questions for existing quiz
    questions = await db.questions.find({"id": {"$in": quiz["questions"]}}).to_list(30000)
    
    return {
        "quiz": Quiz(**quiz),
        "questions": [Question(**q) for q in questions]
    }

# Quiz Attempt Routes
@api_router.post("/quiz-attempts", response_model=QuizAttempt)
async def submit_quiz_attempt(
    attempt: QuizAttemptCreate, 
    current_user: User = Depends(get_current_user)
):
    # Try to find existing quiz first
    quiz = await db.quizzes.find_one({"id": attempt.quiz_id})
    
    if not quiz:
        # Handle dynamic quiz: build hash map
        # IMPORTANT: Set limit high to get ALL questions
        all_questions = await db.questions.find().to_list(30000)
        
        hash_map = {}
        for question in all_questions:
            key = f"{question['specialty']}|{question['topic']}"
            test_id = hashlib.sha256(key.encode()).hexdigest()[:16]
            if test_id not in hash_map:
                hash_map[test_id] = {
                    'specialty': question['specialty'],
                    'topic': question['topic']
                }
        
        if attempt.quiz_id in hash_map:
            match_data = hash_map[attempt.quiz_id]
            questions = await db.questions.find({
                "specialty": match_data['specialty'],
                "topic": match_data['topic']
            }).to_list(30000)
        else:
            raise HTTPException(status_code=404, detail="Quiz not found")
        
        if not questions:
            raise HTTPException(status_code=404, detail="Quiz not found")
    else:
        # Get questions for existing quiz
        questions = await db.questions.find({"id": {"$in": quiz["questions"]}}).to_list(30000)
    
    if not questions:
        raise HTTPException(status_code=404, detail="No questions found for this quiz")
    
    # Calculate score
    correct_answers = 0
    total_questions = len(questions)
    
    for question in questions:
        if attempt.answers.get(question["id"]) == question["correct_answer"]:
            correct_answers += 1
    
    score = (correct_answers / total_questions) * 100 if total_questions > 0 else 0
    
    new_attempt = QuizAttempt(
        user_id=current_user.id,
        quiz_id=attempt.quiz_id,
        answers=attempt.answers,
        score=score,
        total_questions=total_questions,
        correct_answers=correct_answers
    )
    
    await db.quiz_attempts.insert_one(new_attempt.dict())
    
    # If passed (>=80%), update user_progress for pass_count tracking
    if score >= 80:
        await db.user_progress.update_one(
            {"user_id": current_user.id, "item_id": attempt.quiz_id, "item_type": "quiz"},
            {
                "$set": {
                    "user_id": current_user.id,
                    "item_id": attempt.quiz_id,
                    "item_type": "quiz",
                    "passed": True,
                    "completed_at": datetime.now(timezone.utc).isoformat()
                },
                "$inc": {"pass_count": 1}
            },
            upsert=True
        )
    
    # Award points for completing the quiz (regardless of score)
    points_result = await award_points(current_user.id, "quiz", attempt.quiz_id, POINT_VALUES["quiz"])
    
    result = new_attempt.dict()
    result["points_result"] = points_result
    return result

@api_router.get("/quiz-attempts/me")
async def get_my_attempts(
    current_user: User = Depends(get_current_user),
    page: int = 1,
    page_size: int = 50
):
    """Get user's quiz attempts (paginated)"""
    page = max(1, page)
    page_size = min(max(1, page_size), 100)
    skip = (page - 1) * page_size
    
    total_count = await db.quiz_attempts.count_documents({"user_id": current_user.id})
    
    attempts = await db.quiz_attempts.find(
        {"user_id": current_user.id}
    ).sort("completed_at", -1).skip(skip).limit(page_size).to_list(page_size)
    
    return {
        "attempts": [QuizAttempt(**attempt) for attempt in attempts],
        "pagination": {
            "page": page,
            "page_size": page_size,
            "total_count": total_count,
            "total_pages": (total_count + page_size - 1) // page_size if total_count > 0 else 1
        }
    }

# Quiz Progress Routes
@api_router.post("/quiz-progress")
async def save_quiz_progress(
    progress: QuizProgressCreate,
    current_user: User = Depends(get_current_user)
):
    # Check if progress already exists for this user and quiz
    existing_progress = await db.quiz_progress.find_one({
        "user_id": current_user.id,
        "quiz_id": progress.quiz_id
    })
    
    if existing_progress:
        # Update existing progress
        await db.quiz_progress.update_one(
            {"user_id": current_user.id, "quiz_id": progress.quiz_id},
            {
                "$set": {
                    "current_question_index": progress.current_question_index,
                    "answers": progress.answers,
                    "updated_at": datetime.now(timezone.utc)
                }
            }
        )
        return {"message": "Progress updated"}
    else:
        # Create new progress
        new_progress = QuizProgress(
            user_id=current_user.id,
            quiz_id=progress.quiz_id,
            current_question_index=progress.current_question_index,
            answers=progress.answers
        )
        await db.quiz_progress.insert_one(new_progress.dict())
        return {"message": "Progress saved"}

@api_router.get("/quiz-progress/{quiz_id}")
async def get_quiz_progress(
    quiz_id: str,
    current_user: User = Depends(get_current_user)
):
    progress = await db.quiz_progress.find_one({
        "user_id": current_user.id,
        "quiz_id": quiz_id
    })
    
    if progress:
        return QuizProgress(**progress)
    else:
        return None

@api_router.delete("/quiz-progress/{quiz_id}")
async def delete_quiz_progress(
    quiz_id: str,
    current_user: User = Depends(get_current_user)
):
    await db.quiz_progress.delete_one({
        "user_id": current_user.id,
        "quiz_id": quiz_id
    })
    return {"message": "Progress deleted"}

@api_router.get("/leaderboard")
async def get_leaderboard(current_user: User = Depends(get_current_user)):
    pipeline = [
        {
            "$group": {
                "_id": "$user_id",
                "total_score": {"$sum": "$score"},
                "total_attempts": {"$sum": 1},
                "avg_score": {"$avg": "$score"}
            }
        },
        {
            "$sort": {"total_score": -1}
        },
        {
            "$limit": 10
        }
    ]
    
    leaderboard_data = await db.quiz_attempts.aggregate(pipeline).to_list(10)
    
    # Get user details
    for entry in leaderboard_data:
        user = await db.users.find_one({"id": entry["_id"]})
        if user:
            entry["full_name"] = user["full_name"]
            entry["profile_image"] = user.get("profile_image")
        else:
            entry["full_name"] = "Unknown User"
            entry["profile_image"] = None
    
    return leaderboard_data

@api_router.get("/ranking/top10")
async def get_ranking_top10(current_user: User = Depends(get_current_user)):
    """Get top 10 users by gamification points with rank info and university."""
    pipeline = [
        {
            "$group": {
                "_id": "$user_id",
                "total_points": {"$sum": "$points"}
            }
        },
        {
            "$sort": {"total_points": -1}
        },
        {
            "$limit": 10
        }
    ]
    
    top_users = await db.point_transactions.aggregate(pipeline).to_list(10)
    
    result = []
    for entry in top_users:
        user = await db.users.find_one({"id": entry["_id"]}, {"_id": 0})
        if user:
            current_rank, next_rank = get_rank_for_points(entry["total_points"])
            result.append({
                "user_id": entry["_id"],
                "full_name": user.get("full_name", "Unknown"),
                "universidad": user.get("universidad", ""),
                "profile_image": user.get("profile_image"),
                "total_points": entry["total_points"],
                "rank_name": current_rank["name"],
                "rank_key": current_rank["key"]
            })
    
    return result


def get_daily_cutoff_utc():
    """Get the last 9 PM CDMX cutoff as UTC datetime."""
    cdmx_tz = pytz.timezone("America/Mexico_City")
    now_cdmx = datetime.now(cdmx_tz)
    cutoff_cdmx = now_cdmx.replace(hour=21, minute=0, second=0, microsecond=0)
    if now_cdmx < cutoff_cdmx:
        cutoff_cdmx -= timedelta(days=1)
    return cutoff_cdmx.astimezone(timezone.utc)


@api_router.get("/ranking/daily-top10")
async def get_daily_ranking(current_user: User = Depends(get_current_user)):
    """Get top 10 users by daily score since last 9 PM CDMX.
    score = (cuestionarios×25) + (duelos_ganados×5) + (escape_rooms×8) + (simulacros×50) + (imagendx×2)
    """
    cutoff = get_daily_cutoff_utc()

    pipeline = [
        {"$match": {
            "created_at": {"$gte": cutoff},
            "activity_type": {"$in": ["quiz", "duel_win", "escape_room", "simulacro", "imagendx"]}
        }},
        {"$group": {
            "_id": "$user_id",
            "quiz_count": {"$sum": {"$cond": [{"$eq": ["$activity_type", "quiz"]}, 1, 0]}},
            "duel_win_count": {"$sum": {"$cond": [{"$eq": ["$activity_type", "duel_win"]}, 1, 0]}},
            "escape_room_count": {"$sum": {"$cond": [{"$eq": ["$activity_type", "escape_room"]}, 1, 0]}},
            "simulacro_count": {"$sum": {"$cond": [{"$eq": ["$activity_type", "simulacro"]}, 1, 0]}},
            "imagendx_count": {"$sum": {"$cond": [{"$eq": ["$activity_type", "imagendx"]}, 1, 0]}},
        }},
        {"$addFields": {
            "score": {"$add": [
                {"$multiply": ["$quiz_count", 25]},
                {"$multiply": ["$duel_win_count", 5]},
                {"$multiply": ["$escape_room_count", 8]},
                {"$multiply": ["$simulacro_count", 50]},
                {"$multiply": ["$imagendx_count", 2]}
            ]}
        }},
        {"$sort": {"score": -1}},
        {"$limit": 10}
    ]

    top_users = await db.point_transactions.aggregate(pipeline).to_list(10)

    result = []
    for entry in top_users:
        if entry["score"] <= 0:
            continue
        user = await db.users.find_one({"id": entry["_id"]}, {"_id": 0})
        if user:
            total_points = user.get("total_points", 0)
            current_rank, _ = get_rank_for_points(total_points)
            result.append({
                "user_id": entry["_id"],
                "full_name": user.get("full_name", "Unknown"),
                "universidad": user.get("universidad", ""),
                "profile_image": user.get("profile_image"),
                "rank_name": current_rank["name"],
                "score": entry["score"],
                "quiz_count": entry["quiz_count"],
                "duel_win_count": entry["duel_win_count"],
                "escape_room_count": entry["escape_room_count"],
                "simulacro_count": entry["simulacro_count"],
                "imagendx_count": entry["imagendx_count"],
            })

    return result


@api_router.delete("/admin/reset-leaderboard")
async def reset_leaderboard(admin_user: User = Depends(get_admin_user)):
    await db.quiz_attempts.delete_many({})
    return {"message": "Leaderboard reset successfully"}

@api_router.post("/admin/seed-ranking-users")
async def seed_ranking_users(admin_user: User = Depends(get_admin_user)):
    """Create 10 test users with different point levels to populate the ranking."""
    test_users = [
        {"name": "Dr. Carlos Martínez", "uni": "Universidad Nacional Autónoma de México", "points": 45000, "gender": "male"},
        {"name": "Dra. Ana López", "uni": "Universidad de Guadalajara", "points": 32000, "gender": "female"},
        {"name": "Dr. Luis Hernández", "uni": "Universidad Autónoma de Nuevo León", "points": 28500, "gender": "male"},
        {"name": "Dra. María Torres", "uni": "Universidad Autónoma de Puebla", "points": 21000, "gender": "female"},
        {"name": "Dr. Daniel García", "uni": "Universidad de Monterrey", "points": 15000, "gender": "male"},
        {"name": "Dra. Fernanda Ruiz", "uni": "Universidad Panamericana", "points": 9500, "gender": "female"},
        {"name": "Dr. Ricardo Navarro", "uni": "Universidad Anáhuac", "points": 6000, "gender": "male"},
        {"name": "Dra. Valeria Soto", "uni": "Universidad La Salle", "points": 3500, "gender": "female"},
        {"name": "Dr. Jorge Pineda", "uni": "Universidad Veracruzana", "points": 1500, "gender": "male"},
        {"name": "Dra. Sofía Delgado", "uni": "Universidad de Sonora", "points": 800, "gender": "female"},
    ]
    
    created = []
    for tu in test_users:
        email = tu["name"].lower().replace("dr. ", "").replace("dra. ", "").replace(" ", ".") + "@test.com"
        existing = await db.users.find_one({"email": email})
        if existing:
            user_id = existing["id"]
        else:
            hashed = get_password_hash("test123")
            new_user = User(
                full_name=tu["name"],
                email=email,
                hashed_password=hashed,
                is_admin=False,
                is_approved=True,
                gender=tu["gender"],
                universidad=tu["uni"]
            )
            await db.users.insert_one(new_user.dict())
            user_id = new_user.id
        
        # Clear existing points for this user and set new total
        await db.point_transactions.delete_many({"user_id": user_id})
        await db.point_transactions.insert_one({
            "id": str(uuid.uuid4()),
            "user_id": user_id,
            "activity_type": "seed",
            "activity_id": "seed_ranking",
            "points": tu["points"],
            "base_points": tu["points"],
            "is_repeat": False,
            "created_at": datetime.now(timezone.utc)
        })
        
        created.append({"name": tu["name"], "points": tu["points"]})
    
    return {"message": f"Seeded {len(created)} ranking users", "users": created}

# SSE endpoint for real-time notifications
@api_router.get("/notifications/stream/{user_id}")
async def notification_stream(user_id: str, token: str = None):
    """Server-Sent Events endpoint for real-time notifications"""
    
    # Authenticate via token query parameter (EventSource doesn't support custom headers)
    if not token:
        raise HTTPException(status_code=401, detail="Token required")
    
    try:
        # Decode token
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        email = payload.get("sub")
        if not email:
            raise HTTPException(status_code=401, detail="Invalid token")
        
        # Get user from database by email (since JWT contains email in sub field)
        user_data = await db.users.find_one({"email": email})
        if not user_data:
            raise HTTPException(status_code=401, detail="User not found")
        
        current_user = User(**user_data)
        
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")
    
    # Verify user can only access their own notifications
    if current_user.id != user_id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Create SSE connection queue
    queue = sse_manager.add_connection(user_id)
    
    async def event_stream():
        try:
            # Send initial connection confirmation
            yield f"data: {json.dumps({'type': 'connected', 'message': 'Notification stream connected'})}\n\n"
            
            while True:
                try:
                    # Wait for messages with timeout to send heartbeat
                    message = await asyncio.wait_for(queue.get(), timeout=30.0)
                    yield f"data: {json.dumps(message)}\n\n"
                except asyncio.TimeoutError:
                    # Send heartbeat to keep connection alive
                    yield f"data: {json.dumps({'type': 'heartbeat', 'timestamp': datetime.now(timezone.utc).isoformat()})}\n\n"
                
        except Exception as e:
            print(f"SSE stream error for user {user_id}: {e}")
        finally:
            # Clean up connection
            sse_manager.remove_connection(user_id)
            print(f"SSE stream ended for user {user_id}")
    
    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Headers": "Cache-Control"
        }
    )

# =============================================================================
# DUEL QUESTIONS MANAGEMENT (Admin)
# =============================================================================

@api_router.post("/admin/duel-questions/import")
async def import_duel_questions(
    file: UploadFile = File(...),
    specialty: str = Form(...),
    admin_user: User = Depends(get_admin_user)
):
    """Import duel questions from CSV. Specialty is one of the 5 base topics."""
    valid_specialties = ["Ginecología y Obstetricia", "Cirugía", "Pediatría", "Medicina Interna", "Otros"]
    if specialty not in valid_specialties:
        raise HTTPException(status_code=400, detail=f"Especialidad inválida. Debe ser una de: {', '.join(valid_specialties)}")
    
    if not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="El archivo debe ser CSV (.csv)")
    
    content = await file.read()
    # Try multiple encodings
    for encoding in ['utf-8', 'latin-1', 'cp1252']:
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise HTTPException(status_code=400, detail="No se pudo leer el archivo. Verifique la codificación.")
    
    try:
        df = pd.read_csv(io.StringIO(text))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error al parsear CSV: {str(e)}")
    
    # Normalize column names
    df.columns = [col.strip().lower() for col in df.columns]
    
    required_cols = ['materia', 'tema', 'frase', 'opcion a', 'opcion b', 'opcion c', 'letra respuesta correcta']
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise HTTPException(status_code=400, detail=f"Columnas faltantes: {', '.join(missing)}. Esperadas: Materia, Tema, frase, opcion A, opcion B, opcion C, letra respuesta correcta")
    
    imported = 0
    errors = 0
    for _, row in df.iterrows():
        try:
            answer = str(row['letra respuesta correcta']).strip().upper()
            if answer not in ['A', 'B', 'C']:
                errors += 1
                continue
            
            q = DuelQuestion(
                specialty=specialty,
                materia=str(row['materia']).strip(),
                tema=str(row['tema']).strip(),
                question_text=str(row['frase']).strip(),
                option_a=str(row['opcion a']).strip(),
                option_b=str(row['opcion b']).strip(),
                option_c=str(row['opcion c']).strip(),
                correct_answer=answer
            )
            await db.duel_questions.insert_one(q.dict())
            imported += 1
        except Exception:
            errors += 1
    
    return {
        "message": f"Importación completada: {imported} preguntas importadas, {errors} errores",
        "imported": imported,
        "errors": errors,
        "specialty": specialty
    }

@api_router.get("/admin/duel-questions")
async def get_duel_questions(
    specialty: Optional[str] = None,
    page: int = 1,
    page_size: int = 50,
    admin_user: User = Depends(get_admin_user)
):
    """Get duel questions with optional specialty filter and pagination."""
    filter_dict = {}
    if specialty:
        filter_dict["specialty"] = specialty
    
    total = await db.duel_questions.count_documents(filter_dict)
    skip = (page - 1) * page_size
    questions = await db.duel_questions.find(filter_dict, {"_id": 0}).sort("created_at", -1).skip(skip).limit(page_size).to_list(page_size)
    
    return {
        "questions": questions,
        "pagination": {
            "page": page,
            "page_size": page_size,
            "total_count": total,
            "total_pages": (total + page_size - 1) // page_size
        }
    }

@api_router.get("/admin/duel-questions/stats")
async def get_duel_questions_stats(admin_user: User = Depends(get_admin_user)):
    """Get question counts per specialty."""
    pipeline = [
        {"$group": {"_id": "$specialty", "count": {"$sum": 1}}},
        {"$sort": {"_id": 1}}
    ]
    stats = await db.duel_questions.aggregate(pipeline).to_list(10)
    total = sum(s["count"] for s in stats)
    return {"stats": stats, "total": total}

@api_router.delete("/admin/duel-questions/{question_id}")
async def delete_duel_question(question_id: str, admin_user: User = Depends(get_admin_user)):
    result = await db.duel_questions.delete_one({"id": question_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Pregunta no encontrada")
    return {"message": "Pregunta eliminada"}

@api_router.delete("/admin/duel-questions/bulk/{specialty}")
async def delete_duel_questions_bulk(specialty: str, admin_user: User = Depends(get_admin_user)):
    """Delete all duel questions for a specialty."""
    result = await db.duel_questions.delete_many({"specialty": specialty})
    return {"message": f"{result.deleted_count} preguntas eliminadas de {specialty}", "deleted_count": result.deleted_count}

@api_router.get("/duel-questions/{question_id}")
async def get_duel_question(question_id: str, current_user: User = Depends(get_current_user)):
    """Get a single duel question by ID (for gameplay)."""
    question = await db.duel_questions.find_one({"id": question_id}, {"_id": 0})
    if not question:
        raise HTTPException(status_code=404, detail="Pregunta de duelo no encontrada")
    return question

# =============================================================================
# SIMULACROS (Mock Exams)
# =============================================================================

class SimulacroCreate(BaseModel):
    title: str
    description: Optional[str] = None

@api_router.post("/admin/simulacros/import")
async def import_simulacro(
    file: UploadFile = File(...),
    title: str = Form(...),
    admin_user: User = Depends(get_admin_user)
):
    """Import a simulacro from an Excel file."""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="El archivo debe ser Excel (.xlsx)")
    
    import openpyxl
    content = await file.read()
    
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error al leer archivo Excel: {str(e)}")
    
    # Parse questions
    questions = []
    current_case_number = None
    current_case_text = ""
    question_index = 0
    
    for row in range(2, ws.max_row + 1):
        caso_num = ws.cell(row, 1).value
        caso_clinico = ws.cell(row, 2).value
        pregunta = ws.cell(row, 3).value
        opcion_a = ws.cell(row, 4).value
        opcion_b = ws.cell(row, 5).value
        opcion_c = ws.cell(row, 6).value
        opcion_d = ws.cell(row, 7).value
        respuesta = ws.cell(row, 8).value
        especialidad = ws.cell(row, 9).value
        tema = ws.cell(row, 10).value
        
        # Skip empty rows
        if not pregunta:
            continue
        
        # Update case text when new case number appears
        if caso_num is not None:
            current_case_number = caso_num
            current_case_text = str(caso_clinico or '').strip()
        
        questions.append({
            "index": question_index,
            "case_number": current_case_number,
            "case_text": current_case_text,
            "question_text": str(pregunta or '').strip(),
            "option_a": str(opcion_a or '').strip(),
            "option_b": str(opcion_b or '').strip(),
            "option_c": str(opcion_c or '').strip(),
            "option_d": str(opcion_d or '').strip() if opcion_d else '',
            "correct_answer": str(respuesta or '').strip().upper(),
            "especialidad": str(especialidad or '').strip(),
            "tema": str(tema or '').strip()
        })
        question_index += 1
    
    if not questions:
        raise HTTPException(status_code=400, detail="No se encontraron preguntas en el archivo")
    
    # Count unique cases
    unique_cases = len(set(q["case_number"] for q in questions))
    
    # Create simulacro document
    simulacro_id = str(uuid.uuid4())
    simulacro = {
        "id": simulacro_id,
        "title": title,
        "total_questions": len(questions),
        "total_cases": unique_cases,
        "questions": questions,
        "created_at": datetime.now(timezone.utc),
        "created_by": admin_user.id
    }
    
    await db.simulacros.insert_one(simulacro)
    
    return {
        "message": f"Simulacro '{title}' importado: {len(questions)} preguntas, {unique_cases} casos",
        "id": simulacro_id,
        "total_questions": len(questions),
        "total_cases": unique_cases
    }

@api_router.get("/admin/simulacros")
async def get_admin_simulacros(admin_user: User = Depends(get_admin_user)):
    """List all simulacros (admin view)."""
    simulacros = await db.simulacros.find({}, {"_id": 0, "questions": 0}).sort("created_at", -1).to_list(100)
    return simulacros

@api_router.delete("/admin/simulacros/{simulacro_id}")
async def delete_simulacro(simulacro_id: str, admin_user: User = Depends(get_admin_user)):
    result = await db.simulacros.delete_one({"id": simulacro_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Simulacro no encontrado")
    # Also delete all attempts
    await db.simulacro_attempts.delete_many({"simulacro_id": simulacro_id})
    return {"message": "Simulacro eliminado"}

@api_router.get("/simulacros")
async def get_simulacros(current_user: User = Depends(get_current_user)):
    """List available simulacros for users."""
    simulacros = await db.simulacros.find({}, {"_id": 0, "questions": 0}).sort("created_at", -1).to_list(100)
    
    # Get user's attempts for each simulacro
    for sim in simulacros:
        attempt = await db.simulacro_attempts.find_one(
            {"simulacro_id": sim["id"], "user_id": current_user.id},
            {"_id": 0, "status": 1, "score": 1, "total_questions": 1, "completed_at": 1}
        )
        sim["user_attempt"] = attempt
    
    return simulacros

@api_router.get("/simulacros/{simulacro_id}")
async def get_simulacro(simulacro_id: str, current_user: User = Depends(get_current_user)):
    """Get full simulacro with questions (for exam view)."""
    simulacro = await db.simulacros.find_one({"id": simulacro_id}, {"_id": 0})
    if not simulacro:
        raise HTTPException(status_code=404, detail="Simulacro no encontrado")
    
    # Strip correct answers from questions for the user
    safe_questions = []
    for q in simulacro.get("questions", []):
        safe_q = {k: v for k, v in q.items() if k != "correct_answer"}
        safe_questions.append(safe_q)
    simulacro["questions"] = safe_questions
    
    return simulacro

@api_router.post("/simulacros/{simulacro_id}/start")
async def start_simulacro(simulacro_id: str, current_user: User = Depends(get_current_user)):
    """Start a new attempt or resume existing one."""
    simulacro = await db.simulacros.find_one({"id": simulacro_id}, {"_id": 0})
    if not simulacro:
        raise HTTPException(status_code=404, detail="Simulacro no encontrado")
    
    # Check for existing in-progress attempt
    existing = await db.simulacro_attempts.find_one(
        {"simulacro_id": simulacro_id, "user_id": current_user.id, "status": "in_progress"},
        {"_id": 0}
    )
    
    if existing:
        return existing
    
    # Generate randomized case order
    import random
    questions = simulacro.get("questions", [])
    
    # Group question indices by case_number
    case_groups = {}
    for q in questions:
        cn = q["case_number"]
        if cn not in case_groups:
            case_groups[cn] = []
        case_groups[cn].append(q["index"])
    
    # Shuffle the case order, keep questions within each case in original order
    case_numbers = list(case_groups.keys())
    random.shuffle(case_numbers)
    
    question_order = []
    for cn in case_numbers:
        question_order.extend(case_groups[cn])
    
    # Create new attempt
    attempt_id = str(uuid.uuid4())
    attempt = {
        "id": attempt_id,
        "simulacro_id": simulacro_id,
        "user_id": current_user.id,
        "answers": {},
        "marked": [],
        "question_order": question_order,
        "status": "in_progress",
        "started_at": datetime.now(timezone.utc),
        "total_questions": simulacro["total_questions"],
        "time_limit_seconds": 5 * 3600  # 5 hours
    }
    
    await db.simulacro_attempts.insert_one(attempt)
    attempt.pop("_id", None)
    return attempt

@api_router.post("/simulacros/{simulacro_id}/save")
async def save_simulacro_progress(
    simulacro_id: str,
    body: dict,
    current_user: User = Depends(get_current_user)
):
    """Save answers and marked questions in progress."""
    attempt = await db.simulacro_attempts.find_one(
        {"simulacro_id": simulacro_id, "user_id": current_user.id, "status": "in_progress"}
    )
    if not attempt:
        raise HTTPException(status_code=404, detail="No hay intento activo")
    
    update = {}
    if "answers" in body:
        update["answers"] = body["answers"]
    if "marked" in body:
        update["marked"] = body["marked"]
    
    if update:
        await db.simulacro_attempts.update_one(
            {"id": attempt["id"]},
            {"$set": update}
        )
    
    return {"message": "Progreso guardado"}

@api_router.post("/simulacros/{simulacro_id}/finish")
async def finish_simulacro(
    simulacro_id: str,
    body: dict,
    current_user: User = Depends(get_current_user)
):
    """Finish the simulacro and calculate results."""
    simulacro = await db.simulacros.find_one({"id": simulacro_id}, {"_id": 0})
    if not simulacro:
        raise HTTPException(status_code=404, detail="Simulacro no encontrado")
    
    attempt = await db.simulacro_attempts.find_one(
        {"simulacro_id": simulacro_id, "user_id": current_user.id, "status": "in_progress"}
    )
    if not attempt:
        raise HTTPException(status_code=404, detail="No hay intento activo")
    
    # Use latest answers from body or from saved attempt
    answers = body.get("answers", attempt.get("answers", {}))
    
    # Calculate results
    total = len(simulacro["questions"])
    correct = 0
    results_by_especialidad = {}
    results_by_tema = {}
    
    for q in simulacro["questions"]:
        idx = str(q["index"])
        user_answer = answers.get(idx)
        is_correct = user_answer == q["correct_answer"]
        if is_correct:
            correct += 1
        
        esp = q["especialidad"]
        tema = q["tema"]
        
        # By especialidad
        if esp not in results_by_especialidad:
            results_by_especialidad[esp] = {"total": 0, "correct": 0}
        results_by_especialidad[esp]["total"] += 1
        if is_correct:
            results_by_especialidad[esp]["correct"] += 1
        
        # By tema (nested under especialidad)
        key = f"{esp}|{tema}"
        if key not in results_by_tema:
            results_by_tema[key] = {"especialidad": esp, "tema": tema, "total": 0, "correct": 0}
        results_by_tema[key]["total"] += 1
        if is_correct:
            results_by_tema[key]["correct"] += 1
    
    # Calculate percentages
    score_pct = round((correct / total) * 100, 1) if total > 0 else 0
    
    for esp_data in results_by_especialidad.values():
        esp_data["percentage"] = round((esp_data["correct"] / esp_data["total"]) * 100, 1) if esp_data["total"] > 0 else 0
    
    for tema_data in results_by_tema.values():
        tema_data["percentage"] = round((tema_data["correct"] / tema_data["total"]) * 100, 1) if tema_data["total"] > 0 else 0
    
    results = {
        "total_questions": total,
        "correct_answers": correct,
        "score_percentage": score_pct,
        "by_especialidad": results_by_especialidad,
        "by_tema": list(results_by_tema.values())
    }
    
    # Update attempt
    await db.simulacro_attempts.update_one(
        {"id": attempt["id"]},
        {"$set": {
            "status": "completed",
            "answers": answers,
            "score": correct,
            "score_percentage": score_pct,
            "total_questions": total,
            "results": results,
            "completed_at": datetime.now(timezone.utc)
        }}
    )
    
    # Award points only if score >= 60%
    if score_pct >= 60:
        points_result = await award_points(current_user.id, "simulacro", simulacro_id, POINT_VALUES["simulacro"])
        results["points_result"] = points_result
    else:
        results["points_result"] = {
            "message": f"Se requiere al menos 60% de aciertos para obtener puntos. Tu resultado: {score_pct}%",
            "points_awarded": 0
        }
    
    return results

@api_router.get("/simulacros/attempts/{attempt_id}")
async def get_simulacro_attempt(attempt_id: str, current_user: User = Depends(get_current_user)):
    """Get a completed attempt with results."""
    attempt = await db.simulacro_attempts.find_one(
        {"id": attempt_id, "user_id": current_user.id},
        {"_id": 0}
    )
    if not attempt:
        raise HTTPException(status_code=404, detail="Intento no encontrado")
    return attempt

@api_router.get("/simulacros/{simulacro_id}/latest-result")
async def get_latest_simulacro_result(simulacro_id: str, current_user: User = Depends(get_current_user)):
    """Get the latest completed attempt result for a simulacro."""
    attempt = await db.simulacro_attempts.find_one(
        {"simulacro_id": simulacro_id, "user_id": current_user.id, "status": "completed"},
        {"_id": 0},
        sort=[("completed_at", -1)]
    )
    if not attempt:
        raise HTTPException(status_code=404, detail="No hay resultados completados")
    return attempt

# Duel Routes
@api_router.post("/duels/challenge")
async def create_duel_challenge(
    invite: DuelCreate,
    current_user: User = Depends(get_current_user)
):
    # Check if target user exists
    target_user = await db.users.find_one({"email": invite.player2_email, "is_approved": True})
    if not target_user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado o no aprobado")
    
    import random
    specialties = [
        "Ginecología y Obstetricia",
        "Cirugía", 
        "Pediatría",
        "Medicina Interna",
        "Otros"
    ]
    
    duel_topic = invite.duel_topic  # None = General
    round_specialties = []
    round_questions = []
    
    # Get ALL question IDs both players have ever seen (for long-term rotation)
    player_ids = [current_user.id, target_user["id"]]
    all_history = await db.duel_question_history.find(
        {"user_id": {"$in": player_ids}},
        {"question_id": 1, "_id": 0}
    ).to_list(5000)
    all_seen_ids = set(h["question_id"] for h in all_history)
    
    # Generate 5 rounds with variety across materias/subtopics
    selected_question_ids = set()
    used_materias = []  # Track materias used to ensure variety
    
    for i in range(5):
        if duel_topic:
            selected_specialty = duel_topic
        else:
            selected_specialty = random.choice(specialties)
        round_specialties.append(selected_specialty)
        
        # Step 1: Get all distinct materias for this specialty
        all_materias = await db.duel_questions.distinct("materia", {"specialty": selected_specialty})
        all_materias = [m for m in all_materias if m]  # Filter out None
        
        # Step 2: Prefer materias not yet used in this duel for maximum variety
        unused_materias = [m for m in all_materias if m not in used_materias]
        if not unused_materias:
            unused_materias = all_materias  # All used, allow repeats
        
        # Shuffle to randomize materia selection
        random.shuffle(unused_materias)
        
        selected_question = None
        
        # Step 3: Try each materia until we find a good question
        for materia in unused_materias:
            exclude_ids = list(selected_question_ids | all_seen_ids)
            
            # Priority A: Never seen by either player + from this materia
            candidates = await db.duel_questions.find(
                {"specialty": selected_specialty, "materia": materia, "id": {"$nin": exclude_ids}}
            ).to_list(5000)
            
            if candidates:
                selected_question = random.choice(candidates)
                used_materias.append(materia)
                break
        
        # Step 4: Fallback - any materia, never seen
        if not selected_question:
            exclude_ids = list(selected_question_ids | all_seen_ids)
            candidates = await db.duel_questions.find(
                {"specialty": selected_specialty, "id": {"$nin": exclude_ids}}
            ).to_list(5000)
            if candidates:
                selected_question = random.choice(candidates)
        
        # Step 5: All questions seen - reset cycle, pick least used
        if not selected_question:
            candidates = await db.duel_questions.find(
                {"specialty": selected_specialty, "id": {"$nin": list(selected_question_ids)}}
            ).sort("global_usage_count", 1).limit(50).to_list(50)
            if candidates:
                # Pick randomly from the 50 least-used to get variety
                selected_question = random.choice(candidates)
        
        # Step 6: Ultimate fallback - any duel question
        if not selected_question:
            candidates = await db.duel_questions.find(
                {"id": {"$nin": list(selected_question_ids)}}
            ).limit(50).to_list(50)
            if candidates:
                selected_question = random.choice(candidates)
        
        if selected_question:
            round_questions.append(selected_question["id"])
            selected_question_ids.add(selected_question["id"])
            
            # Increment global usage count
            await db.duel_questions.update_one(
                {"id": selected_question["id"]},
                {"$inc": {"global_usage_count": 1}}
            )
    
    if len(round_questions) < 5:
        raise HTTPException(status_code=400, detail="No hay suficientes preguntas de duelo cargadas. El administrador debe importar preguntas primero.")
    
    # Create duel with pre-generated rounds
    new_duel = Duel(
        player1_id=current_user.id,
        player2_id=target_user["id"],
        round_specialties=round_specialties,
        round_questions=round_questions,
        question_source="duel",
        duel_topic=duel_topic,
        status="waiting_player1",
        challenger_message=invite.challenger_message
    )
    
    await db.duels.insert_one(new_duel.dict())
    
    # Send instant notification to target user via WebSocket
    notification_message = {
        "type": "duel_challenge",
        "message": f"Has sido retado por {current_user.full_name}",
        "challenger": {
            "id": current_user.id,
            "name": current_user.full_name,
            "email": current_user.email,
            "profile_image": current_user.profile_image
        },
        "duel_id": new_duel.id,
        "challenger_message": invite.challenger_message,
        "timestamp": datetime.now(timezone.utc).isoformat()
    }
    
    # Try to send instant notification via SSE
    notification_sent = await sse_manager.send_message(target_user["id"], notification_message)
    
    return {
        "message": "¡Duelo creado! Es tu turno de jugar.", 
        "duel_id": new_duel.id,
        "instant_notification_sent": notification_sent,
        "target_user": target_user["full_name"]
    }

@api_router.get("/duels/pending")
async def get_pending_duels(current_user: User = Depends(get_current_user)):
    # Get duels where I'm player2 and it's pending for me to play
    # This includes:
    # 1. New challenges where player1 hasn't played yet (status: waiting_player1)
    # 2. Challenges where player1 completed and now it's my turn (status: waiting_player2, player1_completed: true)
    pending_duels = await db.duels.find({
        "player2_id": current_user.id,
        "$or": [
            {
                # New challenges where player1 hasn't played yet
                "status": "waiting_player1",
                "player1_completed": False,
                "player2_completed": False
            },
            {
                # Challenges ready for me where player1 already played
                "status": "waiting_player2", 
                "player1_completed": True,
                "player2_completed": False
            }
        ]
    }).sort("created_at", -1).to_list(500)  # Sort by newest first
    
    # Clean up and get challenger details
    clean_duels = []
    for duel in pending_duels:
        challenger = await db.users.find_one({"id": duel["player1_id"]})
        
        # Determine the status message
        if duel.get("player1_completed", False):
            status_message = "¡Ya completó sus 5 preguntas! Es tu turno."
        else:
            status_message = "Reto pendiente. El retador aún no ha jugado."
        
        clean_duel = {
            "id": duel["id"],
            "challenger_id": duel["player1_id"],
            "challenger_name": challenger["full_name"] if challenger else "Usuario desconocido",
            "challenger_image": challenger.get("profile_image") if challenger else None,
            "created_at": duel.get("created_at"),
            "player1_score": duel.get("player1_score", 0),
            "status_message": status_message,
            "ready_to_play": duel.get("player1_completed", False),
            "challenger_message": duel.get("challenger_message"),
            "duel_topic": duel.get("duel_topic")
        }
        clean_duels.append(clean_duel)
    
    return clean_duels

@api_router.get("/duels/pending/count")
async def get_pending_duels_count(current_user: User = Depends(get_current_user)):
    """Get count of pending duels for badge notification"""
    count = await db.duels.count_documents({
        "player2_id": current_user.id,
        "$or": [
            {
                "status": "waiting_player1",
                "player1_completed": False,
                "player2_completed": False
            },
            {
                "status": "waiting_player2", 
                "player1_completed": True,
                "player2_completed": False
            }
        ]
    })
    return {"count": count}

@api_router.post("/duels/respond-notification")
async def respond_to_duel_notification(
    response_data: dict,
    current_user: User = Depends(get_current_user)
):
    """Handle response to duel notification (Accept Now vs Accept Later)"""
    duel_id = response_data.get("duel_id")
    action = response_data.get("action")  # "accept_now" or "accept_later"
    
    if not duel_id:
        raise HTTPException(status_code=400, detail="Duel ID required")
    
    # Verify duel exists and user is player2
    duel = await db.duels.find_one({"id": duel_id})
    if not duel:
        raise HTTPException(status_code=404, detail="Duelo no encontrado")
    
    if duel["player2_id"] != current_user.id:
        raise HTTPException(status_code=403, detail="No tienes permiso para este duelo")
    
    if action == "accept_now":
        # Redirect user to duels page with duel ID to start game
        return {"message": "Accediendo al duelo", "redirect": f"/duels?startDuel={duel_id}", "duel_id": duel_id}
    elif action == "accept_later":
        # Just acknowledge - duel remains in pending
        return {"message": "Duelo guardado en pendientes"}
    else:
        raise HTTPException(status_code=400, detail="Acción inválida")

@api_router.post("/duels/start/{duel_id}")
async def start_duel_game(
    duel_id: str,
    current_user: User = Depends(get_current_user)
):
    duel = await db.duels.find_one({"id": duel_id})
    if not duel:
        raise HTTPException(status_code=404, detail="Duelo no encontrado")
    
    # Only player1 (challenger) or player2 (challenged) can start
    if current_user.id not in [duel["player1_id"], duel["player2_id"]]:
        raise HTTPException(status_code=403, detail="No tienes acceso a este duelo")
    
    return {"message": "Duelo iniciado", "duel_id": duel_id}

# Simple test route
@api_router.get("/duels/test")
async def test_duels():
    return {"message": "Duels API working", "status": "ok"}

@api_router.get("/duels/active")
async def get_active_duels(current_user: User = Depends(get_current_user)):
    # Get all duels where I'm either player1 or player2, excluding those that are just pending for me
    duels = await db.duels.find({
        "$or": [
            {"player1_id": current_user.id},
            {"player2_id": current_user.id}
        ],
        "$and": [
            {
                "$or": [
                    {"status": "waiting_player1"},  # I'm player1 and haven't started yet
                    {"status": "waiting_player2"},  # I'm player2 and player1 completed
                    {"status": "completed"}          # Duel is finished
                ]
            },
            {
                # Exclude duels that are only pending for me (where I'm player2 and player1 hasn't completed)
                "$or": [
                    {"player1_id": current_user.id},  # I'm player1 (always show)
                    {
                        "player2_id": current_user.id,
                        "player1_completed": True      # I'm player2 but player1 completed (show in active)
                    },
                    {"status": "completed"}            # Completed duels always show
                ]
            }
        ]
    }).to_list(500)
    
    # Clean up and get player details
    clean_duels = []
    for duel in duels:
        player1 = await db.users.find_one({"id": duel["player1_id"]})
        player2 = await db.users.find_one({"id": duel["player2_id"]})
        
        clean_duel = {
            "id": duel["id"],
            "player1_id": duel["player1_id"],
            "player2_id": duel["player2_id"],
            "player1_completed": duel.get("player1_completed", False),
            "player2_completed": duel.get("player2_completed", False),
            "player1_score": duel.get("player1_score", 0),
            "player2_score": duel.get("player2_score", 0),
            "status": duel.get("status", "waiting_player1"),
            "winner_id": duel.get("winner_id"),
            "forfeit": duel.get("forfeit", False),
            "created_at": duel.get("created_at"),
            "completed_at": duel.get("completed_at"),
            "player1_name": player1["full_name"] if player1 else "Usuario desconocido",
            "player1_image": player1.get("profile_image") if player1 else None,
            "player2_name": player2["full_name"] if player2 else "Usuario desconocido",
            "player2_image": player2.get("profile_image") if player2 else None,
            "current_turn": len(duel.get("player1_answers", [])) + len(duel.get("player2_answers", [])),
            "challenger_message": duel.get("challenger_message"),
            "winner_message": duel.get("winner_message")
        }
        clean_duels.append(clean_duel)
    
    return clean_duels

@api_router.post("/duels/{duel_id}/winner-message")
async def send_winner_message(
    duel_id: str,
    data: dict,
    current_user: User = Depends(get_current_user)
):
    """Send a message from the winner (player2) to the challenger"""
    duel = await db.duels.find_one({"id": duel_id})
    if not duel:
        raise HTTPException(status_code=404, detail="Duelo no encontrado")
    
    # Only the winner (player2) can send this message
    if duel.get("winner_id") != current_user.id:
        raise HTTPException(status_code=403, detail="Solo el ganador puede enviar este mensaje")
    
    # Only player2 can send winner message (when they beat player1)
    if current_user.id != duel["player2_id"]:
        raise HTTPException(status_code=403, detail="Solo el retado ganador puede enviar mensaje")
    
    message = data.get("message", "").strip()
    if not message:
        raise HTTPException(status_code=400, detail="El mensaje no puede estar vacío")
    
    await db.duels.update_one(
        {"id": duel_id},
        {"$set": {"winner_message": message}}
    )
    
    return {"message": "Mensaje enviado correctamente"}

@api_router.get("/duels/completed")  
async def get_completed_duels(
    current_user: User = Depends(get_current_user),
    page: int = 1,
    page_size: int = 20
):
    """Get only completed duels with results for both participants (paginated)"""
    # Validate pagination params
    page = max(1, page)
    page_size = min(max(1, page_size), 50)  # Max 50 per page
    skip = (page - 1) * page_size
    
    # Get total count for pagination metadata
    total_count = await db.duels.count_documents({
        "$or": [
            {"player1_id": current_user.id},
            {"player2_id": current_user.id}
        ],
        "status": "completed",
        "player1_completed": True,
        "player2_completed": True
    })
    
    # Get paginated completed duels with sorting in DB query
    completed_duels = await db.duels.find({
        "$or": [
            {"player1_id": current_user.id},
            {"player2_id": current_user.id}
        ],
        "status": "completed",
        "player1_completed": True,
        "player2_completed": True
    }).sort("completed_at", -1).skip(skip).limit(page_size).to_list(page_size)
    
    # Clean up and get player details with results
    clean_duels = []
    for duel in completed_duels:
        player1 = await db.users.find_one({"id": duel["player1_id"]})
        player2 = await db.users.find_one({"id": duel["player2_id"]})
        
        # Determine winner and result message
        player1_score = duel.get("player1_score", 0)
        player2_score = duel.get("player2_score", 0)
        winner_id = duel.get("winner_id")
        
        # Result message for current user
        if winner_id == current_user.id:
            result_message = "¡Ganaste!"
            result_class = "winner"
        elif winner_id:
            result_message = "Perdiste"
            result_class = "loser"
        else:
            result_message = "Empate"
            result_class = "tie"
        
        clean_duel = {
            "id": duel["id"],
            "player1_id": duel["player1_id"],
            "player2_id": duel["player2_id"],
            "player1_score": player1_score,
            "player2_score": player2_score,
            "winner_id": winner_id,
            "result_message": result_message,
            "result_class": result_class,
            "player1_name": player1["full_name"] if player1 else "Usuario desconocido",
            "player1_image": player1.get("profile_image") if player1 else None,
            "player2_name": player2["full_name"] if player2 else "Usuario desconocido", 
            "player2_image": player2.get("profile_image") if player2 else None,
            "completed_at": duel.get("completed_at"),
            "created_at": duel.get("created_at"),
            "is_current_user_player1": duel["player1_id"] == current_user.id,
            "challenger_message": duel.get("challenger_message"),
            "winner_message": duel.get("winner_message")
        }
        clean_duels.append(clean_duel)
    
    return {
        "duels": clean_duels,
        "pagination": {
            "page": page,
            "page_size": page_size,
            "total_count": total_count,
            "total_pages": (total_count + page_size - 1) // page_size,
            "has_next": skip + page_size < total_count,
            "has_prev": page > 1
        }
    }

@api_router.get("/duels/{duel_id}")
async def get_duel_details(
    duel_id: str,
    current_user: User = Depends(get_current_user)
):
    duel = await db.duels.find_one({"id": duel_id})
    if not duel:
        raise HTTPException(status_code=404, detail="Duelo no encontrado")
    
    # Check if user is part of this duel
    if current_user.id not in [duel["player1_id"], duel["player2_id"]]:
        raise HTTPException(status_code=403, detail="No tienes acceso a este duelo")
    
    # Get player details
    player1 = await db.users.find_one({"id": duel["player1_id"]})
    player2 = await db.users.find_one({"id": duel["player2_id"]})
    
    # Remove MongoDB ObjectId and create clean response
    duel_data = {
        "id": duel["id"],
        "player1_id": duel["player1_id"],
        "player2_id": duel["player2_id"],
        "round_specialties": duel.get("round_specialties", []),
        "round_questions": duel.get("round_questions", []),
        "question_source": duel.get("question_source", "gpc"),
        "duel_topic": duel.get("duel_topic"),
        "player1_answers": duel.get("player1_answers", []),
        "player2_answers": duel.get("player2_answers", []),
        "player1_completed": duel.get("player1_completed", False),
        "player2_completed": duel.get("player2_completed", False),
        "player1_score": duel.get("player1_score", 0),
        "player2_score": duel.get("player2_score", 0),
        "status": duel.get("status", "waiting_player1"),
        "winner_id": duel.get("winner_id"),
        "forfeit": duel.get("forfeit", False),
        "created_at": duel.get("created_at"),
        "completed_at": duel.get("completed_at"),
        "player1_name": player1["full_name"] if player1 else "Usuario desconocido",
        "player1_image": player1.get("profile_image") if player1 else None,
        "player2_name": player2["full_name"] if player2 else "Usuario desconocido",
        "player2_image": player2.get("profile_image") if player2 else None
    }
    
    return duel_data

@api_router.post("/duels/submit-game")
async def submit_player_game(
    data: dict,
    current_user: User = Depends(get_current_user)
):
    duel_id = data.get("duel_id")
    answers = data.get("answers", [])  # List of 5 answers
    
    duel = await db.duels.find_one({"id": duel_id})
    if not duel:
        raise HTTPException(status_code=404, detail="Duelo no encontrado")
    
    is_player1 = current_user.id == duel["player1_id"]
    is_player2 = current_user.id == duel["player2_id"]
    
    if not (is_player1 or is_player2):
        raise HTTPException(status_code=403, detail="No tienes acceso a este duelo")
    
    # Calculate score - check question_source to use correct collection
    score = 0
    question_source = duel.get("question_source", "gpc")
    q_collection = db.duel_questions if question_source == "duel" else db.questions
    
    for i, answer in enumerate(answers):
        question_id = duel["round_questions"][i]
        question = await q_collection.find_one({"id": question_id})
        if question and question["correct_answer"] == answer:
            score += 1
    
    # Track question history for smart selection
    if question_source == "duel":
        history_entries = []
        for qid in duel["round_questions"]:
            history_entries.append({
                "user_id": current_user.id,
                "question_id": qid,
                "duel_id": duel_id,
                "answered_at": datetime.now(timezone.utc)
            })
        if history_entries:
            await db.duel_question_history.insert_many(history_entries)
    
    # Update duel data
    update_data = {}
    current_time = datetime.now(timezone.utc)
    duel_completed = False
    opponent_score = None
    
    if is_player1:
        if duel.get("player1_completed"):
            raise HTTPException(status_code=400, detail="Ya completaste este duelo")
        
        update_data = {
            "player1_answers": answers,
            "player1_completed": True,
            "player1_score": score,
            "player1_completed_at": current_time,
        }
        
        # Check if player2 has also completed
        if duel.get("player2_completed"):
            # Both players done - determine winner
            player2_score = duel["player2_score"]
            opponent_score = player2_score
            duel_completed = True
            update_data["status"] = "completed"
            update_data["completed_at"] = current_time
            
            if score > player2_score:
                update_data["winner_id"] = current_user.id
            elif score < player2_score:
                update_data["winner_id"] = duel["player2_id"]
            # If tied, winner_id remains None
        else:
            # Player1 finished first - wait for player2
            update_data["status"] = "waiting_player2"
    else:  # is_player2
        if duel.get("player2_completed"):
            raise HTTPException(status_code=400, detail="Ya completaste este duelo")
        
        # Player2 can play regardless of player1's status
        update_data = {
            "player2_answers": answers,
            "player2_completed": True,
            "player2_score": score,
            "player2_completed_at": current_time,
        }
        
        # Check if player1 has also completed
        if duel.get("player1_completed"):
            # Both players done - determine winner
            player1_score = duel["player1_score"]
            opponent_score = player1_score
            duel_completed = True
            update_data["status"] = "completed"
            update_data["completed_at"] = current_time
            
            if score > player1_score:
                update_data["winner_id"] = current_user.id
            elif score < player1_score:
                update_data["winner_id"] = duel["player1_id"]
            # If tied, winner_id remains None
        else:
            # Player2 finished first - wait for player1
            update_data["status"] = "waiting_player1"
    
    await db.duels.update_one({"id": duel_id}, {"$set": update_data})
    
    # Award points for duel participation
    points_result = None
    if duel_completed:
        winner_id = update_data.get("winner_id")
        if winner_id == current_user.id:
            points_result = await award_points(current_user.id, "duel_win", duel_id, POINT_VALUES["duel_win"])
        elif winner_id is not None:
            points_result = await award_points(current_user.id, "duel_loss", duel_id, POINT_VALUES["duel_loss"])
        else:
            # Tie - both get loss points
            points_result = await award_points(current_user.id, "duel_loss", duel_id, POINT_VALUES["duel_loss"])
        
        # Also award points to opponent
        opponent_id = duel["player2_id"] if is_player1 else duel["player1_id"]
        if winner_id == opponent_id:
            await award_points(opponent_id, "duel_win", duel_id, POINT_VALUES["duel_win"])
        elif winner_id is not None:
            await award_points(opponent_id, "duel_loss", duel_id, POINT_VALUES["duel_loss"])
        else:
            await award_points(opponent_id, "duel_loss", duel_id, POINT_VALUES["duel_loss"])
    
    return {
        "message": "Juego completado", 
        "score": score, 
        "my_score": score,
        "duel_completed": duel_completed,
        "opponent_score": opponent_score,
        "points_result": points_result
    }

@api_router.post("/duels/abandon/{duel_id}")
async def abandon_duel(
    duel_id: str,
    current_user: User = Depends(get_current_user)
):
    duel = await db.duels.find_one({"id": duel_id})
    if not duel:
        raise HTTPException(status_code=404, detail="Duelo no encontrado")
    
    # Check if user is part of this duel
    if current_user.id not in [duel["player1_id"], duel["player2_id"]]:
        raise HTTPException(status_code=403, detail="No tienes acceso a este duelo")
    
    # Check if duel is already completed
    if duel.get("status") == "completed":
        raise HTTPException(status_code=400, detail="No puedes abandonar un duelo completado")
    
    is_player1 = current_user.id == duel["player1_id"]
    is_player2 = current_user.id == duel["player2_id"]
    
    # If player1 abandons and hasn't completed, delete the duel
    if is_player1 and not duel.get("player1_completed"):
        await db.duels.delete_one({"id": duel_id})
        return {"message": "Duelo cancelado exitosamente"}
    
    # If player2 abandons before completing, mark as forfeit
    if is_player2 and not duel.get("player2_completed"):
        update_data = {
            "player2_completed": True,
            "player2_score": 0,  # Forfeit score
            "player2_completed_at": datetime.now(timezone.utc),
            "status": "completed",
            "completed_at": datetime.now(timezone.utc),
            "winner_id": duel["player1_id"],  # Player1 wins by forfeit
            "forfeit": True
        }
        await db.duels.update_one({"id": duel_id}, {"$set": update_data})
        return {"message": "Has abandonado el duelo. Tu oponente gana por abandono."}
    
    raise HTTPException(status_code=400, detail="No puedes abandonar este duelo en su estado actual")

@api_router.post("/duels/reject/{duel_id}")
async def reject_duel(
    duel_id: str,
    current_user: User = Depends(get_current_user)
):
    duel = await db.duels.find_one({"id": duel_id})
    if not duel:
        raise HTTPException(status_code=404, detail="Duelo no encontrado")
    
    # Only player2 can reject a duel, and only if they haven't started playing
    if current_user.id != duel["player2_id"]:
        raise HTTPException(status_code=403, detail="Solo el usuario retado puede rechazar el duelo")
    
    if duel.get("player2_completed") or duel.get("status") == "completed":
        raise HTTPException(status_code=400, detail="No puedes rechazar un duelo que ya comenzaste o completaste")
    
    # Delete the duel
    await db.duels.delete_one({"id": duel_id})
    return {"message": "Duelo rechazado exitosamente"}

# Specialty Routes
@api_router.get("/specialties")
async def get_specialties(current_user: User = Depends(get_current_user)):
    return [
        "Ginecología y Obstetricia",
        "Cirugía",
        "Pediatría",
        "Medicina Interna",
        "Otros"
    ]

@api_router.get("/users/active")
async def get_active_users(
    current_user: User = Depends(get_current_user),
    limit: int = 20
):
    """Get active/approved users excluding current user (with limit for scalability)"""
    users = await db.users.find({
        "is_approved": True,
        "id": {"$ne": current_user.id}
    }).limit(min(limit, 50)).to_list(min(limit, 50))
    
    return [
        {
            "id": user["id"],
            "full_name": user["full_name"],
            "email": user["email"],
            "profile_image": user.get("profile_image")
        }
        for user in users
    ]

@api_router.get("/users/search")
async def search_users(
    q: str = "",
    limit: int = 20,
    current_user: User = Depends(get_current_user)
):
    """Search approved users by name or email"""
    if not q or len(q) < 2:
        # If no query, return first N users
        users = await db.users.find({
            "is_approved": True,
            "id": {"$ne": current_user.id}
        }).limit(min(limit, 50)).to_list(min(limit, 50))
    else:
        # Search by name or email
        users = await db.users.find({
            "is_approved": True,
            "id": {"$ne": current_user.id},
            "$or": [
                {"full_name": {"$regex": q, "$options": "i"}},
                {"email": {"$regex": q, "$options": "i"}}
            ]
        }).limit(min(limit, 50)).to_list(min(limit, 50))
    
    return [
        {
            "id": user["id"],
            "full_name": user["full_name"],
            "email": user["email"],
            "profile_image": user.get("profile_image")
        }
        for user in users
    ]

# ==================== ESCAPE ROOM / CLINICAL CASES ENDPOINTS ====================

@api_router.get("/cases/themes")
async def get_case_themes():
    """Get available clinical case themes"""
    return {
        "themes": [
            "Cirugía",
            "Medicina Interna",
            "Pediatría",
            "Ginecología y Obstetricia",
            "Otros"
        ]
    }

@api_router.get("/cases")
async def get_cases(
    theme: Optional[str] = None, 
    module: Optional[str] = None,
    submodule: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get all clinical cases, optionally filtered by theme, module, and submodule"""
    query = {}
    if theme:
        query["theme"] = theme
    if module:
        query["module"] = module
    if submodule:
        query["submodule"] = submodule
    
    cases = await db.clinical_cases.find(query).to_list(1000)
    return [ClinicalCase(**case).dict() for case in cases]

@api_router.get("/cases/{case_id}")
async def get_case(case_id: str, current_user: User = Depends(get_current_user)):
    """Get a specific clinical case by ID"""
    case = await db.clinical_cases.find_one({"id": case_id})
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")
    return ClinicalCase(**case).dict()

@api_router.get("/user/escape-room-progress")
async def get_escape_room_progress(current_user: User = Depends(get_current_user)):
    """Get user's escape room progress and rank"""
    user_data = await db.users.find_one({"id": current_user.id})
    if not user_data:
        raise HTTPException(status_code=404, detail="User not found")
    
    successful_cases = user_data.get("cases_successful", 0)
    total_cases = user_data.get("cases_completed", 0)
    current_rank = user_data.get("current_rank", "Estudiante Universitario")
    
    rank_progress = get_rank_progress(successful_cases)
    
    return {
        "username": user_data.get("full_name", "Usuario"),
        "total_cases_completed": total_cases,
        "successful_cases": successful_cases,
        "current_rank": current_rank,
        "rank_progress": rank_progress,
        "success_rate": round((successful_cases / total_cases * 100) if total_cases > 0 else 0, 1)
    }

@api_router.post("/game/start")
async def start_game(case_id: str, current_user: User = Depends(get_current_user)):
    """Start a new game session for a clinical case"""
    case = await db.clinical_cases.find_one({"id": case_id})
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")
    
    # End any existing sessions for this user and case
    await db.game_sessions.update_many(
        {"user_id": current_user.id, "case_id": case_id, "is_completed": False},
        {"$set": {"is_completed": True}}
    )
    
    # Create new session
    session = GameSession(
        user_id=current_user.id,
        case_id=case_id
    )
    
    await db.game_sessions.insert_one(session.dict())
    
    return {
        "session_id": session.id,
        "case": ClinicalCase(**case).dict(),
        "current_step": 1,
        "steps_taken": 0
    }

@api_router.post("/game/answer")
async def submit_answer(answer: CaseAnswer, current_user: User = Depends(get_current_user)):
    """Submit an answer for a clinical case step"""
    # Get session - IMPORTANTE: Obtener el estado MÁS RECIENTE
    session = await db.game_sessions.find_one({"id": answer.session_id})
    if not session:
        raise HTTPException(status_code=404, detail="Game session not found")
    
    # Verificar si ya está completada (prevenir requests duplicadas)
    if session.get("is_completed", False):
        return {
            "is_correct": False,
            "message": "El caso ya ha sido completado",
            "game_ended": True,
            "is_successful": session.get("is_successful", False)
        }
    
    if session["user_id"] != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    if session["is_completed"]:
        raise HTTPException(status_code=400, detail="Game session already completed")
    
    case = await db.clinical_cases.find_one({"id": session["case_id"]})
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")
    
    # Get current step data
    current_step = session["current_step"]
    step_data = case["steps"][current_step - 1]  # 0-indexed
    
    # Check if answer is correct
    if answer.is_alternative:
        correct_answer = step_data.get("RespuestaCorrecta_Alternativa", "")
        is_correct = answer.selected_option == correct_answer
    else:
        correct_answer = step_data.get("RespuestaCorrecta_Principal", "")
        is_correct = answer.selected_option == correct_answer
    
    # Update session based on answer
    session_obj = GameSession(**session)
    session_obj.steps_taken += 1
    
    response_data = {
        "is_correct": is_correct,
        "message": "",
        "next_step": None,
        "show_alternative": False,
        "game_ended": False,
        "is_successful": False
    }
    
    # REGLA: >=9 pasos = muerte automática (7 principales + 1 alternativa permitida = 8 pasos máximo)
    if session_obj.steps_taken >= 9:
        session_obj.is_completed = True
        session_obj.final_message = case["global_messages"]["MensajeMuerte"]
        response_data["message"] = case["global_messages"]["MensajeMuerte"]
        response_data["game_ended"] = True
        
        # Update user progress for failed completion
        await db.users.update_one(
            {"id": current_user.id},
            {"$inc": {"cases_completed": 1}}
        )
        
        # Update session in database
        await db.game_sessions.update_one(
            {"id": answer.session_id},
            {"$set": session_obj.dict()}
        )
        
        return response_data
    
    if answer.is_alternative:
        if is_correct:
            # Correct alternative answer - return to main path
            response_data["message"] = step_data.get("MensajeCorrecto_Alternativa", "¡Correcto! Volviendo al camino principal.")
            session_obj.current_step += 1
            response_data["next_step"] = session_obj.current_step
        else:
            # REGLA NUEVA: Fallar alternativa incrementa el contador de errores
            session_obj.errors_count += 1
            print(f"DEBUG: Failed alternative - errors_count now = {session_obj.errors_count}")
            
            # REGLA NUEVA: Solo muerte si ya tiene 2 o más errores totales
            if session_obj.errors_count >= 2:
                session_obj.is_completed = True
                session_obj.final_message = case["global_messages"]["MensajeMuerte"]
                response_data["message"] = case["global_messages"]["MensajeMuerte"]
                response_data["game_ended"] = True
                
                # Update user progress for failed completion
                await db.users.update_one(
                    {"id": current_user.id},
                    {"$inc": {"cases_completed": 1}}
                )
            else:
                # Primer error - mostrar mensaje pero permitir continuar
                response_data["message"] = step_data.get("MensajeIncorrecto_Alternativa", "Respuesta incorrecta en alternativa. Continuando al siguiente paso.")
                session_obj.current_step += 1
                response_data["next_step"] = session_obj.current_step
    else:
        # Main path answer
        if is_correct:
            # REGLA: Completar el último paso (7)
            if current_step >= 7:
                # DEBUG: Log para ver el estado
                print(f"DEBUG: Completing step 7 - errors_count={session_obj.errors_count}, steps_taken={session_obj.steps_taken}")
                
                # Verificar SOLO errores totales: 0-1 error = éxito, 2+ errores = fallo
                # La regla de pasos se verifica antes en línea 1660
                if session_obj.errors_count <= 1:
                    # Game completed successfully! (0 o 1 error)
                    session_obj.is_completed = True
                    session_obj.is_successful = True
                    session_obj.completed_at = datetime.now(timezone.utc)
                    session_obj.final_message = case["global_messages"]["MensajeCuracion"]
                    response_data["message"] = case["global_messages"]["MensajeCuracion"]
                    response_data["game_ended"] = True
                    response_data["is_successful"] = True
                else:
                    # Tiene 2+ errores → reprobado
                    session_obj.is_completed = True
                    session_obj.final_message = case["global_messages"]["MensajeMuerte"]
                    response_data["message"] = case["global_messages"]["MensajeMuerte"]
                    response_data["game_ended"] = True
                    response_data["is_successful"] = False
                    
                    # Update user progress for failed completion
                    await db.users.update_one(
                        {"id": current_user.id},
                        {"$inc": {"cases_completed": 1}}
                    )
                    
                    # Update session in database
                    await db.game_sessions.update_one(
                        {"id": answer.session_id},
                        {"$set": session_obj.dict()}
                    )
                    
                    return response_data
                
                # Update user progress
                user_update = await db.users.find_one_and_update(
                    {"id": current_user.id},
                    {
                        "$inc": {
                            "cases_completed": 1,
                            "cases_successful": 1
                        }
                    },
                    return_document=True
                )
                
                # Update rank
                new_successful_cases = user_update.get("cases_successful", 1)
                old_rank = user_update.get("current_rank", "Estudiante Universitario")
                new_rank = get_rank_from_successful_cases(new_successful_cases)
                
                if new_rank != old_rank:
                    await db.users.update_one(
                        {"id": current_user.id},
                        {"$set": {"current_rank": new_rank}}
                    )
                    response_data["rank_up"] = {
                        "old_rank": old_rank,
                        "new_rank": new_rank,
                        "successful_cases": new_successful_cases
                    }
            else:
                # Move to next step
                response_data["message"] = step_data.get("MensajeCorrecto_Principal", "¡Correcto! Continúa al siguiente paso.")
                session_obj.current_step += 1
                response_data["next_step"] = session_obj.current_step
        else:
            # REGLA NUEVA: Incrementar contador de errores totales
            session_obj.errors_count += 1
            session_obj.alternative_branches_used.append(current_step)
            print(f"DEBUG: Failed main question - errors_count now = {session_obj.errors_count}")
            
            # REGLA NUEVA: Solo muerte si ya tiene 2 o más errores totales
            if session_obj.errors_count >= 2:
                # Muerte por 2+ errores totales
                session_obj.is_completed = True
                session_obj.final_message = case["global_messages"]["MensajeMuerte"]
                response_data["message"] = case["global_messages"]["MensajeMuerte"]
                response_data["game_ended"] = True
                
                # Update user progress for failed completion
                await db.users.update_one(
                    {"id": current_user.id},
                    {"$inc": {"cases_completed": 1}}
                )
            else:
                # Primer error - mostrar alternativa
                response_data["message"] = step_data.get("MensajeIncorrecto_Principal", "Respuesta incorrecta. Intenta la pregunta alternativa.")
                response_data["show_alternative"] = True
    
    # Update session in database
    await db.game_sessions.update_one(
        {"id": answer.session_id},
        {"$set": session_obj.dict()}
    )
    
    return response_data

@api_router.post("/admin/cases/upload")
async def upload_case(
    title: str = Form(...),
    theme: str = Form(...),
    description: str = Form(...),
    module: str = Form(...),
    submodule: str = Form(...),
    file: UploadFile = File(...),
    current_user: User = Depends(get_admin_user)
):
    """Upload a clinical case from Excel or CSV file (admin only)"""
    try:
        contents = await file.read()
        
        # Detect file type and read accordingly
        filename = file.filename.lower()
        if filename.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(contents))
        elif filename.endswith(('.xlsx', '.xls')):
            df = pd.read_excel(io.BytesIO(contents), engine='openpyxl')
        else:
            raise HTTPException(status_code=400, detail="File must be CSV or Excel (.xlsx, .xls)")
        
        # Validate required columns
        required_columns = [
            "Paso", "Pregunta_Principal", "A_Principal", "B_Principal", 
            "C_Principal", "D_Principal", "RespuestaCorrecta_Principal"
        ]
        
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise HTTPException(
                status_code=400, 
                detail=f"Missing required columns: {', '.join(missing_columns)}"
            )
        
        # Parse steps
        steps = []
        for _, row in df.iterrows():
            step = {
                "Paso": int(row["Paso"]),
                "Pregunta_Principal": str(row["Pregunta_Principal"]),
                "A_Principal": str(row["A_Principal"]),
                "B_Principal": str(row["B_Principal"]),
                "C_Principal": str(row["C_Principal"]),
                "D_Principal": str(row["D_Principal"]),
                "RespuestaCorrecta_Principal": str(row["RespuestaCorrecta_Principal"]),
                "MensajeCorrecto_Principal": str(row.get("MensajeCorrecto_Principal", "¡Correcto!")),
                "MensajeIncorrecto_Principal": str(row.get("MensajeIncorrecto_Principal", "Incorrecto")),
                "Pregunta_Alternativa": str(row.get("Pregunta_Alternativa", "")),
                "A_Alternativa": str(row.get("A_Alternativa", "")),
                "B_Alternativa": str(row.get("B_Alternativa", "")),
                "C_Alternativa": str(row.get("C_Alternativa", "")),
                "D_Alternativa": str(row.get("D_Alternativa", "")),
                "RespuestaCorrecta_Alternativa": str(row.get("RespuestaCorrecta_Alternativa", "")),
                "MensajeCorrecto_Alternativa": str(row.get("MensajeCorrecto_Alternativa", "¡Correcto!"))
            }
            steps.append(step)
        
        # Get global messages (from first row or default)
        global_messages = {
            "MensajeCuracion": str(df.iloc[0].get("MensajeCuracion", "¡Felicidades! Has completado el caso exitosamente.")),
            "MensajeMuerte": str(df.iloc[0].get("MensajeMuerte", "El paciente ha fallecido. Intenta nuevamente."))
        }
        
        # Create clinical case
        clinical_case = ClinicalCase(
            title=title,
            theme=theme,
            description=description,
            module=module,
            submodule=submodule,
            steps=steps,
            global_messages=global_messages,
            created_by=current_user.id
        )
        
        await db.clinical_cases.insert_one(clinical_case.dict())
        
        return {"message": "Case uploaded successfully", "case_id": clinical_case.id}
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error processing file: {str(e)}")

@api_router.delete("/admin/cases/{case_id}")
async def delete_case(case_id: str, current_user: User = Depends(get_admin_user)):
    """Delete a clinical case (admin only)"""
    result = await db.clinical_cases.delete_one({"id": case_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Case not found")
    
    # Also delete any game sessions for this case
    await db.game_sessions.delete_many({"case_id": case_id})
    
    return {"message": "Case deleted successfully"}


# ==================== IMAGEN DX ENDPOINTS ====================

@api_router.get("/imagendx/systems")
async def get_imagendx_systems(current_user: User = Depends(get_current_user)):
    """Get all unique systems (subtemas) for Imagen DX"""
    systems = await db.imagen_dx_cases.distinct("system", {"is_published": True})
    return {"systems": sorted(systems)}

@api_router.get("/imagendx/cases")
async def get_imagendx_cases(
    system: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get all Imagen DX cases, optionally filtered by system"""
    query = {"is_published": True}
    if system:
        query["system"] = system
    
    cases = await db.imagen_dx_cases.find(query).to_list(30000)
    return [ImagenDXCase(**case).dict() for case in cases]

@api_router.get("/imagendx/cases/{case_id}")
async def get_imagendx_case(
    case_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get a specific Imagen DX case"""
    case = await db.imagen_dx_cases.find_one({"id": case_id})
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")
    return ImagenDXCase(**case).dict()

@api_router.post("/imagendx/submit")
async def submit_imagendx_answer(
    answer: ImagenDXAnswer,
    current_user: User = Depends(get_current_user)
):
    """Submit an answer to an Imagen DX case"""
    # Get case
    case = await db.imagen_dx_cases.find_one({"id": answer.case_id})
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")
    
    case_obj = ImagenDXCase(**case)
    
    # Map selected option to actual text
    option_map = {
        "A": case_obj.option_a,
        "B": case_obj.option_b,
        "C": case_obj.option_c,
        "D": case_obj.option_d
    }
    
    selected_text = option_map.get(answer.selected_option.upper())
    
    # Check if correct (compare with finding_or_sign)
    is_correct = selected_text == case_obj.finding_or_sign
    
    # Save/update progress if correct - increment pass_count
    points_result = None
    if is_correct:
        await db.user_progress.update_one(
            {"user_id": current_user.id, "item_id": answer.case_id, "item_type": "imagendx"},
            {
                "$set": {
                    "user_id": current_user.id,
                    "item_id": answer.case_id,
                    "item_type": "imagendx",
                    "passed": True,
                    "completed_at": datetime.now(timezone.utc).isoformat()
                },
                "$inc": {"pass_count": 1}
            },
            upsert=True
        )
        points_result = await award_points(current_user.id, "imagendx", answer.case_id, POINT_VALUES["imagendx"])
    
    return {
        "is_correct": is_correct,
        "correct_answer": case_obj.finding_or_sign,
        "feedback": {
            "modality": case_obj.modality,
            "interpretation_clave": case_obj.interpretation_clave,
            "diagnostico_sugerido": case_obj.diagnostico_sugerido
        },
        "points_result": points_result
    }

# Progress Tracking Endpoints
@api_router.get("/progress")
async def get_user_progress(current_user: User = Depends(get_current_user)):
    """Get all progress for the current user (passed quizzes, escape rooms, imagendx)"""
    
    # Get quiz attempts with score >= 80%
    quiz_attempts = await db.quiz_attempts.find({
        "user_id": current_user.id,
        "score": {"$gte": 80}
    }, {"_id": 0}).to_list(1000)
    
    passed_quizzes = list(set([attempt["quiz_id"] for attempt in quiz_attempts]))
    
    # Get successful escape room sessions
    game_sessions = await db.game_sessions.find({
        "user_id": current_user.id,
        "is_completed": True,
        "is_successful": True
    }, {"_id": 0}).to_list(1000)
    
    passed_escape_rooms = list(set([session["case_id"] for session in game_sessions]))
    
    # Get passed imagendx cases
    imagendx_progress = await db.user_progress.find({
        "user_id": current_user.id,
        "item_type": "imagendx",
        "passed": True
    }, {"_id": 0}).to_list(1000)
    
    passed_imagendx = [item["item_id"] for item in imagendx_progress]
    
    return {
        "quizzes": passed_quizzes,
        "escape_rooms": passed_escape_rooms,
        "imagendx": passed_imagendx
    }

@api_router.get("/progress/quizzes")
async def get_passed_quizzes(current_user: User = Depends(get_current_user)):
    """Get passed quiz IDs (score >= 80%)"""
    attempts = await db.quiz_attempts.find({
        "user_id": current_user.id,
        "score": {"$gte": 80}
    }, {"_id": 0}).to_list(1000)
    
    passed_quiz_ids = list(set([attempt["quiz_id"] for attempt in attempts]))
    return {"passed": passed_quiz_ids}

@api_router.get("/progress/escape-rooms")
async def get_passed_escape_rooms(current_user: User = Depends(get_current_user)):
    """Get passed escape room case IDs (completed successfully)"""
    sessions = await db.game_sessions.find({
        "user_id": current_user.id,
        "is_completed": True,
        "is_successful": True
    }, {"_id": 0}).to_list(1000)
    
    passed_case_ids = list(set([session["case_id"] for session in sessions]))
    return {"passed": passed_case_ids}

@api_router.post("/progress/escape-room-complete")
async def record_escape_room_completion(
    data: dict,
    current_user: User = Depends(get_current_user)
):
    """Record a successful escape room completion from the retro UI"""
    case_id = data.get("case_id")
    if not case_id:
        raise HTTPException(status_code=400, detail="case_id is required")
    
    # Verify the case exists
    case = await db.clinical_cases.find_one({"id": case_id})
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")
    
    # Create a new game session marked as completed and successful
    session = GameSession(
        user_id=current_user.id,
        case_id=case_id,
        is_completed=True,
        is_successful=True,
        completed_at=datetime.now(timezone.utc)
    )
    
    await db.game_sessions.insert_one(session.dict())
    
    # Update user_progress for pass_count tracking
    await db.user_progress.update_one(
        {"user_id": current_user.id, "item_id": case_id, "item_type": "escape_room"},
        {
            "$set": {
                "user_id": current_user.id,
                "item_id": case_id,
                "item_type": "escape_room",
                "passed": True,
                "completed_at": datetime.now(timezone.utc).isoformat()
            },
            "$inc": {"pass_count": 1}
        },
        upsert=True
    )
    
    # Update user stats
    await db.users.update_one(
        {"id": current_user.id},
        {
            "$inc": {
                "cases_completed": 1,
                "cases_successful": 1
            }
        }
    )
    
    # Check for rank up
    user = await db.users.find_one({"id": current_user.id})
    successful_cases = user.get("cases_successful", 0)
    old_rank = user.get("current_rank", "Estudiante Universitario")
    new_rank = get_rank_from_successful_cases(successful_cases)
    
    rank_up = None
    if new_rank != old_rank:
        await db.users.update_one(
            {"id": current_user.id},
            {"$set": {"current_rank": new_rank}}
        )
        rank_up = {
            "old_rank": old_rank,
            "new_rank": new_rank,
            "successful_cases": successful_cases
        }
    
    return {
        "success": True,
        "session_id": session.id,
        "rank_up": rank_up,
        "points_result": await award_points(current_user.id, "escape_room", case_id, POINT_VALUES["escape_room"])
    }

@api_router.get("/progress/pass-counts")
async def get_pass_counts(current_user: User = Depends(get_current_user)):
    """Get pass counts for all items (quizzes, escape rooms, imagendx)"""
    progress = await db.user_progress.find({
        "user_id": current_user.id,
        "passed": True
    }, {"_id": 0, "item_id": 1, "item_type": 1, "pass_count": 1}).to_list(1000)
    
    # Convert to dict format: { item_id: pass_count }
    result = {}
    for item in progress:
        result[item["item_id"]] = item.get("pass_count", 1)
    
    return {"pass_counts": result}

@api_router.get("/progress/imagendx")
async def get_imagendx_progress(current_user: User = Depends(get_current_user)):
    """Get passed imagendx case IDs"""
    progress = await db.user_progress.find({
        "user_id": current_user.id,
        "item_type": "imagendx",
        "passed": True
    }, {"_id": 0}).to_list(1000)
    
    passed_ids = [item["item_id"] for item in progress]
    return {"passed": passed_ids}

# Admin endpoints for Imagen DX
@api_router.post("/admin/imagendx/upload")
async def upload_imagendx_cases(
    system: str = Form(...),
    file: UploadFile = File(...),
    replace_existing: bool = Form(False),
    current_user: User = Depends(get_admin_user)
):
    """Upload Imagen DX cases from CSV/Excel"""
    try:
        # Read file
        contents = await file.read()
        
        # Determine file type and read accordingly
        if file.filename.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(contents))
        elif file.filename.endswith(('.xlsx', '.xls')):
            df = pd.read_excel(io.BytesIO(contents))
        else:
            raise HTTPException(
                status_code=400,
                detail="Invalid file format. Only CSV and Excel files are allowed"
            )
        
        # Validate required columns (case-sensitive)
        required_columns = [
            'system', 'modality', 'finding_or_sign', 'interpretation_clave',
            'diagnostico_sugerido', 'CASO CLINICO', 'A', 'B', 'C', 'D'
        ]
        
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise HTTPException(
                status_code=400,
                detail=f"Missing required columns: {', '.join(missing_columns)}"
            )
        
        # If replace_existing, delete all cases for this system
        if replace_existing:
            await db.imagen_dx_cases.delete_many({"system": system})
        
        # Process and insert cases
        imported_count = 0
        errors = []
        
        for index, row in df.iterrows():
            try:
                # Validate row data
                if pd.isna(row['finding_or_sign']) or not str(row['finding_or_sign']).strip():
                    errors.append(f"Row {index + 2}: finding_or_sign is empty")
                    continue
                
                # Create case object
                case = ImagenDXCase(
                    system=system,  # Use system from form, not from file
                    modality=str(row['modality']).strip(),
                    finding_or_sign=str(row['finding_or_sign']).strip(),
                    interpretation_clave=str(row['interpretation_clave']).strip(),
                    diagnostico_sugerido=str(row['diagnostico_sugerido']).strip(),
                    caso_clinico=str(row['CASO CLINICO']).strip(),
                    option_a=str(row['A']).strip(),
                    option_b=str(row['B']).strip(),
                    option_c=str(row['C']).strip(),
                    option_d=str(row['D']).strip(),
                    link_imagen_referencia=str(row.get('link_imagen_referencia', '')).strip() if not pd.isna(row.get('link_imagen_referencia')) else None,
                    created_by=current_user.id
                )
                
                # Insert into database
                await db.imagen_dx_cases.insert_one(case.dict())
                imported_count += 1
                
            except Exception as e:
                errors.append(f"Row {index + 2}: {str(e)}")
        
        return {
            "message": f"Successfully imported {imported_count} cases",
            "imported_count": imported_count,
            "errors": errors if errors else None
        }
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")

@api_router.post("/admin/imagendx/upload-image/{case_id}")
async def upload_imagendx_image(
    case_id: str,
    file: UploadFile = File(...),
    current_user: User = Depends(get_admin_user)
):
    """Upload an image for a specific Imagen DX case"""
    # Validate file type
    if not file.content_type.startswith('image/'):
        raise HTTPException(status_code=400, detail="File must be an image (JPG/PNG)")
    
    # Validate file size (5 MB limit)
    contents = await file.read()
    if len(contents) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Image size must be less than 5 MB")
    
    # Get case
    case = await db.imagen_dx_cases.find_one({"id": case_id})
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")
    
    # Save image to disk
    image_dir = "/app/frontend/public/imagendx"
    os.makedirs(image_dir, exist_ok=True)
    
    # Generate unique filename
    file_extension = file.filename.split('.')[-1]
    image_filename = f"{case_id}.{file_extension}"
    image_path = os.path.join(image_dir, image_filename)
    
    # Write file
    with open(image_path, 'wb') as f:
        f.write(contents)
    
    # Update case with image URL
    image_url = f"/imagendx/{image_filename}"
    await db.imagen_dx_cases.update_one(
        {"id": case_id},
        {"$set": {
            "image_url": image_url,
            "updated_at": datetime.now(timezone.utc),
            "updated_by": current_user.id
        }}
    )
    
    return {"message": "Image uploaded successfully", "image_url": image_url}

@api_router.post("/admin/imagendx/upload-multiple-images/{case_id}")
async def upload_imagendx_multiple_images(
    case_id: str,
    files: List[UploadFile] = File(...),
    current_user: User = Depends(get_admin_user)
):
    """Upload multiple images for Dermatología/Patología cases (up to 3)"""
    # Get case
    case = await db.imagen_dx_cases.find_one({"id": case_id})
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")
    
    # Validate system is Dermatología or Patología
    if case.get("system") not in ["Dermatología", "Patología"]:
        raise HTTPException(status_code=400, detail="Multiple images only allowed for Dermatología or Patología")
    
    # Validate number of files
    if len(files) > 3:
        raise HTTPException(status_code=400, detail="Maximum 3 images allowed")
    
    image_dir = "/app/frontend/public/imagendx"
    os.makedirs(image_dir, exist_ok=True)
    
    uploaded_urls = []
    
    for idx, file in enumerate(files):
        # Validate file type
        if not file.content_type.startswith('image/'):
            raise HTTPException(status_code=400, detail=f"File {idx+1} must be an image (JPG/PNG)")
        
        # Validate file size (5 MB limit)
        contents = await file.read()
        if len(contents) > 5 * 1024 * 1024:
            raise HTTPException(status_code=400, detail=f"Image {idx+1} size must be less than 5 MB")
        
        # Generate unique filename
        file_extension = file.filename.split('.')[-1] if '.' in file.filename else 'jpg'
        image_filename = f"{case_id}_{idx+1}.{file_extension}"
        image_path = os.path.join(image_dir, image_filename)
        
        # Write file
        with open(image_path, 'wb') as f:
            f.write(contents)
        
        uploaded_urls.append(f"/imagendx/{image_filename}")
    
    # Update case with image URLs
    update_data = {
        "updated_at": datetime.now(timezone.utc),
        "updated_by": current_user.id
    }
    
    # First image goes to image_url for backwards compatibility
    if uploaded_urls:
        update_data["image_url"] = uploaded_urls[0]
        update_data["image_urls"] = uploaded_urls
    
    await db.imagen_dx_cases.update_one(
        {"id": case_id},
        {"$set": update_data}
    )
    
    return {"message": f"{len(uploaded_urls)} images uploaded successfully", "image_urls": uploaded_urls}

@api_router.delete("/admin/imagendx/cases/{case_id}")
async def delete_imagendx_case(
    case_id: str,
    current_user: User = Depends(get_admin_user)
):
    """Delete an Imagen DX case"""
    result = await db.imagen_dx_cases.delete_one({"id": case_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Case not found")
    return {"message": "Case deleted successfully"}


# ==================== FLASHCARDS ENDPOINTS ====================

@api_router.post("/flashcards")
async def save_flashcard(
    flashcard_data: FlashcardCreate,
    current_user: User = Depends(get_current_user)
):
    """Save a quiz question as a flashcard"""
    # Check if this flashcard already exists for this user
    existing = await db.flashcards.find_one({
        "user_id": current_user.id,
        "question_id": flashcard_data.question_id
    })
    
    if existing:
        raise HTTPException(status_code=400, detail="Esta pregunta ya está guardada como flashcard")
    
    new_flashcard = Flashcard(
        user_id=current_user.id,
        question_id=flashcard_data.question_id,
        quiz_id=flashcard_data.quiz_id,
        quiz_title=flashcard_data.quiz_title,
        specialty=flashcard_data.specialty,
        topic=flashcard_data.topic,
        question_text=flashcard_data.question_text,
        answer_text=flashcard_data.answer_text,
        explanation=flashcard_data.explanation,
        is_manual=False
    )
    
    await db.flashcards.insert_one(new_flashcard.dict())
    return {"message": "Flashcard guardada exitosamente", "flashcard_id": new_flashcard.id}


@api_router.delete("/flashcards/deck")
async def delete_flashcard_deck(
    topic: str,
    quiz_title: str,
    current_user: User = Depends(get_current_user)
):
    """Delete all flashcards in a deck (by topic and quiz_title)"""
    result = await db.flashcards.delete_many({
        "user_id": current_user.id,
        "topic": topic,
        "quiz_title": quiz_title
    })
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="No se encontraron flashcards en este paquete")
    
    return {"message": f"Se eliminaron {result.deleted_count} flashcards", "deleted_count": result.deleted_count}


@api_router.delete("/flashcards/{flashcard_id}")
async def delete_flashcard(
    flashcard_id: str,
    current_user: User = Depends(get_current_user)
):
    """Delete a flashcard"""
    result = await db.flashcards.delete_one({
        "id": flashcard_id,
        "user_id": current_user.id
    })
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Flashcard no encontrada")
    
    return {"message": "Flashcard eliminada exitosamente"}


@api_router.delete("/flashcards/by-question/{question_id}")
async def delete_flashcard_by_question(
    question_id: str,
    current_user: User = Depends(get_current_user)
):
    """Delete a flashcard by question ID (used in quiz view toggle)"""
    result = await db.flashcards.delete_one({
        "question_id": question_id,
        "user_id": current_user.id
    })
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Flashcard no encontrada")
    
    return {"message": "Flashcard eliminada exitosamente"}


@api_router.get("/flashcards/check/{question_id}")
async def check_flashcard_exists(
    question_id: str,
    current_user: User = Depends(get_current_user)
):
    """Check if a question is already saved as a flashcard"""
    existing = await db.flashcards.find_one({
        "user_id": current_user.id,
        "question_id": question_id
    })
    return {"exists": existing is not None}


@api_router.get("/flashcards")
async def get_user_flashcards(current_user: User = Depends(get_current_user)):
    """Get all flashcards for the current user"""
    flashcards = await db.flashcards.find(
        {"user_id": current_user.id},
        {"_id": 0}
    ).sort("created_at", -1).to_list(5000)
    
    return flashcards


@api_router.get("/flashcards/grouped")
async def get_flashcards_grouped(current_user: User = Depends(get_current_user)):
    """Get flashcards grouped by specialty > topic > quiz_title"""
    flashcards = await db.flashcards.find(
        {"user_id": current_user.id},
        {"_id": 0}
    ).sort("created_at", -1).to_list(5000)
    
    # Helper function to normalize specialty to one of the 5 base categories
    def normalize_specialty(specialty: str) -> str:
        specialty_lower = (specialty or "").lower()
        if "ginecología" in specialty_lower or "ginecologia" in specialty_lower or "obstetricia" in specialty_lower:
            return "Ginecología y Obstetricia"
        elif "cirugía" in specialty_lower or "cirugia" in specialty_lower or "angiología" in specialty_lower or "angiologia" in specialty_lower:
            return "Cirugía"
        elif "pediatría" in specialty_lower or "pediatria" in specialty_lower:
            return "Pediatría"
        elif ("medicina interna" in specialty_lower or "cardiología" in specialty_lower or 
              "cardiologia" in specialty_lower or "endocrinología" in specialty_lower or
              "nefrología" in specialty_lower or "neumología" in specialty_lower or
              "gastroenterología" in specialty_lower or "neurología" in specialty_lower or
              "reumatología" in specialty_lower or "hematología" in specialty_lower or
              "infectología" in specialty_lower or "dermatología" in specialty_lower):
            return "Medicina Interna"
        else:
            return "Otros"
    
    # Group by normalized specialty -> topic -> quiz_title
    grouped = {}
    for fc in flashcards:
        raw_specialty = fc.get("specialty", "Otros")
        specialty = normalize_specialty(raw_specialty)
        topic = fc.get("topic", "General")
        quiz_title = fc.get("quiz_title", "Sin título")
        
        if specialty not in grouped:
            grouped[specialty] = {"topics": {}, "count": 0}
        
        if topic not in grouped[specialty]["topics"]:
            grouped[specialty]["topics"][topic] = {"quizzes": {}, "count": 0}
        
        if quiz_title not in grouped[specialty]["topics"][topic]["quizzes"]:
            grouped[specialty]["topics"][topic]["quizzes"][quiz_title] = {"cards": [], "count": 0}
        
        grouped[specialty]["topics"][topic]["quizzes"][quiz_title]["cards"].append(fc)
        grouped[specialty]["topics"][topic]["quizzes"][quiz_title]["count"] += 1
        grouped[specialty]["topics"][topic]["count"] += 1
        grouped[specialty]["count"] += 1
    
    return grouped


@api_router.get("/flashcards/by-deck")
async def get_flashcards_by_deck(
    specialty: str,
    topic: str,
    quiz_title: str,
    current_user: User = Depends(get_current_user)
):
    """Get flashcards for a specific deck (specialty + topic + quiz_title)"""
    # Helper function to normalize specialty
    def normalize_specialty(s: str) -> str:
        s_lower = (s or "").lower()
        if "ginecología" in s_lower or "ginecologia" in s_lower or "obstetricia" in s_lower:
            return "Ginecología y Obstetricia"
        elif "cirugía" in s_lower or "cirugia" in s_lower or "angiología" in s_lower or "angiologia" in s_lower:
            return "Cirugía"
        elif "pediatría" in s_lower or "pediatria" in s_lower:
            return "Pediatría"
        elif ("medicina interna" in s_lower or "cardiología" in s_lower or 
              "cardiologia" in s_lower or "endocrinología" in s_lower or
              "nefrología" in s_lower or "neumología" in s_lower or
              "gastroenterología" in s_lower or "neurología" in s_lower or
              "reumatología" in s_lower or "hematología" in s_lower or
              "infectología" in s_lower or "dermatología" in s_lower):
            return "Medicina Interna"
        else:
            return "Otros"
    
    # Get all flashcards for this user with matching topic and quiz_title
    flashcards = await db.flashcards.find(
        {
            "user_id": current_user.id,
            "topic": topic,
            "quiz_title": quiz_title
        },
        {"_id": 0}
    ).sort("created_at", 1).to_list(5000)
    
    # Filter by normalized specialty
    filtered = [fc for fc in flashcards if normalize_specialty(fc.get("specialty", "")) == specialty]
    
    return filtered


@api_router.post("/flashcards/manual")
async def create_manual_flashcard(
    flashcard_data: FlashcardManualCreate,
    current_user: User = Depends(get_current_user)
):
    """Create a manual flashcard"""
    new_flashcard = Flashcard(
        user_id=current_user.id,
        question_id=None,
        quiz_id=None,
        quiz_title=flashcard_data.quiz_title,
        specialty=flashcard_data.specialty,
        topic=flashcard_data.topic,
        question_text=flashcard_data.question_text,
        answer_text=flashcard_data.answer_text,
        personal_notes=flashcard_data.personal_notes,
        is_manual=True
    )
    
    await db.flashcards.insert_one(new_flashcard.dict())
    return {"message": "Flashcard creada exitosamente", "flashcard_id": new_flashcard.id}


@api_router.put("/flashcards/{flashcard_id}")
async def update_flashcard(
    flashcard_id: str,
    update_data: FlashcardUpdate,
    current_user: User = Depends(get_current_user)
):
    """Update a manual flashcard (only manual cards can be edited)"""
    flashcard = await db.flashcards.find_one({
        "id": flashcard_id,
        "user_id": current_user.id
    })
    
    if not flashcard:
        raise HTTPException(status_code=404, detail="Flashcard no encontrada")
    
    if not flashcard.get("is_manual"):
        raise HTTPException(status_code=400, detail="Solo las flashcards manuales pueden ser editadas")
    
    update_dict = {k: v for k, v in update_data.dict().items() if v is not None}
    
    if update_dict:
        await db.flashcards.update_one(
            {"id": flashcard_id},
            {"$set": update_dict}
        )
    
    return {"message": "Flashcard actualizada exitosamente"}


@api_router.get("/flashcards/specialties")
async def get_flashcard_specialties():
    """Get available specialties for flashcards"""
    return [
        "Ginecología y Obstetricia",
        "Cirugía",
        "Pediatría",
        "Medicina Interna",
        "Otros"
    ]


@api_router.get("/flashcards/topics/{specialty}")
async def get_topics_for_specialty(specialty: str, current_user: User = Depends(get_current_user)):
    """Get available topics for a specialty based on existing questions"""
    # Get unique topics from questions
    questions = await db.questions.find(
        {"specialty": specialty},
        {"topic": 1, "_id": 0}
    ).to_list(30000)
    
    topics = list(set(q.get("topic") for q in questions if q.get("topic")))
    topics.sort()
    
    return topics


@api_router.get("/flashcards/decks/{specialty}/{topic}")
async def get_decks_for_topic(
    specialty: str,
    topic: str,
    current_user: User = Depends(get_current_user)
):
    """Get available deck names for a specialty/topic"""
    flashcards = await db.flashcards.find(
        {
            "user_id": current_user.id,
            "specialty": specialty,
            "topic": topic
        },
        {"quiz_title": 1, "_id": 0}
    ).to_list(5000)
    
    decks = list(set(fc.get("quiz_title") for fc in flashcards if fc.get("quiz_title")))
    decks.sort()
    
    return decks


# Initialize admin user on startup
@app.on_event("startup")
async def startup_event():
    # Check if admin user already exists
    existing_admin = await db.users.find_one({"email": "admin@puertoenarm.com"})
    
    if not existing_admin:
        # Create new admin user with profile image only if doesn't exist
        admin_password = get_password_hash("admin123")
        # Simple avatar image (green circle with white "A")
        admin_image = "data:image/svg+xml;base64,PHN2ZyB3aWR0aD0iMTAwIiBoZWlnaHQ9IjEwMCIgdmlld0JveD0iMCAwIDEwMCAxMDAiIGZpbGw9Im5vbmUiIHhtbG5zPSJodHRwOi8vd3d3LnczLm9yZy8yMDAwL3N2ZyI+CjxjaXJjbGUgY3g9IjUwIiBjeT0iNTAiIHI9IjUwIiBmaWxsPSIjMGQ5NDg4Ii8+Cjx0ZXh0IHg9IjUwIiB5PSI2NSIgZm9udC1mYW1pbHk9IkFyaWFsLCBzYW5zLXNlcmlmIiBmb250LXNpemU9IjQwIiBmaWxsPSJ3aGl0ZSIgdGV4dC1hbmNob3I9Im1pZGRsZSI+QTwvdGV4dD4KPC9zdmc+"
    
        new_admin = User(
            full_name="Administrador Puerto ENARM",
            email="admin@puertoenarm.com",
            hashed_password=admin_password,
            is_admin=True,
            is_approved=True,
            profile_image=admin_image
        )
        await db.users.insert_one(new_admin.dict())
        print("Admin user created with profile image: admin@puertoenarm.com / admin123")
    else:
        print("Admin user already exists, skipping creation")
    
    # Initialize object storage for presentations
    try:
        presentation_storage.init_storage()
    except Exception as e:
        print(f"⚠️ Object storage init: {e}")
    
    # =============================================================================
    # DATABASE INDEXES - Critical for performance with large datasets
    # =============================================================================
    print("🔧 Creating/verifying database indexes...")
    
    try:
        # USERS collection indexes
        await db.users.create_index("id", unique=True)
        await db.users.create_index("email", unique=True)
        await db.users.create_index("is_approved")
        await db.users.create_index("is_admin")
        print("  ✅ users indexes created")
    except Exception as e:
        print(f"  ⚠️ users indexes: {e}")
    
    try:
        # DUELS collection indexes - heavily queried
        await db.duels.create_index("id", unique=True)
        await db.duels.create_index([("player2_id", 1), ("status", 1)])  # Pending duels query
        await db.duels.create_index([("status", 1), ("created_at", -1)])  # Completed duels sorted
        await db.duels.create_index("player1_id")
        await db.duels.create_index("created_at")
        print("  ✅ duels indexes created")
    except Exception as e:
        print(f"  ⚠️ duels indexes: {e}")
    
    try:
        # QUESTIONS collection indexes
        await db.questions.create_index("id", unique=True)
        await db.questions.create_index("specialty")
        await db.questions.create_index([("specialty", 1), ("topic", 1)])
        print("  ✅ questions indexes created")
    except Exception as e:
        print(f"  ⚠️ questions indexes: {e}")
    
    try:
        # QUIZ_ATTEMPTS collection indexes
        await db.quiz_attempts.create_index("id", unique=True)
        await db.quiz_attempts.create_index("user_id")
        await db.quiz_attempts.create_index([("user_id", 1), ("quiz_id", 1)])
        await db.quiz_attempts.create_index("quiz_id")
        print("  ✅ quiz_attempts indexes created")
    except Exception as e:
        print(f"  ⚠️ quiz_attempts indexes: {e}")
    
    try:
        # USER_PROGRESS collection indexes - used for pass counts
        await db.user_progress.create_index([("user_id", 1), ("item_type", 1)])
        await db.user_progress.create_index([("user_id", 1), ("item_id", 1), ("item_type", 1)], unique=True)
        print("  ✅ user_progress indexes created")
    except Exception as e:
        print(f"  ⚠️ user_progress indexes: {e}")
    
    try:
        # FLASHCARDS collection indexes
        await db.flashcards.create_index("id", unique=True)
        await db.flashcards.create_index("user_id")
        await db.flashcards.create_index([("user_id", 1), ("specialty", 1)])
        await db.flashcards.create_index([("user_id", 1), ("question_id", 1)])
        print("  ✅ flashcards indexes created")
    except Exception as e:
        print(f"  ⚠️ flashcards indexes: {e}")
    
    try:
        # CLINICAL_CASES collection indexes
        await db.clinical_cases.create_index("id", unique=True)
        await db.clinical_cases.create_index("theme")
        await db.clinical_cases.create_index("is_published")
        print("  ✅ clinical_cases indexes created")
    except Exception as e:
        print(f"  ⚠️ clinical_cases indexes: {e}")
    
    try:
        # GAME_SESSIONS collection indexes
        await db.game_sessions.create_index("id", unique=True)
        await db.game_sessions.create_index("user_id")
        await db.game_sessions.create_index([("user_id", 1), ("case_id", 1)])
        print("  ✅ game_sessions indexes created")
    except Exception as e:
        print(f"  ⚠️ game_sessions indexes: {e}")
    
    try:
        # POINT_TRANSACTIONS collection indexes
        await db.point_transactions.create_index("id", unique=True)
        await db.point_transactions.create_index("user_id")
        await db.point_transactions.create_index([("user_id", 1), ("activity_type", 1), ("activity_id", 1)])
        await db.point_transactions.create_index([("user_id", 1), ("activity_type", 1), ("created_at", -1)])
        print("  ✅ point_transactions indexes created")
    except Exception as e:
        print(f"  ⚠️ point_transactions indexes: {e}")
    
    try:
        # IMAGEN_DX_CASES collection indexes (expanded from existing)
        await db.imagen_dx_cases.create_index("id", unique=True)
        await db.imagen_dx_cases.create_index([
            ("system", 1),
            ("is_published", 1),
            ("created_at", -1)
        ])
        await db.imagen_dx_cases.create_index([("finding_or_sign", "text")])
        print("  ✅ imagen_dx_cases indexes created")
    except Exception as e:
        print(f"  ⚠️ imagen_dx_cases indexes: {e}")
    
    try:
        # PEARL related collections indexes
        await db.pearl_modules.create_index("id", unique=True)
        await db.pearl_subtopics.create_index("id", unique=True)
        await db.pearl_subtopics.create_index("module_id")
        await db.pearls.create_index("id", unique=True)
        await db.pearls.create_index("subtopic_id")
        await db.pearls.create_index("is_active")
        await db.pearl_preferences.create_index([("user_id", 1)], unique=True)
        await db.pearl_delivery_logs.create_index([("user_id", 1), ("sent_at", -1)])
        await db.pearl_delivery_logs.create_index("pearl_id")
        await db.push_subscriptions.create_index("user_id")
        print("  ✅ pearl collections indexes created")
    except Exception as e:
        print(f"  ⚠️ pearl collections indexes: {e}")
    
    try:
        # PRESENTATIONS collection indexes
        await db.presentations.create_index("id", unique=True)
        await db.presentations.create_index("module")
        await db.presentations.create_index([("module", 1), ("submodule", 1)])
        await db.presentations.create_index("created_at")
        print("  ✅ presentations indexes created")
    except Exception as e:
        print(f"  ⚠️ presentations indexes: {e}")
    
    print("✅ Database indexing complete!")
    
    try:
        # DUEL_QUESTIONS collection indexes
        await db.duel_questions.create_index("id", unique=True)
        await db.duel_questions.create_index("specialty")
        await db.duel_questions.create_index([("specialty", 1), ("global_usage_count", 1)])
        # DUEL_QUESTION_HISTORY collection indexes
        await db.duel_question_history.create_index([("user_id", 1), ("answered_at", -1)])
        await db.duel_question_history.create_index("question_id")
        print("  ✅ duel_questions indexes created")
    except Exception as e:
        print(f"  ⚠️ duel_questions indexes: {e}")
    
    try:
        # SIMULACROS collection indexes
        await db.simulacros.create_index("id", unique=True)
        await db.simulacro_attempts.create_index("id", unique=True)
        await db.simulacro_attempts.create_index([("simulacro_id", 1), ("user_id", 1)])
        await db.simulacro_attempts.create_index("user_id")
        print("  ✅ simulacros indexes created")
    except Exception as e:
        print(f"  ⚠️ simulacros indexes: {e}")
    
    # =============================================================================
    
    # Create pilot clinical case if it doesn't exist
    pilot_exists = await db.clinical_cases.find_one({"title": "Glomerulonefritis lúpica clase III–IV rápidamente progresiva"})
    if not pilot_exists:
        pilot_case = ClinicalCase(
            title="Glomerulonefritis lúpica clase III–IV rápidamente progresiva",
            theme="Medicina Interna",
            description="Paciente femenina de 25 años con historia de lupus eritematoso sistémico presenta deterioro de función renal agudo.",
            steps=[
                {
                    "Paso": 1,
                    "Pregunta_Principal": "Paciente femenina de 25 años, con diagnóstico previo de lupus eritematoso sistémico, acude a urgencias por edema en miembros inferiores y disminución del volumen urinario en los últimos 5 días. ¿Cuál es el siguiente paso más apropiado?",
                    "A_Principal": "Iniciar diuréticos inmediatamente",
                    "B_Principal": "Solicitar exámenes de función renal y uroanálisis",
                    "C_Principal": "Referir a nefrología en 2 semanas",
                    "D_Principal": "Iniciar restricción de líquidos",
                    "RespuestaCorrecta_Principal": "B",
                    "MensajeCorrecto_Principal": "¡Correcto! La evaluación de función renal es prioritaria en un paciente con LES y oliguria.",
                    "MensajeIncorrecto_Principal": "Respuesta incorrecta. Primero debemos evaluar la función renal.",
                    "Pregunta_Alternativa": "Los resultados muestran creatinina elevada. ¿Qué estudio complementario es más urgente?",
                    "A_Alternativa": "Tomografía de abdomen",
                    "B_Alternativa": "Biopsia renal",
                    "C_Alternativa": "Ultrasonido renal",
                    "D_Alternativa": "Gammagrafía renal",
                    "RespuestaCorrecta_Alternativa": "C",
                    "MensajeCorrecto_Alternativa": "Correcto. El ultrasonido permite evaluar tamaño renal y descartar obstrucción."
                },
                {
                    "Paso": 2,
                    "Pregunta_Principal": "Los laboratorios muestran: Cr 3.5 mg/dL, BUN 60 mg/dL, proteinuria 3+. Complemento C3 bajo, C4 bajo. ¿Cuál es el diagnóstico más probable?",
                    "A_Principal": "Nefritis lúpica",
                    "B_Principal": "Insuficiencia renal aguda prerenal",
                    "C_Principal": "Glomerulonefritis post-estreptocócica",
                    "D_Principal": "Síndrome urémico hemolítico",
                    "RespuestaCorrecta_Principal": "A",
                    "MensajeCorrecto_Principal": "¡Excelente! El contexto de LES con hipocomplementemia y proteinuria sugiere nefritis lúpica.",
                    "MensajeIncorrecto_Principal": "Incorrecto. Considera el contexto de LES y los complementos bajos.",
                    "Pregunta_Alternativa": "¿Qué hallazgo serológico adicional apoyaría más el diagnóstico de nefritis lúpica?",
                    "A_Alternativa": "Anti-DNA positivo",
                    "B_Alternativa": "ANCA positivo",
                    "C_Alternativa": "Anti-MBG positivo",
                    "D_Alternativa": "ASLO elevado",
                    "RespuestaCorrecta_Alternativa": "A",
                    "MensajeCorrecto_Alternativa": "Correcto. Los anti-DNA son específicos de LES y se correlacionan con actividad renal."
                },
                {
                    "Paso": 3,
                    "Pregunta_Principal": "Se confirma nefritis lúpica. ¿Cuál es el siguiente paso diagnóstico esencial antes de iniciar tratamiento inmunosupresor?",
                    "A_Principal": "Resonancia magnética renal",
                    "B_Principal": "Biopsia renal",
                    "C_Principal": "PET-CT",
                    "D_Principal": "Iniciar tratamiento empírico",
                    "RespuestaCorrecta_Principal": "B",
                    "MensajeCorrecto_Principal": "¡Correcto! La biopsia es crucial para clasificar el tipo de nefritis lúpica y guiar el tratamiento.",
                    "MensajeIncorrecto_Principal": "Incorrecto. Necesitamos clasificar histológicamente antes de tratar.",
                    "Pregunta_Alternativa": "¿Qué parámetro de la biopsia es más importante para el pronóstico?",
                    "A_Alternativa": "Grado de inflamación",
                    "B_Alternativa": "Índice de actividad y cronicidad",
                    "C_Alternativa": "Tamaño del glomérulo",
                    "D_Alternativa": "Número de células mesangiales",
                    "RespuestaCorrecta_Alternativa": "B",
                    "MensajeCorrecto_Alternativa": "Correcto. Estos índices determinan pronóstico y agresividad del tratamiento."
                },
                {
                    "Paso": 4,
                    "Pregunta_Principal": "La biopsia reporta: Glomerulonefritis lúpica clase IV (difusa). ¿Cuál es el tratamiento de inducción más apropiado?",
                    "A_Principal": "Prednisona oral únicamente",
                    "B_Principal": "Micofenolato de mofetilo",
                    "C_Principal": "Ciclofosfamida + corticosteroides",
                    "D_Principal": "Azatioprina",
                    "RespuestaCorrecta_Principal": "C",
                    "MensajeCorrecto_Principal": "¡Excelente! La clase IV requiere terapia agresiva con ciclofosfamida y esteroides.",
                    "MensajeIncorrecto_Principal": "Incorrecto. La clase IV requiere inmunosupresión más agresiva.",
                    "Pregunta_Alternativa": "Si la paciente desea preservar fertilidad, ¿qué alternativa es válida?",
                    "A_Alternativa": "Esperar sin tratamiento",
                    "B_Alternativa": "Micofenolato como primera línea",
                    "C_Alternativa": "Rituximab",
                    "D_Alternativa": "Metotrexato",
                    "RespuestaCorrecta_Alternativa": "B",
                    "MensajeCorrecto_Alternativa": "Correcto. El micofenolato es efectivo y menos tóxico para la fertilidad que ciclofosfamida."
                },
                {
                    "Paso": 5,
                    "Pregunta_Principal": "La paciente inicia ciclofosfamida IV mensual + metilprednisolona. ¿Qué complicación debe monitorizarse más cercanamente?",
                    "A_Principal": "Hipertensión",
                    "B_Principal": "Cistitis hemorrágica",
                    "C_Principal": "Hipoglucemia",
                    "D_Principal": "Trombosis",
                    "RespuestaCorrecta_Principal": "B",
                    "MensajeCorrecto_Principal": "¡Correcto! La cistitis hemorrágica es una complicación grave de la ciclofosfamida.",
                    "MensajeIncorrecto_Principal": "Aunque importante, hay una complicación más específica de la ciclofosfamida.",
                    "Pregunta_Alternativa": "¿Qué medida preventiva reduce el riesgo de cistitis hemorrágica?",
                    "A_Alternativa": "Antibióticos profilácticos",
                    "B_Alternativa": "Mesna y buena hidratación",
                    "C_Alternativa": "Anticoagulación",
                    "D_Alternativa": "Restricción de líquidos",
                    "RespuestaCorrecta_Alternativa": "B",
                    "MensajeCorrecto_Alternativa": "Correcto. Mesna protege la vejiga del metabolito tóxico de ciclofosfamida."
                },
                {
                    "Paso": 6,
                    "Pregunta_Principal": "Después de 6 meses de tratamiento de inducción, la paciente muestra mejoría (Cr 1.2, proteinuria reducida). ¿Cuál es el siguiente paso?",
                    "A_Principal": "Suspender toda medicación",
                    "B_Principal": "Continuar ciclofosfamida indefinidamente",
                    "C_Principal": "Cambiar a terapia de mantenimiento (azatioprina o micofenolato)",
                    "D_Principal": "Aumentar dosis de ciclofosfamida",
                    "RespuestaCorrecta_Principal": "C",
                    "MensajeCorrecto_Principal": "¡Perfecto! Tras la inducción exitosa, se cambia a terapia de mantenimiento menos tóxica.",
                    "MensajeIncorrecto_Principal": "Incorrecto. El manejo debe pasar a una fase de mantenimiento.",
                    "Pregunta_Alternativa": "¿Por cuánto tiempo mínimo debe continuar la terapia de mantenimiento?",
                    "A_Alternativa": "6 meses",
                    "B_Alternativa": "1 año",
                    "C_Alternativa": "2-3 años mínimo",
                    "D_Alternativa": "De por vida",
                    "RespuestaCorrecta_Alternativa": "C",
                    "MensajeCorrecto_Alternativa": "Correcto. Se recomienda al menos 2-3 años para prevenir recaídas."
                },
                {
                    "Paso": 7,
                    "Pregunta_Principal": "Durante el seguimiento, la paciente pregunta sobre embarazo. ¿Cuál es la recomendación más apropiada?",
                    "A_Principal": "Puede embarazarse inmediatamente",
                    "B_Principal": "Debe esperar al menos 6-12 meses con enfermedad inactiva y suspender micofenolato",
                    "C_Principal": "No puede embarazarse nunca",
                    "D_Principal": "Solo con fertilización in vitro",
                    "RespuestaCorrecta_Principal": "B",
                    "MensajeCorrecto_Principal": "¡Excelente manejo! Es crucial que la enfermedad esté inactiva y cambiar a medicamentos seguros en embarazo.",
                    "MensajeIncorrecto_Principal": "Incorrecto. Hay consideraciones específicas sobre el embarazo en LES.",
                    "Pregunta_Alternativa": "¿Qué inmunosupresor es más seguro durante el embarazo en nefritis lúpica?",
                    "A_Alternativa": "Micofenolato",
                    "B_Alternativa": "Ciclofosfamida",
                    "C_Alternativa": "Azatioprina",
                    "D_Alternativa": "Metotrexato",
                    "RespuestaCorrecta_Alternativa": "C",
                    "MensajeCorrecto_Alternativa": "Correcto. La azatioprina es relativamente segura en embarazo, junto con hidroxicloroquina."
                }
            ],
            global_messages={
                "MensajeCuracion": "🎉 ¡Felicidades! Has manejado exitosamente un caso complejo de nefritis lúpica clase IV. Demostraste conocimiento en diagnóstico, clasificación, tratamiento de inducción y mantenimiento, así como manejo de complicaciones. La paciente ha mejorado significativamente y tiene buen pronóstico.",
                "MensajeMuerte": "❌ La paciente desarrolló complicaciones graves y progresó a enfermedad renal terminal. El manejo de la nefritis lúpica requiere diagnóstico temprano, tratamiento agresivo apropiado y monitorización cercana. Revisa las guías KDIGO y estudia más sobre glomerulonefritis lúpica."
            },
            created_by="system"
        )
        await db.clinical_cases.insert_one(pilot_case.dict())
        print("✅ Pilot clinical case created: Glomerulonefritis lúpica")

# ==================== PERLAS DIARIAS ENDPOINTS ====================

# VAPID keys for Web Push (generated once, store securely in production)
VAPID_PUBLIC_KEY = "BNvk4OjeutrI6dVHnkzEBgYsod3g1aLD_yLRwJX989p68r1-_6OQtFfRDXfXmbpJmWaeX-7Kb8WS76i7ftgt_oQ"
VAPID_PRIVATE_KEY = "DQQZcRDC6_HbUxA0oJ5xYQg9vyO8iqIduhW5WVcGRYo"
VAPID_CLAIMS = {"sub": "mailto:admin@puertoenarm.com"}

@api_router.get("/perlas/modules")
async def get_pearl_modules(current_user: User = Depends(get_current_user)):
    """Get all pearl modules (cached for 5 minutes)"""
    cache_key = "pearl_modules"
    cached = cache.get(cache_key, ttl_seconds=300)
    if cached:
        return cached
    
    modules = await db.pearl_modules.find({"is_active": True}, {"_id": 0}).to_list(100)
    cache.set(cache_key, modules)
    return modules

@api_router.get("/perlas/subtopics/{module_id}")
async def get_pearl_subtopics(module_id: str, current_user: User = Depends(get_current_user)):
    """Get subtopics for a module (cached for 5 minutes)"""
    cache_key = f"pearl_subtopics_{module_id}"
    cached = cache.get(cache_key, ttl_seconds=300)
    if cached:
        return cached
    
    subtopics = await db.pearl_subtopics.find(
        {"module_id": module_id, "is_active": True}, 
        {"_id": 0}
    ).to_list(500)
    cache.set(cache_key, subtopics)
    return subtopics

@api_router.get("/perlas/all-subtopics")
async def get_all_pearl_subtopics(current_user: User = Depends(get_current_user)):
    """Get all subtopics grouped by module (cached for 5 minutes)"""
    cache_key = "pearl_all_subtopics"
    cached = cache.get(cache_key, ttl_seconds=300)
    if cached:
        return cached
    
    modules = await db.pearl_modules.find({"is_active": True}, {"_id": 0}).to_list(100)
    result = []
    for module in modules:
        subtopics = await db.pearl_subtopics.find(
            {"module_id": module["id"], "is_active": True},
            {"_id": 0}
        ).to_list(500)
        result.append({
            "module": module,
            "subtopics": subtopics
        })
    cache.set(cache_key, result)
    return result

@api_router.get("/perlas/pearls/{subtopic_id}")
async def get_pearls_by_subtopic(subtopic_id: str, current_user: User = Depends(get_current_user)):
    """Get pearls for a subtopic"""
    pearls = await db.pearls.find(
        {"subtopic_id": subtopic_id, "is_active": True},
        {"_id": 0}
    ).to_list(1000)
    return pearls

@api_router.get("/perlas/vapid-public-key")
async def get_vapid_public_key():
    """Get VAPID public key for push subscription"""
    return {"publicKey": VAPID_PUBLIC_KEY}

@api_router.get("/perlas/user/preference")
async def get_user_pearl_preference(current_user: User = Depends(get_current_user)):
    """Get user's current pearl subscription preference"""
    preference = await db.pearl_preferences.find_one(
        {"user_id": current_user.id, "is_active": True},
        {"_id": 0}
    )
    if preference:
        # Get module and subtopic details
        module = await db.pearl_modules.find_one({"id": preference["module_id"]}, {"_id": 0})
        subtopic = await db.pearl_subtopics.find_one({"id": preference["subtopic_id"]}, {"_id": 0})
        preference["module"] = module
        preference["subtopic"] = subtopic
    return preference

@api_router.get("/perlas/user/delivery-history")
async def get_user_delivery_history(
    limit: int = 50,
    current_user: User = Depends(get_current_user)
):
    """Get user's pearl delivery history"""
    logs = await db.pearl_delivery_logs.find(
        {"user_id": current_user.id},
        {"_id": 0}
    ).sort("sent_at", -1).limit(limit).to_list(limit)
    
    # Get pearl details
    for log in logs:
        pearl = await db.pearls.find_one({"id": log["pearl_id"]}, {"_id": 0})
        log["pearl"] = pearl
    
    return logs

@api_router.get("/perlas/subscription-status")
async def get_subscription_status(current_user: User = Depends(get_current_user)):
    """Get user's push subscription status - useful for re-subscribing if expired"""
    # Check if user has active preferences
    preference = await db.pearl_preferences.find_one(
        {"user_id": current_user.id, "is_active": True},
        {"_id": 0}
    )
    
    # Check push subscriptions status
    active_subscriptions = await db.push_subscriptions.count_documents(
        {"user_id": current_user.id, "is_active": True}
    )
    
    inactive_subscriptions = await db.push_subscriptions.find(
        {"user_id": current_user.id, "is_active": False},
        {"_id": 0, "deactivated_at": 1, "deactivate_reason": 1, "error_count": 1}
    ).sort("deactivated_at", -1).to_list(5)
    
    # Count recent delivery failures
    today_start = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    failed_deliveries = await db.pearl_delivery_logs.count_documents({
        "user_id": current_user.id,
        "status": "failed",
        "sent_at": {"$gte": today_start.isoformat()}
    })
    
    return {
        "has_active_preference": preference is not None,
        "is_paused": preference.get("is_paused", False) if preference else False,
        "active_push_subscriptions": active_subscriptions,
        "needs_resubscribe": active_subscriptions == 0 and preference is not None,
        "recent_failures": failed_deliveries,
        "inactive_subscription_reasons": inactive_subscriptions
    }

@api_router.get("/perlas/{pearl_id}")
async def get_pearl_detail(pearl_id: str, current_user: User = Depends(get_current_user)):
    """Get full pearl detail (for deep link)"""
    pearl = await db.pearls.find_one({"id": pearl_id}, {"_id": 0})
    if not pearl:
        raise HTTPException(status_code=404, detail="Perla no encontrada")
    
    # Get subtopic and module info
    subtopic = await db.pearl_subtopics.find_one({"id": pearl["subtopic_id"]}, {"_id": 0})
    module = None
    if subtopic:
        module = await db.pearl_modules.find_one({"id": subtopic["module_id"]}, {"_id": 0})
    
    return {
        "pearl": pearl,
        "subtopic": subtopic,
        "module": module
    }

@api_router.post("/perlas/subscribe")
async def subscribe_to_pearls(
    preference: PearlPreferenceCreate,
    current_user: User = Depends(get_current_user)
):
    """Subscribe to daily pearls"""
    # Deactivate any existing preferences
    await db.pearl_preferences.update_many(
        {"user_id": current_user.id},
        {"$set": {"is_active": False}}
    )
    
    # Create new preference
    new_preference = UserPearlPreference(
        user_id=current_user.id,
        module_id=preference.module_id,
        subtopic_id=preference.subtopic_id,
        interval_minutes=preference.interval_minutes,
        start_hour=preference.start_hour,
        end_hour=preference.end_hour,
        timezone=preference.timezone
    )
    
    await db.pearl_preferences.insert_one(new_preference.dict())
    
    return {"message": "Suscripción activada", "preference_id": new_preference.id}

@api_router.post("/perlas/push-subscription")
async def save_push_subscription(
    subscription: PushSubscriptionCreate,
    current_user: User = Depends(get_current_user)
):
    """Save push subscription for notifications"""
    # Deactivate existing subscriptions for this endpoint
    await db.push_subscriptions.update_many(
        {"user_id": current_user.id, "endpoint": subscription.endpoint},
        {"$set": {"is_active": False}}
    )
    
    # Create new subscription
    new_sub = PushSubscription(
        user_id=current_user.id,
        endpoint=subscription.endpoint,
        keys=subscription.keys
    )
    
    await db.push_subscriptions.insert_one(new_sub.dict())
    
    return {"message": "Push subscription guardada", "subscription_id": new_sub.id}

@api_router.put("/perlas/pause")
async def pause_pearl_subscription(current_user: User = Depends(get_current_user)):
    """Pause pearl notifications"""
    result = await db.pearl_preferences.update_one(
        {"user_id": current_user.id, "is_active": True},
        {"$set": {"is_paused": True, "updated_at": datetime.now(timezone.utc)}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="No hay suscripción activa")
    return {"message": "Notificaciones pausadas"}

@api_router.put("/perlas/resume")
async def resume_pearl_subscription(current_user: User = Depends(get_current_user)):
    """Resume pearl notifications"""
    result = await db.pearl_preferences.update_one(
        {"user_id": current_user.id, "is_active": True},
        {"$set": {"is_paused": False, "updated_at": datetime.now(timezone.utc)}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="No hay suscripción activa")
    return {"message": "Notificaciones reanudadas"}

@api_router.delete("/perlas/unsubscribe")
async def unsubscribe_from_pearls(current_user: User = Depends(get_current_user)):
    """Unsubscribe from pearls"""
    await db.pearl_preferences.update_many(
        {"user_id": current_user.id},
        {"$set": {"is_active": False, "updated_at": datetime.now(timezone.utc)}}
    )
    await db.push_subscriptions.update_many(
        {"user_id": current_user.id},
        {"$set": {"is_active": False}}
    )
    return {"message": "Suscripción cancelada"}

# ==================== ADMIN PERLAS ENDPOINTS ====================

@api_router.post("/admin/perlas/modules")
async def create_pearl_module(
    data: dict,
    admin_user: User = Depends(get_admin_user)
):
    """Create a new pearl module (admin only)"""
    module = PearlModule(
        name=data["name"],
        description=data.get("description")
    )
    await db.pearl_modules.insert_one(module.dict())
    # Invalidate cache
    cache.invalidate("pearl_modules")
    cache.invalidate("pearl_all_subtopics")
    return {"message": "Módulo creado", "id": module.id}

@api_router.put("/admin/perlas/modules/{module_id}")
async def update_pearl_module(
    module_id: str,
    data: dict,
    admin_user: User = Depends(get_admin_user)
):
    """Update a pearl module"""
    update_data = {}
    if "name" in data:
        update_data["name"] = data["name"]
    if "description" in data:
        update_data["description"] = data["description"]
    if "is_active" in data:
        update_data["is_active"] = data["is_active"]
    
    result = await db.pearl_modules.update_one(
        {"id": module_id},
        {"$set": update_data}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Módulo no encontrado")
    # Invalidate cache
    cache.invalidate("pearl_modules")
    cache.invalidate("pearl_all_subtopics")
    return {"message": "Módulo actualizado"}

@api_router.delete("/admin/perlas/modules/{module_id}")
async def delete_pearl_module(
    module_id: str,
    admin_user: User = Depends(get_admin_user)
):
    """Delete a pearl module (soft delete)"""
    await db.pearl_modules.update_one(
        {"id": module_id},
        {"$set": {"is_active": False}}
    )
    # Invalidate cache
    cache.invalidate("pearl_modules")
    cache.invalidate("pearl_all_subtopics")
    return {"message": "Módulo eliminado"}

@api_router.post("/admin/perlas/subtopics")
async def create_pearl_subtopic(
    data: dict,
    admin_user: User = Depends(get_admin_user)
):
    """Create a new pearl subtopic (admin only)"""
    subtopic = PearlSubtopic(
        module_id=data["module_id"],
        name=data["name"],
        description=data.get("description")
    )
    await db.pearl_subtopics.insert_one(subtopic.dict())
    # Invalidate cache
    cache.invalidate_prefix("pearl_subtopics_")
    cache.invalidate("pearl_all_subtopics")
    return {"message": "Subtema creado", "id": subtopic.id}

@api_router.put("/admin/perlas/subtopics/{subtopic_id}")
async def update_pearl_subtopic(
    subtopic_id: str,
    data: dict,
    admin_user: User = Depends(get_admin_user)
):
    """Update a pearl subtopic"""
    update_data = {}
    if "name" in data:
        update_data["name"] = data["name"]
    if "description" in data:
        update_data["description"] = data["description"]
    if "is_active" in data:
        update_data["is_active"] = data["is_active"]
    
    result = await db.pearl_subtopics.update_one(
        {"id": subtopic_id},
        {"$set": update_data}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Subtema no encontrado")
    # Invalidate cache
    cache.invalidate_prefix("pearl_subtopics_")
    cache.invalidate("pearl_all_subtopics")
    return {"message": "Subtema actualizado"}

@api_router.delete("/admin/perlas/subtopics/{subtopic_id}")
async def delete_pearl_subtopic(
    subtopic_id: str,
    admin_user: User = Depends(get_admin_user)
):
    """Delete a pearl subtopic (soft delete)"""
    await db.pearl_subtopics.update_one(
        {"id": subtopic_id},
        {"$set": {"is_active": False}}
    )
    # Invalidate cache
    cache.invalidate_prefix("pearl_subtopics_")
    cache.invalidate("pearl_all_subtopics")
    return {"message": "Subtema eliminado"}

@api_router.post("/admin/perlas/pearls")
async def create_pearl(
    data: PearlCreate,
    admin_user: User = Depends(get_admin_user)
):
    """Create a new pearl (admin only)"""
    pearl = Pearl(
        subtopic_id=data.subtopic_id,
        title=data.title[:40],  # Enforce max 40 chars
        message=data.message[:200],  # Enforce max 200 chars
        long_body=data.long_body,
        tags=data.tags,
        difficulty=data.difficulty,
        is_active=data.is_active,
        priority=data.priority
    )
    await db.pearls.insert_one(pearl.dict())
    return {"message": "Perla creada", "id": pearl.id}

@api_router.put("/admin/perlas/pearls/{pearl_id}")
async def update_pearl(
    pearl_id: str,
    data: dict,
    admin_user: User = Depends(get_admin_user)
):
    """Update a pearl"""
    update_data = {"updated_at": datetime.now(timezone.utc)}
    
    for field in ["title", "message", "long_body", "tags", "difficulty", "is_active", "priority"]:
        if field in data:
            if field == "title":
                update_data[field] = data[field][:40]
            elif field == "message":
                update_data[field] = data[field][:200]
            else:
                update_data[field] = data[field]
    
    result = await db.pearls.update_one(
        {"id": pearl_id},
        {"$set": update_data}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Perla no encontrada")
    return {"message": "Perla actualizada"}

@api_router.delete("/admin/perlas/pearls/{pearl_id}")
async def delete_pearl(
    pearl_id: str,
    admin_user: User = Depends(get_admin_user)
):
    """Delete a pearl (soft delete)"""
    await db.pearls.update_one(
        {"id": pearl_id},
        {"$set": {"is_active": False}}
    )
    return {"message": "Perla eliminada"}

@api_router.get("/admin/perlas/all")
async def get_all_pearls_admin(admin_user: User = Depends(get_admin_user)):
    """Get all pearls with module/subtopic info (admin only)"""
    pearls = await db.pearls.find({}, {"_id": 0}).to_list(5000)
    
    # Get all modules and subtopics for lookup
    modules = {m["id"]: m for m in await db.pearl_modules.find({}, {"_id": 0}).to_list(100)}
    subtopics = {s["id"]: s for s in await db.pearl_subtopics.find({}, {"_id": 0}).to_list(500)}
    
    for pearl in pearls:
        subtopic = subtopics.get(pearl.get("subtopic_id"))
        if subtopic:
            pearl["subtopic_name"] = subtopic.get("name")
            module = modules.get(subtopic.get("module_id"))
            if module:
                pearl["module_name"] = module.get("name")
    
    return pearls

@api_router.get("/admin/perlas/quiz-categories")
async def get_quiz_categories(admin_user: User = Depends(get_admin_user)):
    """Get all specialties grouped into 5 main modules for pearl import"""
    
    # Define the 5 main modules
    MAIN_MODULES = {
        "Ginecología y Obstetricia": [],
        "Pediatría": [],
        "Cirugía": [],
        "Medicina Interna": [],
        "Otros": []
    }
    
    # Specialties that belong to "Otros"
    OTROS_SUBTEMAS = [
        "Alergia e Inmunología",
        "Analgesia y Anestesia", 
        "Enfermedades Lisosomales",
        "Genética",
        "Geriatría",
        "Medicina Familiar",
        "Medicina Física y Rehabilitación",
        "Nutriología",
        "Psiquiatría",
        "Urgencias Médico-Quirúrgicas"
    ]
    
    # Get all unique specialties from questions
    specialties = await db.questions.distinct("specialty")
    specialties = [s for s in specialties if s]
    
    # Process each specialty
    cirugia_subtemas = set()
    medicina_interna_subtemas = set()
    otros_subtemas = set()
    
    for specialty in specialties:
        # Check if it's a compound specialty (e.g., "Medicina Interna - Cardiología")
        if " - " in specialty:
            parts = specialty.split(" - ", 1)
            main_module = parts[0].strip()
            subtema = parts[1].strip()
            
            if main_module == "Cirugía":
                cirugia_subtemas.add(subtema)
            elif main_module == "Medicina Interna":
                medicina_interna_subtemas.add(subtema)
            # Ignore others with " - " pattern
        else:
            # It's a standalone specialty
            if specialty == "Ginecología y Obstetricia" or specialty == "Pediatría":
                # These don't have subtemas
                pass
            elif specialty in OTROS_SUBTEMAS or specialty.startswith("TEST-"):
                # Belongs to "Otros" or is a test
                if not specialty.startswith("TEST-"):
                    otros_subtemas.add(specialty)
            else:
                # Check if it's a known "Otros" subtema by partial match
                matched = False
                for otros_sub in OTROS_SUBTEMAS:
                    if otros_sub.lower() in specialty.lower() or specialty.lower() in otros_sub.lower():
                        otros_subtemas.add(otros_sub)
                        matched = True
                        break
                if not matched and specialty not in ["Ginecología y Obstetricia", "Pediatría"]:
                    # Unknown standalone - add to Otros
                    otros_subtemas.add(specialty)
    
    # Build the result
    result = [
        {
            "module": "Ginecología y Obstetricia",
            "subtemas": []  # No subtemas
        },
        {
            "module": "Pediatría", 
            "subtemas": []  # No subtemas
        },
        {
            "module": "Cirugía",
            "subtemas": sorted(list(cirugia_subtemas))
        },
        {
            "module": "Medicina Interna",
            "subtemas": sorted(list(medicina_interna_subtemas))
        },
        {
            "module": "Otros",
            "subtemas": sorted(list(otros_subtemas)) if otros_subtemas else OTROS_SUBTEMAS
        }
    ]
    
    return result

@api_router.post("/admin/perlas/bulk-import")
async def bulk_import_pearls(
    file: UploadFile = File(...),
    specialty: str = Form(None),
    topic: str = Form(None),
    admin_user: User = Depends(get_admin_user)
):
    """Bulk import pearls from CSV (max 20 pearls per file)
    
    If specialty is provided (and topic if applicable), the CSV only needs: titulo, mensaje, explicacion
    For Ginecología/Pediatría, topic is not required
    """
    if not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="El archivo debe ser CSV")
    
    try:
        content = await file.read()
        
        # Try multiple encodings
        encodings = ['utf-8', 'utf-8-sig', 'latin1', 'cp1252']
        df = None
        
        for encoding in encodings:
            try:
                df = pd.read_csv(io.StringIO(content.decode(encoding)))
                break
            except:
                continue
        
        if df is None:
            raise HTTPException(status_code=400, detail="No se pudo leer el archivo CSV")
        
        # Validate max 20 pearls
        if len(df) > 20:
            raise HTTPException(status_code=400, detail=f"Máximo 20 perlas por archivo. Tu archivo tiene {len(df)} filas.")
        
        # Check if specialty was provided (simplified mode)
        if specialty:
            # Simplified CSV mode - only needs titulo, mensaje, explicacion
            required = ['titulo', 'mensaje']
            missing = [c for c in required if c not in df.columns]
            if missing:
                raise HTTPException(status_code=400, detail=f"Columnas faltantes: {', '.join(missing)}")
            
            # Find or create module for the specialty
            module = await db.pearl_modules.find_one({"name": specialty})
            if not module:
                module = PearlModule(name=specialty)
                await db.pearl_modules.insert_one(module.dict())
                module = module.dict()
            
            # For modules without subtemas (Ginecología, Pediatría), use module name as subtopic
            subtopic_name = topic if topic else specialty
            
            # Find or create subtopic
            subtopic = await db.pearl_subtopics.find_one({
                "module_id": module["id"],
                "name": subtopic_name
            })
            if not subtopic:
                subtopic = PearlSubtopic(module_id=module["id"], name=subtopic_name)
                await db.pearl_subtopics.insert_one(subtopic.dict())
                subtopic = subtopic.dict()
            
            imported = 0
            errors = []
            
            for idx, row in df.iterrows():
                try:
                    long_body = None
                    if 'explicacion' in df.columns and pd.notna(row.get('explicacion')):
                        long_body = str(row['explicacion']).strip()
                    
                    pearl = Pearl(
                        subtopic_id=subtopic["id"],
                        title=str(row['titulo']).strip()[:40],
                        message=str(row['mensaje']).strip()[:200],
                        long_body=long_body,
                        is_active=True
                    )
                    await db.pearls.insert_one(pearl.dict())
                    imported += 1
                    
                except Exception as e:
                    errors.append(f"Fila {idx + 2}: {str(e)}")
            
            location = f"{specialty}" + (f" > {topic}" if topic else "")
            return {
                "message": f"Importación completada. {imported} perlas importadas a {location}.",
                "imported_count": imported,
                "errors": errors[:10]
            }
        
        else:
            # Full CSV mode - needs modulo, subtema, titulo, mensaje
            required = ['modulo', 'subtema', 'titulo', 'mensaje']
            missing = [c for c in required if c not in df.columns]
            if missing:
                raise HTTPException(status_code=400, detail=f"Columnas faltantes: {', '.join(missing)}")
            
            imported = 0
            errors = []
            
            for idx, row in df.iterrows():
                try:
                    # Find or create module
                    module_name = str(row['modulo']).strip()
                    module = await db.pearl_modules.find_one({"name": module_name})
                    if not module:
                        module = PearlModule(name=module_name)
                        await db.pearl_modules.insert_one(module.dict())
                        module = module.dict()
                    
                    # Find or create subtopic
                    subtopic_name = str(row['subtema']).strip()
                    subtopic = await db.pearl_subtopics.find_one({
                        "module_id": module["id"],
                        "name": subtopic_name
                    })
                    if not subtopic:
                        subtopic = PearlSubtopic(module_id=module["id"], name=subtopic_name)
                        await db.pearl_subtopics.insert_one(subtopic.dict())
                        subtopic = subtopic.dict()
                    
                    # Create pearl
                    long_body = None
                    if 'explicacion' in df.columns and pd.notna(row.get('explicacion')):
                        long_body = str(row['explicacion']).strip()
                    elif 'cuerpo_largo' in df.columns and pd.notna(row.get('cuerpo_largo')):
                        long_body = str(row['cuerpo_largo']).strip()
                    
                    pearl = Pearl(
                        subtopic_id=subtopic["id"],
                        title=str(row['titulo']).strip()[:40],
                        message=str(row['mensaje']).strip()[:200],
                        long_body=long_body,
                        tags=[t.strip() for t in str(row.get('tags', '')).split(',') if t.strip()] if pd.notna(row.get('tags')) else [],
                        difficulty=str(row.get('dificultad', '')).strip() if pd.notna(row.get('dificultad')) else None,
                        is_active=True
                    )
                    await db.pearls.insert_one(pearl.dict())
                    imported += 1
                    
                except Exception as e:
                    errors.append(f"Fila {idx + 2}: {str(e)}")
        
        return {
            "message": f"Importación completada. {imported} perlas importadas.",
            "imported_count": imported,
            "errors": errors[:10]
        }
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error procesando archivo: {str(e)}")

@api_router.get("/admin/perlas/analytics")
async def get_pearl_analytics(admin_user: User = Depends(get_admin_user)):
    """Get pearl subscription analytics (admin only)"""
    # Total subscribed users
    total_subscribed = await db.pearl_preferences.count_documents({"is_active": True})
    
    # Subscriptions by module
    pipeline = [
        {"$match": {"is_active": True}},
        {"$group": {"_id": "$module_id", "count": {"$sum": 1}}}
    ]
    by_module_raw = await db.pearl_preferences.aggregate(pipeline).to_list(100)
    
    # Get module names
    modules = {m["id"]: m["name"] for m in await db.pearl_modules.find({}, {"_id": 0}).to_list(100)}
    by_module = [{
        "module_id": item["_id"],
        "module_name": modules.get(item["_id"], "Desconocido"),
        "count": item["count"]
    } for item in by_module_raw]
    
    # Total pearls
    total_pearls = await db.pearls.count_documents({"is_active": True})
    
    # Delivery stats
    total_delivered = await db.pearl_delivery_logs.count_documents({"status": "delivered"})
    total_failed = await db.pearl_delivery_logs.count_documents({"status": "failed"})
    total_clicked = await db.pearl_delivery_logs.count_documents({"status": "clicked"})
    
    return {
        "total_subscribed_users": total_subscribed,
        "subscriptions_by_module": by_module,
        "total_pearls": total_pearls,
        "delivery_stats": {
            "delivered": total_delivered,
            "failed": total_failed,
            "clicked": total_clicked,
            "click_rate": round((total_clicked / total_delivered * 100) if total_delivered > 0 else 0, 2)
        }
    }

# Seed sample pearl data on startup
@app.on_event("startup")
async def seed_pearl_data():
    """Seed sample modules, subtopics and pearls"""
    # Check if already seeded
    existing = await db.pearl_modules.count_documents({})
    if existing > 0:
        print("✅ Pearl data already exists, skipping seed")
        return
    
    print("🌱 Seeding pearl sample data...")
    
    # Create modules
    modules_data = [
        {"name": "Ginecología y Obstetricia", "description": "Perlas de ginecología y obstetricia"},
        {"name": "Cirugía", "description": "Perlas de cirugía general y especializada"},
        {"name": "Pediatría", "description": "Perlas de pediatría"},
        {"name": "Medicina Interna", "description": "Perlas de medicina interna"},
        {"name": "Otros", "description": "Otras especialidades"}
    ]
    
    created_modules = []
    for m_data in modules_data:
        module = PearlModule(**m_data)
        await db.pearl_modules.insert_one(module.dict())
        created_modules.append(module)
    
    # Create subtopics with sample pearls
    subtopics_and_pearls = {
        "Ginecología y Obstetricia": [
            {
                "name": "Embarazo de alto riesgo",
                "pearls": [
                    {"title": "Preeclampsia", "message": "La preeclampsia se define por HTA >140/90 + proteinuria >300mg/24h después de semana 20. Considerar sulfato de magnesio para prevención de eclampsia.", "long_body": "La preeclampsia es un síndrome multisistémico que complica 2-8% de embarazos. Se caracteriza por hipertensión de novo después de la semana 20 de gestación con proteinuria significativa. El sulfato de magnesio es el fármaco de elección para prevenir convulsiones eclámpticas."},
                    {"title": "HELLP", "message": "Síndrome HELLP: Hemólisis, Elevated Liver enzymes, Low Platelets. Emergencia obstétrica que requiere terminación del embarazo.", "long_body": "El síndrome HELLP es una complicación severa de la preeclampsia. Se presenta con hemólisis microangiopática, elevación de transaminasas hepáticas y trombocitopenia. El tratamiento definitivo es la terminación del embarazo."},
                    {"title": "Diabetes gestacional", "message": "Tamizaje de DMG con carga de 50g glucosa entre semanas 24-28. Si ≥140 mg/dL, realizar curva de tolerancia.", "long_body": "La diabetes gestacional aumenta el riesgo de macrosomía fetal, distocia de hombros y complicaciones neonatales. El tamizaje universal está recomendado entre las semanas 24-28."}
                ]
            },
            {
                "name": "Trabajo de parto",
                "pearls": [
                    {"title": "Fases del trabajo de parto", "message": "Fase latente: dilatación 0-6cm. Fase activa: 6-10cm con velocidad ≥1cm/hora en nulíparas.", "long_body": "La fase latente puede durar hasta 20 horas en nulíparas. La fase activa inicia típicamente a los 6cm y debe progresar al menos 1cm por hora. La detención del trabajo de parto se diagnostica con ≥4 horas sin cambio cervical con contracciones adecuadas."},
                    {"title": "Partograma", "message": "El partograma es herramienta esencial para monitorizar progreso del trabajo de parto. Alerta si cruza línea de acción.", "long_body": "El partograma de la OMS incluye líneas de alerta y acción. Si la curva de dilatación cruza la línea de acción, indica necesidad de intervención."}
                ]
            }
        ],
        "Medicina Interna": [
            {
                "name": "Cardiología",
                "pearls": [
                    {"title": "Infarto STEMI", "message": "Tiempo puerta-balón meta <90 min. Tiempo puerta-aguja <30 min si no hay ICP disponible.", "long_body": "En el infarto con elevación del ST, la reperfusión coronaria temprana es crítica. La angioplastia primaria es el tratamiento de elección si está disponible en <120 minutos del primer contacto médico."},
                    {"title": "Insuficiencia cardíaca", "message": "Pilares del tratamiento de IC con FE reducida: IECA/ARA2/ARNI + betabloqueador + antagonista mineralocorticoide + iSGLT2.", "long_body": "La terapia cuádruple ha demostrado reducir mortalidad y hospitalizaciones en IC con fracción de eyección reducida. Los iSGLT2 son la adición más reciente con beneficio demostrado."},
                    {"title": "Fibrilación auricular", "message": "CHA2DS2-VASc ≥2 en hombres o ≥3 en mujeres indica anticoagulación. DOACs preferidos sobre warfarina.", "long_body": "La fibrilación auricular aumenta 5x el riesgo de EVC. Los anticoagulantes orales directos tienen mejor perfil de seguridad que warfarina y no requieren monitoreo de INR."}
                ]
            },
            {
                "name": "Endocrinología",
                "pearls": [
                    {"title": "Crisis tirotóxica", "message": "Escala de Burch-Wartofsky >45 puntos sugiere crisis tirotóxica. Iniciar PTU, yodo, betabloqueador y glucocorticoides.", "long_body": "La tormenta tiroidea es emergencia endocrina con mortalidad >20%. El tratamiento incluye bloqueo de síntesis (PTU), bloqueo de liberación (yodo después de PTU), control de síntomas (betabloqueador) y prevención de conversión periférica (hidrocortisona)."},
                    {"title": "CAD vs EHH", "message": "CAD: pH<7.3, bicarbonato<18, cetonas positivas. EHH: glucosa>600, osmolaridad>320, sin cetosis significativa.", "long_body": "La cetoacidosis diabética es más común en DM1, mientras que el estado hiperosmolar hiperglucémico predomina en DM2. Ambos requieren hidratación agresiva, insulina IV y corrección de electrolitos."}
                ]
            }
        ],
        "Pediatría": [
            {
                "name": "Neonatología",
                "pearls": [
                    {"title": "Apgar", "message": "Apgar evalúa: Aspecto, Pulso, Gesticulación, Actividad, Respiración. Se mide al minuto 1 y 5. <7 indica necesidad de intervención.", "long_body": "El puntaje de Apgar es una evaluación rápida de la transición neonatal. Un Apgar de 1 minuto bajo indica necesidad de reanimación. El Apgar de 5 minutos tiene mejor correlación con pronóstico."},
                    {"title": "Ictericia neonatal", "message": "Bilirrubina indirecta >15 mg/dL en RN a término requiere fototerapia. Niveles >25 mg/dL riesgo de kernicterus.", "long_body": "La hiperbilirrubinemia neonatal es frecuente pero potencialmente peligrosa. La fototerapia convierte la bilirrubina en formas hidrosolubles que se excretan sin conjugación hepática."}
                ]
            },
            {
                "name": "Infectología pediátrica",
                "pearls": [
                    {"title": "Meningitis bacteriana", "message": "LCR turbio, glucosa baja, proteínas altas, PMN predominio = meningitis bacteriana. Iniciar ceftriaxona + vancomicina empíricamente.", "long_body": "La meningitis bacteriana en niños es emergencia que requiere tratamiento antibiótico inmediato. Los agentes más comunes varían por edad: GBS y E. coli en neonatos, S. pneumoniae y N. meningitidis en mayores."}
                ]
            }
        ],
        "Cirugía": [
            {
                "name": "Abdomen agudo",
                "pearls": [
                    {"title": "Apendicitis", "message": "Punto de McBurney: unión del tercio externo con dos tercios internos de línea umbílico-iliaca derecha. Signo de Rovsing patognomónico.", "long_body": "La apendicitis aguda es la emergencia quirúrgica abdominal más común. El diagnóstico es principalmente clínico, apoyado por leucocitosis y estudios de imagen. La apendicectomía laparoscópica es el estándar actual."},
                    {"title": "Colecistitis aguda", "message": "Criterios de Tokio para colecistitis: fiebre, dolor en CSD, Murphy positivo, leucocitosis, PCR elevada + hallazgos de imagen.", "long_body": "La colecistitis aguda litiásica es la complicación más común de la colelitiasis. La colecistectomía laparoscópica temprana (primeras 72h) es preferible a la cirugía diferida."}
                ]
            }
        ],
        "Otros": [
            {
                "name": "Urgencias",
                "pearls": [
                    {"title": "ACLS - Ritmos desfibrilables", "message": "FV y TV sin pulso: desfibrilar inmediatamente. Adrenalina cada 3-5 min. Amiodarona después de 3 choques.", "long_body": "En paro cardíaco, la desfibrilación temprana es el determinante más importante de supervivencia para ritmos desfibrilables. La RCP de alta calidad debe interrumpirse mínimamente."},
                    {"title": "Trauma - ABCDE", "message": "Evaluación primaria de trauma: Airway, Breathing, Circulation, Disability, Exposure. Hemorragia es causa prevenible #1 de muerte.", "long_body": "La evaluación sistemática ABCDE permite identificar y tratar lesiones que amenazan la vida en orden de prioridad. El control de hemorragia externa es prioritario en la evaluación inicial."}
                ]
            }
        ]
    }
    
    for module in created_modules:
        if module.name in subtopics_and_pearls:
            for subtopic_data in subtopics_and_pearls[module.name]:
                subtopic = PearlSubtopic(
                    module_id=module.id,
                    name=subtopic_data["name"]
                )
                await db.pearl_subtopics.insert_one(subtopic.dict())
                
                for pearl_data in subtopic_data["pearls"]:
                    pearl = Pearl(
                        subtopic_id=subtopic.id,
                        title=pearl_data["title"],
                        message=pearl_data["message"],
                        long_body=pearl_data.get("long_body")
                    )
                    await db.pearls.insert_one(pearl.dict())
    
    print("✅ Pearl sample data seeded successfully")

# ==================== JOURNAL SYSTEM ====================

@api_router.post("/admin/journal/upload")
async def upload_journal_articles(
    file: UploadFile = File(...),
    admin_user: User = Depends(get_admin_user)
):
    """Upload Excel file with journal article topics. Columns: Tema, Antecedentes, Metodos, Resultados, Conclusiones"""
    import openpyxl
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    required = ['Tema', 'Antecedentes', 'Metodos', 'Resultados', 'Conclusiones']
    for r in required:
        if r not in headers:
            raise HTTPException(status_code=400, detail=f"Columna '{r}' no encontrada. Columnas requeridas: {required}")

    col_idx = {h: i for i, h in enumerate(headers)}
    articles = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        tema = row[col_idx['Tema']]
        if not tema:
            continue
        articles.append({
            "index": row_idx,
            "tema": str(tema).strip(),
            "antecedentes": str(row[col_idx['Antecedentes']] or "").strip(),
            "metodos": str(row[col_idx['Metodos']] or "").strip(),
            "resultados": str(row[col_idx['Resultados']] or "").strip(),
            "conclusiones": str(row[col_idx['Conclusiones']] or "").strip(),
        })

    # Replace all articles in the collection
    await db.journal_articles.delete_many({})
    if articles:
        await db.journal_articles.insert_many(articles)

    # Reset used tracking
    await db.journal_used.delete_many({})

    # Select the first article immediately
    await select_next_journal_article()

    return {"message": f"Se cargaron {len(articles)} artículos del journal", "count": len(articles)}


@api_router.get("/admin/journal/status")
async def get_journal_status(admin_user: User = Depends(get_admin_user)):
    """Get the current journal system status."""
    total = await db.journal_articles.count_documents({})
    used = await db.journal_used.count_documents({})
    current = await db.journal_current.find_one({}, {"_id": 0})
    return {
        "total_articles": total,
        "used_articles": used,
        "remaining_articles": total - used,
        "current_article": current
    }


@api_router.get("/journal/today")
async def get_journal_today(current_user: User = Depends(get_current_user)):
    """Get today's journal article with authors."""
    current = await db.journal_current.find_one({}, {"_id": 0})
    if not current:
        return None
    return current


@api_router.get("/journal/history")
async def get_journal_history(current_user: User = Depends(get_current_user)):
    """Get last 30 published journals."""
    history = await db.journal_history.find(
        {},
        {"_id": 0}
    ).sort("selected_at", -1).to_list(30)
    return history


@api_router.post("/admin/journal/rotate")
async def manual_journal_rotate(admin_user: User = Depends(get_admin_user)):
    """Manually trigger journal article rotation (for testing)."""
    result = await select_next_journal_article(snapshot_authors=True)
    return result


async def select_next_journal_article(snapshot_authors=True):
    """Select the next random unused article for the journal, snapshot daily ranking as authors."""
    import random

    total = await db.journal_articles.count_documents({})
    if total == 0:
        return {"message": "No hay artículos cargados"}

    used_docs = await db.journal_used.find({}, {"_id": 0, "index": 1}).to_list(5000)
    used_indices = {d["index"] for d in used_docs}

    # If all used, reset
    if len(used_indices) >= total:
        await db.journal_used.delete_many({})
        used_indices = set()

    # Get unused articles
    unused = await db.journal_articles.find(
        {"index": {"$nin": list(used_indices)}},
        {"_id": 0}
    ).to_list(5000)

    if not unused:
        return {"message": "No hay artículos disponibles"}

    chosen = random.choice(unused)

    # Mark as used
    await db.journal_used.insert_one({"index": chosen["index"], "used_at": datetime.now(timezone.utc)})

    # Get current issue number
    prev = await db.journal_current.find_one({}, {"_id": 0})
    issue_number = (prev.get("issue_number", 142) if prev else 142) + 1

    # Archive current journal to history before replacing
    if prev and prev.get("tema"):
        prev_history = {k: v for k, v in prev.items() if k != "_id"}
        await db.journal_history.insert_one(prev_history)

    # Get CDMX date
    cdmx_tz = pytz.timezone("America/Mexico_City")
    now_cdmx = datetime.now(cdmx_tz)
    date_str = now_cdmx.strftime("%B %d, %Y")

    # Snapshot daily ranking top 10 as authors
    # Use the PREVIOUS cutoff (yesterday's 9 PM to today's 9 PM)
    # because at 9 PM rotation time, the current cutoff resets to NOW
    authors = []
    if snapshot_authors:
        current_cutoff = get_daily_cutoff_utc()
        previous_cutoff = current_cutoff - timedelta(days=1)
        pipeline = [
            {"$match": {
                "created_at": {"$gte": previous_cutoff, "$lt": current_cutoff},
                "activity_type": {"$in": ["quiz", "duel_win", "escape_room", "simulacro", "imagendx"]}
            }},
            {"$group": {
                "_id": "$user_id",
                "quiz_count": {"$sum": {"$cond": [{"$eq": ["$activity_type", "quiz"]}, 1, 0]}},
                "duel_win_count": {"$sum": {"$cond": [{"$eq": ["$activity_type", "duel_win"]}, 1, 0]}},
                "escape_room_count": {"$sum": {"$cond": [{"$eq": ["$activity_type", "escape_room"]}, 1, 0]}},
                "simulacro_count": {"$sum": {"$cond": [{"$eq": ["$activity_type", "simulacro"]}, 1, 0]}},
                "imagendx_count": {"$sum": {"$cond": [{"$eq": ["$activity_type", "imagendx"]}, 1, 0]}},
            }},
            {"$addFields": {
                "score": {"$add": [
                    {"$multiply": ["$quiz_count", 25]},
                    {"$multiply": ["$duel_win_count", 5]},
                    {"$multiply": ["$escape_room_count", 8]},
                    {"$multiply": ["$simulacro_count", 50]},
                    {"$multiply": ["$imagendx_count", 2]}
                ]}
            }},
            {"$match": {"score": {"$gt": 0}}},
            {"$sort": {"score": -1}},
            {"$limit": 10}
        ]
        top_users = await db.point_transactions.aggregate(pipeline).to_list(10)
        for entry in top_users:
            user = await db.users.find_one({"id": entry["_id"]}, {"_id": 0})
            if user:
                authors.append({
                    "user_id": entry["_id"],
                    "full_name": user.get("full_name", "Unknown"),
                    "universidad": user.get("universidad", ""),
                    "score": entry["score"]
                })

    # If no daily authors, fallback to overall top 10
    if not authors:
        fallback_pipeline = [
            {"$group": {"_id": "$user_id", "total_points": {"$sum": "$points"}}},
            {"$sort": {"total_points": -1}},
            {"$limit": 10}
        ]
        fallback_users = await db.point_transactions.aggregate(fallback_pipeline).to_list(10)
        for entry in fallback_users:
            user = await db.users.find_one({"id": entry["_id"]}, {"_id": 0})
            if user:
                authors.append({
                    "user_id": entry["_id"],
                    "full_name": user.get("full_name", "Unknown"),
                    "universidad": user.get("universidad", ""),
                    "score": entry["total_points"]
                })

    # Save as current
    current_article = {
        "tema": chosen["tema"],
        "antecedentes": chosen["antecedentes"],
        "metodos": chosen["metodos"],
        "resultados": chosen["resultados"],
        "conclusiones": chosen["conclusiones"],
        "issue_number": issue_number,
        "date_str": date_str,
        "selected_at": datetime.now(timezone.utc).isoformat(),
        "article_index": chosen["index"],
        "authors": authors
    }

    await db.journal_current.delete_many({})
    await db.journal_current.insert_one(current_article)

    return {"message": f"Artículo #{chosen['index']} seleccionado", "article": {k: v for k, v in current_article.items() if k != "_id"}}


async def journal_daily_rotation():
    """Scheduled job: rotate journal article daily at 9 PM CDMX."""
    try:
        logger.info("📰 Running daily journal rotation...")
        await select_next_journal_article(snapshot_authors=True)
        logger.info("✅ Journal article rotated successfully")
    except Exception as e:
        logger.error(f"❌ Journal rotation error: {e}")

# ==================== PERLAS SCHEDULER ====================

# Initialize scheduler
scheduler = AsyncIOScheduler()

async def send_push_notification(subscription_data: dict, payload: dict) -> bool:
    """Send a push notification using pywebpush"""
    endpoint = subscription_data.get("endpoint", "unknown")
    user_id = subscription_data.get("user_id", "unknown")
    
    try:
        logger.info(f"📤 Sending push to user {user_id}, endpoint: {endpoint[:50]}...")
        webpush(
            subscription_info={
                "endpoint": subscription_data["endpoint"],
                "keys": subscription_data["keys"]
            },
            data=json.dumps(payload),
            vapid_private_key=VAPID_PRIVATE_KEY,
            vapid_claims=VAPID_CLAIMS
        )
        logger.info(f"✅ Push sent successfully to user {user_id}")
        return True
    except WebPushException as e:
        status_code = e.response.status_code if e.response else "N/A"
        logger.error(f"❌ Push failed for user {user_id}: Status {status_code} - {e}")
        # If subscription is invalid (expired or unsubscribed), mark it inactive
        if e.response and e.response.status_code in [404, 410]:
            logger.warning(f"🗑️ Marking subscription as inactive for user {user_id} (expired/unsubscribed)")
            await db.push_subscriptions.update_one(
                {"endpoint": subscription_data["endpoint"]},
                {"$set": {"is_active": False, "deactivated_at": datetime.now(timezone.utc).isoformat(), "deactivate_reason": f"HTTP {status_code}"}, "$inc": {"error_count": 1}}
            )
        return False
    except Exception as e:
        logger.error(f"❌ Unexpected push error for user {user_id}: {e}")
        return False

async def get_next_pearl_for_user(user_id: str, subtopic_id: str) -> Optional[dict]:
    """Get the next pearl that hasn't been sent to this user"""
    # Get all pearls for the subtopic
    all_pearls = await db.pearls.find(
        {"subtopic_id": subtopic_id, "is_active": True},
        {"_id": 0}
    ).sort("priority", -1).to_list(1000)
    
    if not all_pearls:
        return None
    
    # Get pearls already sent to this user today
    today_start = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    sent_today = await db.pearl_delivery_logs.find(
        {
            "user_id": user_id,
            "sent_at": {"$gte": today_start}
        },
        {"pearl_id": 1}
    ).to_list(1000)
    sent_today_ids = {log["pearl_id"] for log in sent_today}
    
    # Get all pearls ever sent to this user (for rotation)
    all_sent = await db.pearl_delivery_logs.find(
        {"user_id": user_id},
        {"pearl_id": 1}
    ).to_list(10000)
    all_sent_ids = {log["pearl_id"] for log in all_sent}
    
    # First, try to find a pearl not sent today
    for pearl in all_pearls:
        if pearl["id"] not in sent_today_ids:
            # Prefer pearls never sent
            if pearl["id"] not in all_sent_ids:
                return pearl
    
    # If all pearls have been sent at least once, but not today, pick any unsent today
    for pearl in all_pearls:
        if pearl["id"] not in sent_today_ids:
            return pearl
    
    # If all pearls sent today (shouldn't happen with proper frequency), return None
    return None

def is_within_user_hours(preference: dict) -> bool:
    """Check if current time is within user's configured hours"""
    try:
        user_tz = pytz.timezone(preference.get("timezone", "America/Mexico_City"))
        now_user = datetime.now(user_tz)
        
        start_parts = preference.get("start_hour", "08:00").split(":")
        end_parts = preference.get("end_hour", "22:00").split(":")
        
        start_hour = int(start_parts[0])
        end_hour = int(end_parts[0])
        current_hour = now_user.hour
        
        return start_hour <= current_hour < end_hour
    except Exception as e:
        logger.error(f"Error checking user hours: {e}")
        return False

def should_send_pearl_now(preference: dict) -> bool:
    """Check if enough time has passed since last pearl was sent based on interval_minutes"""
    try:
        interval_minutes = preference.get("interval_minutes", 60)
        last_sent_str = preference.get("last_sent_at")
        
        if not last_sent_str:
            # Never sent before, should send now
            return True
        
        # Parse last sent time
        try:
            last_sent = datetime.fromisoformat(last_sent_str.replace('Z', '+00:00'))
        except:
            return True
        
        # Calculate time since last send
        now = datetime.now(timezone.utc)
        minutes_since_last = (now - last_sent).total_seconds() / 60
        
        # Check if enough time has passed
        return minutes_since_last >= interval_minutes
    except Exception as e:
        logger.error(f"Error checking send interval: {e}")
        return False

async def process_pearl_notifications():
    """Main scheduler job - process and send pearl notifications"""
    try:
        logger.info("🔔 Running pearl notification scheduler...")
        
        # Get all active, non-paused preferences
        preferences = await db.pearl_preferences.find(
            {"is_active": True, "is_paused": False},
            {"_id": 0}
        ).to_list(10000)
        
        logger.info(f"Found {len(preferences)} active preferences")
        
        sent_count = 0
        
        for pref in preferences:
            try:
                # Check if within user's hours
                if not is_within_user_hours(pref):
                    continue
                
                # Check if enough time has passed since last pearl (based on interval_minutes)
                if not should_send_pearl_now(pref):
                    continue
                
                # Get next pearl
                pearl = await get_next_pearl_for_user(pref["user_id"], pref["subtopic_id"])
                if not pearl:
                    logger.warning(f"No pearls available for user {pref['user_id']}")
                    continue
                
                # Get user's push subscriptions
                subscriptions = await db.push_subscriptions.find(
                    {"user_id": pref["user_id"], "is_active": True},
                    {"_id": 0}
                ).to_list(10)
                
                if not subscriptions:
                    continue
                
                # Prepare notification payload
                payload = {
                    "title": f"💎 {pearl['title']}",
                    "message": pearl["message"],
                    "pearl_id": pearl["id"],
                    "icon": "/icons/icon-192x192.png",
                    "badge": "/icons/icon-96x96.png"
                }
                
                # Send to all user's subscriptions
                success = False
                for sub in subscriptions:
                    if await send_push_notification(sub, payload):
                        success = True
                
                # Log delivery
                log_entry = PearlDeliveryLog(
                    user_id=pref["user_id"],
                    pearl_id=pearl["id"],
                    preference_id=pref["id"],
                    status="delivered" if success else "failed"
                )
                await db.pearl_delivery_logs.insert_one(log_entry.dict())
                
                if success:
                    sent_count += 1
                    # Update last_sent_at for this preference
                    await db.pearl_preferences.update_one(
                        {"id": pref["id"]},
                        {"$set": {"last_sent_at": datetime.now(timezone.utc).isoformat()}}
                    )
                    logger.info(f"✅ Sent pearl '{pearl['title']}' to user {pref['user_id']}")
                
            except Exception as e:
                logger.error(f"Error processing preference {pref.get('id')}: {e}")
                continue
        
        logger.info(f"🔔 Scheduler complete. Sent {sent_count} notifications.")
        
    except Exception as e:
        logger.error(f"Scheduler error: {e}")

@app.on_event("startup")
async def start_scheduler():
    """Start the pearl notification scheduler and journal rotation scheduler"""
    # Pearl notifications - every 15 minutes
    scheduler.add_job(
        process_pearl_notifications,
        IntervalTrigger(minutes=15),
        id="pearl_notifications",
        replace_existing=True
    )
    # Journal rotation - daily at 9 PM CDMX (UTC-6 = 3 AM UTC)
    cdmx_tz = pytz.timezone("America/Mexico_City")
    scheduler.add_job(
        journal_daily_rotation,
        CronTrigger(hour=21, minute=0, timezone=cdmx_tz),
        id="journal_rotation",
        replace_existing=True
    )
    scheduler.start()
    logger.info("🚀 Schedulers started (pearls: 15min, journal: 9PM CDMX daily)")

# Manual trigger endpoint for testing
@api_router.post("/admin/perlas/trigger-scheduler")
async def trigger_scheduler(admin_user: User = Depends(get_admin_user)):
    """Manually trigger the pearl notification scheduler (admin only)"""
    await process_pearl_notifications()
    return {"message": "Scheduler triggered manually"}

# Track pearl click (called from service worker)
@api_router.post("/perlas/track-click/{pearl_id}")
async def track_pearl_click(pearl_id: str):
    """Track when a user clicks on a pearl notification"""
    await db.pearl_delivery_logs.update_many(
        {"pearl_id": pearl_id, "status": "delivered"},
        {"$set": {"status": "clicked"}}
    )
    return {"message": "Click tracked"}

# ==================== PRESENTACIONES (PDFs) ENDPOINTS ====================

# Pydantic models for presentations
class PresentationCreate(BaseModel):
    title: str
    module: str  # Ginecología y Obstetricia, Cirugía, Pediatría, Medicina Interna, Otros
    submodule: Optional[str] = None  # For modules with submodules
    description: Optional[str] = None

class Presentation(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    title: str
    module: str
    submodule: Optional[str] = None
    description: Optional[str] = None
    file_path: str
    file_name: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Presentation modules structure (same as specialtyStructure in frontend)
PRESENTATION_MODULES = {
    "Ginecología y Obstetricia": {"hasSubmodules": False, "submodules": []},
    "Pediatría": {"hasSubmodules": False, "submodules": []},
    "Cirugía": {
        "hasSubmodules": True,
        "submodules": [
            "Angiología", "Cirugía General", "Cirugía Maxilofacial", "Cirugía Plástica",
            "Neurocirugía", "Oftalmología", "Otorrinolaringología", "Trasplantes",
            "Traumatología y Ortopedia", "Urología"
        ]
    },
    "Medicina Interna": {
        "hasSubmodules": True,
        "submodules": [
            "Cardiología", "Dermatología", "Endocrinología", "Gastroenterología",
            "Hematología", "Infectología", "Medicina Interna (general)", "Nefrología",
            "Neumología", "Neurología", "Oncología", "Reumatología"
        ]
    },
    "Otros": {
        "hasSubmodules": True,
        "submodules": [
            "Alergia e Inmunología", "Analgesia y Anestesia", "Enfermedades Lisosomales",
            "Genética", "Geriatría", "Medicina Familiar", "Medicina Física y Rehabilitación",
            "Nutriología", "Psiquiatría", "Urgencias Médico-Quirúrgicas"
        ]
    }
}

# Upload directory for presentations
PRESENTATIONS_DIR = ROOT_DIR / "uploads" / "presentations"
if PRESENTATIONS_DIR.is_symlink():
    # Symlink exists but target might not — ensure target dir exists
    target = PRESENTATIONS_DIR.resolve()
    target.mkdir(parents=True, exist_ok=True)
elif not PRESENTATIONS_DIR.exists():
    PRESENTATIONS_DIR.mkdir(parents=True, exist_ok=True)

# Logo path for watermark
LOGO_PATH = ROOT_DIR / "uploads" / "logo_watermark.png"

def add_watermark_to_pdf_fast(input_path: Path, output_path: Path, logo_path: Path) -> bool:
    """Add logo watermark to PDF - optimized version for large files
    
    Only watermarks first 3 pages for speed on large PDFs.
    """
    try:
        from PyPDF2 import PdfReader, PdfWriter
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import letter
        import io
        
        # Create watermark PDF in memory (faster than disk)
        watermark_buffer = io.BytesIO()
        c = canvas.Canvas(watermark_buffer, pagesize=letter)
        
        # Check if logo exists
        if logo_path.exists():
            page_width, page_height = letter
            # Make logo larger to cover more of the page
            logo_width = 300
            logo_height = 150
            x = (page_width - logo_width) / 2
            y = (page_height - logo_height) / 2
            c.saveState()
            c.setFillAlpha(0.12)  # Semi-transparent
            try:
                c.drawImage(str(logo_path), x, y, width=logo_width, height=logo_height, 
                           preserveAspectRatio=True, mask='auto')
            except Exception as img_err:
                logging.warning(f"Could not add image watermark: {img_err}")
                # Fallback to text watermark
                c.setFillAlpha(0.1)
                c.setFont("Helvetica-Bold", 40)
                c.translate(page_width/2, page_height/2)
                c.rotate(45)
                c.drawCentredString(0, 0, "Proyecto Residente")
            c.restoreState()
        
        c.save()
        watermark_buffer.seek(0)
        
        # Read watermark from memory
        watermark_reader = PdfReader(watermark_buffer)
        watermark_page = watermark_reader.pages[0]
        
        # Process original PDF
        reader = PdfReader(str(input_path))
        writer = PdfWriter()
        
        total_pages = len(reader.pages)
        
        for i, page in enumerate(reader.pages):
            # Only watermark first 5 pages and every 10th page after for speed
            if i < 5 or i % 10 == 0:
                page.merge_page(watermark_page)
            writer.add_page(page)
        
        # Write output
        with open(output_path, 'wb') as output_file:
            writer.write(output_file)
        
        logging.info(f"Watermark added to PDF ({total_pages} pages)")
        return True
        
    except Exception as e:
        logging.error(f"Error adding watermark: {e}")
        # If watermark fails, just copy the original
        import shutil
        shutil.copy(input_path, output_path)
        return False

@api_router.get("/presentations/modules")
async def get_presentation_modules(current_user: User = Depends(get_current_user)):
    """Get all presentation modules structure"""
    return PRESENTATION_MODULES

@api_router.get("/presentations/list")
async def list_presentations(
    module: Optional[str] = None,
    submodule: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """List presentations, optionally filtered by module/submodule"""
    query = {}
    if module:
        query["module"] = module
    if submodule:
        query["submodule"] = submodule
    
    presentations = await db.presentations.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return presentations

@api_router.get("/presentations/view/{presentation_id}")
async def view_presentation(
    request: Request,
    presentation_id: str,
    token: Optional[str] = None
):
    """Stream PDF for viewing (without download)
    
    Supports Range requests for progressive loading.
    Priority: Object Storage > GridFS > Disk
    """
    # Validate token from query parameter
    if not token:
        raise HTTPException(status_code=401, detail="Token requerido")
    
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
        user_identifier = payload.get("sub")
        if not user_identifier:
            raise HTTPException(status_code=401, detail="Token inválido")
        
        # Try to find user by id first, then by email (for compatibility)
        user = await db.users.find_one({"id": user_identifier})
        if not user:
            user = await db.users.find_one({"email": user_identifier})
        if not user:
            raise HTTPException(status_code=401, detail="Usuario no encontrado")
        
        if not user.get("is_approved", False):
            raise HTTPException(status_code=403, detail="Usuario no aprobado")
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expirado")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token inválido")
    
    presentation = await db.presentations.find_one({"id": presentation_id})
    if not presentation:
        raise HTTPException(status_code=404, detail="Presentación no encontrada")
    
    file_name = presentation.get("file_name", "presentation.pdf")
    
    # Priority 1: Object Storage (if migrated)
    if presentation.get("storage_type") == "external" and presentation.get("pdf_key"):
        file_data = presentation_storage.get_pdf(presentation["pdf_key"])
        if file_data:
            return _serve_pdf_bytes(request, file_data, file_name)
    
    # Priority 2: GridFS (existing behavior)
    gridfs_meta = await db.presentations_fs.files.find_one({"metadata.presentation_id": presentation_id})
    if gridfs_meta:
        grid_out = await fs_bucket.open_download_stream(gridfs_meta["_id"])
        file_data = await grid_out.read()
        return _serve_pdf_bytes(request, file_data, file_name)
    
    # Priority 3: Fallback to disk
    file_path = PRESENTATIONS_DIR / presentation["file_path"]
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Archivo de presentación no encontrado en el servidor. Necesita ser subido nuevamente.")
    
    with open(file_path, "rb") as f:
        file_data = f.read()
    return _serve_pdf_bytes(request, file_data, file_name)


def _serve_pdf_bytes(request: Request, data: bytes, filename: str) -> Response:
    """Serve PDF bytes with proper headers, including Range request support."""
    total = len(data)
    range_header = request.headers.get("range")
    
    headers = {
        "Content-Disposition": f'inline; filename="{filename}"',
        "Accept-Ranges": "bytes",
        "X-Content-Type-Options": "nosniff",
        "Cache-Control": "private, max-age=3600",
    }
    
    if range_header:
        # Parse Range: bytes=start-end
        try:
            range_spec = range_header.replace("bytes=", "")
            parts = range_spec.split("-")
            start = int(parts[0]) if parts[0] else 0
            end = int(parts[1]) if parts[1] else total - 1
            end = min(end, total - 1)
            length = end - start + 1
            headers["Content-Range"] = f"bytes {start}-{end}/{total}"
            headers["Content-Length"] = str(length)
            return Response(
                content=data[start:end + 1],
                status_code=206,
                media_type="application/pdf",
                headers=headers,
            )
        except (ValueError, IndexError):
            pass
    
    headers["Content-Length"] = str(total)
    return Response(
        content=data,
        status_code=200,
        media_type="application/pdf",
        headers=headers,
    )

@api_router.post("/admin/presentations/upload")
async def upload_presentation(
    title: str = Form(...),
    module: str = Form(...),
    submodule: Optional[str] = Form(None),
    description: Optional[str] = Form(None),
    file: UploadFile = File(...),
    admin_user: User = Depends(get_admin_user)
):
    """Upload a new presentation (admin only)"""
    # Validate module
    if module not in PRESENTATION_MODULES:
        raise HTTPException(status_code=400, detail="Módulo inválido")
    
    # Validate submodule if module has submodules
    module_info = PRESENTATION_MODULES[module]
    if module_info["hasSubmodules"]:
        if not submodule or submodule not in module_info["submodules"]:
            raise HTTPException(status_code=400, detail="Submódulo requerido o inválido")
    else:
        submodule = None  # Clear submodule for modules without submodules
    
    # Validate file type
    if not file.filename.lower().endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Solo se permiten archivos PDF")
    
    # Generate unique filename
    file_id = str(uuid.uuid4())
    safe_filename = f"{file_id}.pdf"
    original_filename = file.filename
    
    # Read file content
    content = await file.read()
    
    # Save original file temporarily for watermarking
    temp_path = PRESENTATIONS_DIR / f"temp_{safe_filename}"
    final_path = PRESENTATIONS_DIR / safe_filename
    
    with open(temp_path, "wb") as f:
        f.write(content)
    
    # Add watermark (optimized version)
    add_watermark_to_pdf_fast(temp_path, final_path, LOGO_PATH)
    
    # Remove temp file
    if temp_path.exists():
        temp_path.unlink()
    
    # Read watermarked file and store in GridFS for persistence
    if final_path.exists():
        with open(final_path, "rb") as f:
            watermarked_content = f.read()
    else:
        watermarked_content = content
    
    # Create presentation record
    presentation = Presentation(
        title=title,
        module=module,
        submodule=submodule,
        description=description,
        file_path=safe_filename,
        file_name=original_filename
    )
    
    await db.presentations.insert_one(presentation.dict())
    
    # Store file in GridFS
    await fs_bucket.upload_from_stream(
        original_filename,
        watermarked_content,
        metadata={"presentation_id": presentation.id}
    )
    
    return {"message": "Presentación subida correctamente", "id": presentation.id}

@api_router.post("/admin/presentations/{presentation_id}/reupload")
async def reupload_presentation_file(
    presentation_id: str,
    file: UploadFile = File(...),
    admin_user: User = Depends(get_admin_user)
):
    """Re-upload a PDF file for an existing presentation record"""
    presentation = await db.presentations.find_one({"id": presentation_id})
    if not presentation:
        raise HTTPException(status_code=404, detail="Presentación no encontrada")
    
    if not file.filename.lower().endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Solo se permiten archivos PDF")
    
    content = await file.read()
    
    # Save and watermark
    safe_filename = presentation.get("file_path", f"{str(uuid.uuid4())}.pdf")
    temp_path = PRESENTATIONS_DIR / f"temp_{safe_filename}"
    final_path = PRESENTATIONS_DIR / safe_filename
    
    with open(temp_path, "wb") as f:
        f.write(content)
    
    add_watermark_to_pdf_fast(temp_path, final_path, LOGO_PATH)
    
    if temp_path.exists():
        temp_path.unlink()
    
    # Read watermarked file and store in GridFS
    if final_path.exists():
        with open(final_path, "rb") as f:
            watermarked_content = f.read()
    else:
        watermarked_content = content
    
    # Delete old GridFS file if exists
    old_file = await db.presentations_fs.files.find_one({"metadata.presentation_id": presentation_id})
    if old_file:
        await fs_bucket.delete(old_file["_id"])
    
    # Store new file in GridFS
    await fs_bucket.upload_from_stream(
        presentation.get("file_name", file.filename),
        watermarked_content,
        metadata={"presentation_id": presentation_id}
    )
    
    return {"message": "Archivo re-subido correctamente"}


@api_router.delete("/admin/presentations/{presentation_id}")
async def delete_presentation(
    presentation_id: str,
    admin_user: User = Depends(get_admin_user)
):
    """Delete a presentation (admin only)"""
    presentation = await db.presentations.find_one({"id": presentation_id})
    if not presentation:
        raise HTTPException(status_code=404, detail="Presentación no encontrada")
    
    # Delete file from disk
    file_path = PRESENTATIONS_DIR / presentation["file_path"]
    if file_path.exists():
        file_path.unlink()
    
    # Delete file from GridFS
    old_file = await db.presentations_fs.files.find_one({"metadata.presentation_id": presentation_id})
    if old_file:
        await fs_bucket.delete(old_file["_id"])
    
    # Delete from database
    await db.presentations.delete_one({"id": presentation_id})
    
    return {"message": "Presentación eliminada"}

@api_router.get("/admin/presentations/all")
async def get_all_presentations(admin_user: User = Depends(get_admin_user)):
    """Get all presentations for admin panel"""
    presentations = await db.presentations.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    # Check which presentations have file data in GridFS
    pres_ids_with_files = set()
    async for doc in db.presentations_fs.files.find({}, {"metadata.presentation_id": 1, "_id": 0}):
        pid = doc.get("metadata", {}).get("presentation_id")
        if pid:
            pres_ids_with_files.add(pid)
    
    for p in presentations:
        file_path = PRESENTATIONS_DIR / p.get("file_path", "")
        p["file_available"] = (
            p.get("storage_type") == "external"
            or p["id"] in pres_ids_with_files
            or file_path.exists()
        )
    
    return presentations


class MigrateRequest(BaseModel):
    presentation_ids: Optional[List[str]] = None
    batch_size: int = 10

@api_router.post("/admin/presentations/migrate-to-storage")
async def migrate_presentations_to_storage(
    req: MigrateRequest,
    admin_user: User = Depends(get_admin_user),
):
    """Migrate presentations from GridFS to external Object Storage."""
    presentation_ids = req.presentation_ids
    batch_size = req.batch_size
    if presentation_ids:
        query = {"id": {"$in": presentation_ids}, "storage_type": {"$ne": "external"}}
    else:
        query = {"storage_type": {"$ne": "external"}}
    
    to_migrate = await db.presentations.find(query, {"_id": 0}).to_list(batch_size)
    
    results = {"migrated": [], "failed": [], "skipped": []}
    
    for pres in to_migrate:
        pres_id = pres["id"]
        title = pres.get("title", pres_id)
        
        # Read from GridFS
        gridfs_meta = await db.presentations_fs.files.find_one(
            {"metadata.presentation_id": pres_id}
        )
        
        pdf_data = None
        if gridfs_meta:
            try:
                grid_out = await fs_bucket.open_download_stream(gridfs_meta["_id"])
                pdf_data = await grid_out.read()
            except Exception as e:
                # GridFS read failed — check if already in Object Storage
                expected_key = f"proyecto-residente/presentations/{pres_id}.pdf"
                existing_data = presentation_storage.get_pdf(expected_key)
                if existing_data and existing_data[:5] == b"%PDF-":
                    # File already exists in Object Storage, just update metadata
                    await db.presentations.update_one(
                        {"id": pres_id},
                        {"$set": {
                            "storage_type": "external",
                            "pdf_key": expected_key,
                            "pdf_size": len(existing_data),
                            "migrated_at": datetime.now(timezone.utc).isoformat(),
                        }}
                    )
                    results["migrated"].append({
                        "id": pres_id, "title": title,
                        "size": len(existing_data), "key": expected_key,
                        "source": "existing_object_storage",
                    })
                    continue
                results["failed"].append({"id": pres_id, "title": title, "error": f"GridFS read: {e}"})
                continue
        
        if not gridfs_meta:
            # No GridFS file — check if already in Object Storage
            expected_key = f"proyecto-residente/presentations/{pres_id}.pdf"
            existing_data = presentation_storage.get_pdf(expected_key)
            if existing_data and existing_data[:5] == b"%PDF-":
                await db.presentations.update_one(
                    {"id": pres_id},
                    {"$set": {
                        "storage_type": "external",
                        "pdf_key": expected_key,
                        "pdf_size": len(existing_data),
                        "migrated_at": datetime.now(timezone.utc).isoformat(),
                    }}
                )
                results["migrated"].append({
                    "id": pres_id, "title": title,
                    "size": len(existing_data), "key": expected_key,
                    "source": "existing_object_storage",
                })
                continue
            results["skipped"].append({"id": pres_id, "title": title, "reason": "No GridFS file and not in Object Storage"})
            continue
        
        # Validate it's a real PDF
        if not pdf_data[:5] == b"%PDF-":
            results["failed"].append({"id": pres_id, "title": title, "error": "Not a valid PDF"})
            continue
        
        # Upload to Object Storage (exact bytes, no modification)
        upload_result = presentation_storage.put_pdf(pres_id, pdf_data)
        if not upload_result:
            results["failed"].append({"id": pres_id, "title": title, "error": "Upload failed"})
            continue
        
        # Validate uploaded size matches
        if upload_result.get("size") != len(pdf_data):
            results["failed"].append({
                "id": pres_id, "title": title,
                "error": f"Size mismatch: uploaded {upload_result.get('size')} vs original {len(pdf_data)}"
            })
            continue
        
        # Update presentation metadata (add new fields, keep everything else)
        await db.presentations.update_one(
            {"id": pres_id},
            {"$set": {
                "storage_type": "external",
                "pdf_key": upload_result["path"],
                "pdf_size": len(pdf_data),
                "migrated_at": datetime.now(timezone.utc).isoformat(),
            }}
        )
        
        results["migrated"].append({
            "id": pres_id,
            "title": title,
            "size": len(pdf_data),
            "key": upload_result["path"],
        })
    
    remaining = await db.presentations.count_documents({"storage_type": {"$ne": "external"}})
    results["remaining"] = remaining
    results["total_migrated_now"] = len(results["migrated"])
    
    return results


class SyncStorageRequest(BaseModel):
    entries: List[Dict[str, Any]]

@api_router.post("/admin/presentations/sync-storage-metadata")
async def sync_storage_metadata(
    req: SyncStorageRequest,
    admin_user: User = Depends(get_admin_user),
):
    """Sync Object Storage metadata for presentations that are already uploaded.
    
    Use this when PDFs are already in Object Storage but the DB doesn't have
    the storage_type/pdf_key fields (e.g. after deploy from a different environment).
    
    Each entry: {"id": "pres-id", "pdf_key": "path/in/storage", "pdf_size": 12345}
    """
    updated = 0
    failed = []
    
    for entry in req.entries:
        pres_id = entry.get("id")
        pdf_key = entry.get("pdf_key")
        pdf_size = entry.get("pdf_size", 0)
        
        if not pres_id or not pdf_key:
            failed.append({"id": pres_id, "error": "Missing id or pdf_key"})
            continue
        
        # Verify the presentation exists
        pres = await db.presentations.find_one({"id": pres_id})
        if not pres:
            failed.append({"id": pres_id, "error": "Presentation not found"})
            continue
        
        # Update metadata (trust the caller — files were already uploaded)
        await db.presentations.update_one(
            {"id": pres_id},
            {"$set": {
                "storage_type": "external",
                "pdf_key": pdf_key,
                "pdf_size": pdf_size,
                "migrated_at": datetime.now(timezone.utc).isoformat(),
            }}
        )
        updated += 1
    
    return {"updated": updated, "failed": failed, "total_entries": len(req.entries)}


class RollbackRequest(BaseModel):
    presentation_ids: Optional[List[str]] = None

@api_router.post("/admin/presentations/rollback-storage")
async def rollback_storage_migration(
    req: RollbackRequest,
    admin_user: User = Depends(get_admin_user),
):
    """Rollback: remove external storage fields, revert to GridFS serving."""
    presentation_ids = req.presentation_ids
    if presentation_ids:
        query = {"id": {"$in": presentation_ids}, "storage_type": "external"}
    else:
        query = {"storage_type": "external"}
    
    result = await db.presentations.update_many(
        query,
        {"$unset": {"storage_type": "", "pdf_key": "", "pdf_size": "", "migrated_at": ""}}
    )
    
    return {
        "rolled_back": result.modified_count,
        "message": f"Reverted {result.modified_count} presentations to GridFS serving"
    }


# =============================================================================
# ENARM MATCH ENDPOINTS
# =============================================================================

class ENARMMatchProgress(BaseModel):
    paso: int
    respuestasPares: Optional[List[Dict]] = []
    tolerancias: Optional[Dict] = {}
    subsObjetivo: Optional[List[str]] = []
    destinoSeleccionado: Optional[Dict] = None
    confirmacionTronco: Optional[str] = None
    noTolerancias: Optional[List[str]] = []

class ENARMMatchResult(BaseModel):
    respuestasPares: List[Dict]
    tolerancias: Dict
    subsObjetivo: List[str]
    destinoSeleccionado: Optional[str] = None
    confirmacionTronco: Optional[str] = None
    noTolerancias: List[str]
    perfilUsuario: Dict
    flags: Dict
    top2: List[Dict]
    confianza: Dict
    completado_at: str

@api_router.get("/enarm-match/progress")
async def get_enarm_match_progress(current_user: User = Depends(get_current_user)):
    """Get user's ENARM Match progress"""
    progress = await db.enarm_match_progress.find_one(
        {"user_id": current_user.id},
        {"_id": 0}
    )
    if progress:
        return progress
    return {}

@api_router.post("/enarm-match/progress")
async def save_enarm_match_progress(
    progress: ENARMMatchProgress,
    current_user: User = Depends(get_current_user)
):
    """Save user's ENARM Match progress"""
    progress_data = progress.dict()
    progress_data["user_id"] = current_user.id
    progress_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.enarm_match_progress.update_one(
        {"user_id": current_user.id},
        {"$set": progress_data},
        upsert=True
    )
    
    return {"message": "Progreso guardado"}

@api_router.delete("/enarm-match/progress")
async def delete_enarm_match_progress(current_user: User = Depends(get_current_user)):
    """Delete user's ENARM Match progress"""
    await db.enarm_match_progress.delete_one({"user_id": current_user.id})
    return {"message": "Progreso eliminado"}

@api_router.post("/enarm-match/results")
async def save_enarm_match_results(
    result: ENARMMatchResult,
    current_user: User = Depends(get_current_user)
):
    """Save user's ENARM Match results"""
    result_data = result.dict()
    result_data["id"] = str(uuid.uuid4())
    result_data["user_id"] = current_user.id
    result_data["created_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.enarm_match_results.insert_one(result_data)
    
    return {"message": "Resultados guardados", "id": result_data["id"]}

@api_router.get("/enarm-match/results")
async def get_enarm_match_results(current_user: User = Depends(get_current_user)):
    """Get user's ENARM Match results history"""
    results = await db.enarm_match_results.find(
        {"user_id": current_user.id},
        {"_id": 0}
    ).sort("created_at", -1).to_list(50)
    
    return results

@api_router.get("/enarm-match/results/{result_id}")
async def get_enarm_match_result_by_id(
    result_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get a specific ENARM Match result"""
    result = await db.enarm_match_results.find_one(
        {"id": result_id, "user_id": current_user.id},
        {"_id": 0}
    )
    if not result:
        raise HTTPException(status_code=404, detail="Resultado no encontrado")
    return result

# Admin endpoint to get all ENARM Match results (for analytics)
@api_router.get("/admin/enarm-match/analytics")
async def get_enarm_match_analytics(admin_user: User = Depends(get_admin_user)):
    """Get ENARM Match analytics for admin"""
    # Total results
    total_results = await db.enarm_match_results.count_documents({})
    
    # Results by specialty (top recommendations)
    pipeline = [
        {"$unwind": "$top2"},
        {"$group": {
            "_id": "$top2.especialidadId",
            "count": {"$sum": 1},
            "nombre": {"$first": "$top2.nombre"}
        }},
        {"$sort": {"count": -1}},
        {"$limit": 10}
    ]
    top_specialties = await db.enarm_match_results.aggregate(pipeline).to_list(10)
    
    # Confidence distribution
    confidence_pipeline = [
        {"$group": {
            "_id": "$confianza.nivel",
            "count": {"$sum": 1}
        }}
    ]
    confidence_dist = await db.enarm_match_results.aggregate(confidence_pipeline).to_list(10)
    
    return {
        "total_results": total_results,
        "top_specialties": top_specialties,
        "confidence_distribution": confidence_dist
    }

# =============================================================================

# Include the planner module
from planner import planner_router, set_db as planner_set_db
planner_set_db(db)
app.include_router(planner_router)

# Include presentation storage module
import presentation_storage

# Include the MercadoPago payment module
from mercadopago_routes import mp_router, set_db as mp_set_db, set_activation_callback
mp_set_db(db)
api_router.include_router(mp_router)

# Include the support module
from support import (
    create_ticket_handler, get_my_tickets_handler, get_all_tickets_handler,
    admin_reply_handler, user_reply_handler, get_unread_count_admin_handler,
    get_unread_count_user_handler, mark_ticket_read_admin_handler
)
import support as support_module
support_module.set_db(db)

class CreateTicketRequest(BaseModel):
    message: str
    category: Optional[str] = "general"

class ReplyRequest(BaseModel):
    message: str

@api_router.post("/support/create")
async def create_support_ticket(req: CreateTicketRequest, current_user: User = Depends(get_current_user)):
    return await create_ticket_handler(req.message, req.category, {
        "user_id": current_user.id,
        "full_name": current_user.full_name,
        "email": current_user.email
    })

@api_router.get("/support/my-tickets")
async def get_my_support_tickets(current_user: User = Depends(get_current_user)):
    return await get_my_tickets_handler({
        "user_id": current_user.id
    })

@api_router.get("/support/all-tickets")
async def get_all_support_tickets(admin: User = Depends(get_admin_user)):
    return await get_all_tickets_handler()

@api_router.post("/support/reply/{ticket_id}")
async def admin_reply_to_ticket(ticket_id: str, req: ReplyRequest, admin: User = Depends(get_admin_user)):
    return await admin_reply_handler(ticket_id, req.message)

@api_router.post("/support/user-reply/{ticket_id}")
async def user_reply_to_ticket(ticket_id: str, req: ReplyRequest, current_user: User = Depends(get_current_user)):
    return await user_reply_handler(ticket_id, req.message, {
        "user_id": current_user.id
    })

@api_router.get("/support/unread-count")
async def get_support_unread_count(current_user: User = Depends(get_current_user)):
    if current_user.is_admin:
        return await get_unread_count_admin_handler()
    return await get_unread_count_user_handler({
        "user_id": current_user.id
    })

@api_router.put("/support/mark-read/{ticket_id}")
async def mark_ticket_read(ticket_id: str, admin: User = Depends(get_admin_user)):
    return await mark_ticket_read_admin_handler(ticket_id)

# =============================================================================
# ADMIN ACCOUNT CHANGE (with email verification)
# =============================================================================
from admin_account import (
    send_verification_email, create_change_request, confirm_change,
    set_db as admin_account_set_db
)
admin_account_set_db(db)

ADMIN_VERIFY_EMAIL = os.environ.get('ADMIN_EMAIL', 'maggiani@gmail.com')

class AdminChangeRequest(BaseModel):
    change_type: str  # "name" or "password"
    new_value: str

@api_router.post("/admin/request-account-change")
async def request_account_change(req: AdminChangeRequest, admin: User = Depends(get_admin_user)):
    if req.change_type not in ("name", "password"):
        raise HTTPException(status_code=400, detail="Tipo de cambio invalido")
    if not req.new_value or len(req.new_value.strip()) < 3:
        raise HTTPException(status_code=400, detail="El valor debe tener al menos 3 caracteres")

    token = await create_change_request(req.change_type, req.new_value.strip(), admin.id)
    
    app_url = os.environ.get('FRONTEND_URL', os.environ.get('APP_URL', ''))
    
    # Use the API URL for the confirmation endpoint
    confirm_url = f"{app_url}/api/admin/confirm-change/{token}"
    
    desc = f"Cambiar nombre a: {req.new_value}" if req.change_type == "name" else "Cambiar contrasena de administrador"
    
    try:
        send_verification_email(ADMIN_VERIFY_EMAIL, confirm_url, desc)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al enviar correo: {str(e)}")
    
    return {"ok": True, "message": f"Se envio un enlace de confirmacion a {ADMIN_VERIFY_EMAIL[:3]}***@gmail.com"}

@app.get("/api/admin/confirm-change/{token}")
async def confirm_account_change(token: str):
    from fastapi.responses import HTMLResponse
    change_type, error = await confirm_change(token)
    if error:
        html = f"""
        <html><body style="font-family:Arial;max-width:500px;margin:80px auto;text-align:center;">
        <h2 style="color:red;">Error</h2>
        <p>{error}</p>
        </body></html>
        """
        return HTMLResponse(content=html, status_code=400)
    
    label = "nombre" if change_type == "name" else "contrasena"
    html = f"""
    <html><body style="font-family:Arial;max-width:500px;margin:80px auto;text-align:center;">
    <h2 style="color:green;">Cambio confirmado</h2>
    <p>El {label} del administrador ha sido actualizado correctamente.</p>
    <p style="margin-top:30px;"><a href="/" style="background:#000;color:#fff;padding:12px 30px;border-radius:8px;text-decoration:none;">Volver a la app</a></p>
    </body></html>
    """
    return HTMLResponse(content=html)

# =============================================================================
# POST-PAYMENT EMAIL ACTIVATION
# =============================================================================
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

def send_activation_email(to_email: str, user_name: str, activation_url: str, user_password: str = ""):
    """Send activation email after successful payment."""
    gmail_user = os.environ.get('GMAIL_USER')
    gmail_password = os.environ.get('GMAIL_APP_PASSWORD')
    if not gmail_user or not gmail_password:
        raise Exception("Gmail credentials not configured")

    msg = MIMEMultipart('alternative')
    msg['Subject'] = 'Proyecto Residente - Activa tu Cuenta'
    msg['From'] = gmail_user
    msg['To'] = to_email

    password_section = ""
    if user_password:
        password_section = f'<p style="font-size: 14px; color: #166534; margin: 4px 0 0 0;">Contrasena: {user_password}</p>'

    html = f"""
    <html>
    <body style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px; background-color: #f5f5f5;">
        <div style="background: #ffffff; border-radius: 16px; padding: 40px; box-shadow: 0 2px 12px rgba(0,0,0,0.08);">
            <div style="text-align: center; margin-bottom: 30px;">
                <h1 style="color: #000; font-size: 24px; margin: 0;">Proyecto Residente</h1>
                <p style="color: #666; font-size: 14px; margin-top: 8px;">Tu pago ha sido confirmado</p>
            </div>
            <div style="background: #f9f9f9; border-radius: 12px; padding: 24px; margin-bottom: 24px;">
                <p style="font-size: 16px; color: #333; margin: 0 0 12px 0;">Hola <strong>{user_name}</strong>,</p>
                <p style="font-size: 14px; color: #555; line-height: 1.6; margin: 0;">
                    Tu pago se ha procesado exitosamente. Para completar tu registro y comenzar a usar la plataforma, 
                    activa tu cuenta haciendo clic en el boton de abajo.
                </p>
            </div>
            <div style="background: #f0fdf4; border-radius: 12px; padding: 16px; margin-bottom: 24px; border: 1px solid #bbf7d0;">
                <p style="font-size: 14px; color: #166534; margin: 0;"><strong>Datos de tu cuenta:</strong></p>
                <p style="font-size: 14px; color: #166534; margin: 8px 0 0 0;">Correo: {to_email}</p>
                <p style="font-size: 14px; color: #166534; margin: 4px 0 0 0;">Nombre: {user_name}</p>
                {password_section}
                <p style="font-size: 14px; color: #166534; margin: 4px 0 0 0;">Plan: 6 meses de acceso completo</p>
            </div>
            <div style="text-align: center; margin: 32px 0;">
                <a href="{activation_url}" 
                   style="background: #000; color: #fff; padding: 16px 48px; border-radius: 50px; text-decoration: none; font-size: 16px; font-weight: bold; display: inline-block;">
                    Activar Cuenta
                </a>
            </div>
            <p style="color: #999; font-size: 12px; text-align: center; margin-top: 24px;">
                Este enlace es valido por 7 dias. Si no solicitaste este registro, ignora este correo.
            </p>
        </div>
    </body>
    </html>
    """
    msg.attach(MIMEText(html, 'html'))

    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
        server.login(gmail_user, gmail_password)
        server.sendmail(gmail_user, to_email, msg.as_string())


async def generate_activation_token(user_id: str) -> str:
    """Generate and store an activation token for a user."""
    token = str(uuid.uuid4())
    await db.activation_tokens.insert_one({
        "token": token,
        "user_id": user_id,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "expires_at": (datetime.now(timezone.utc) + timedelta(days=7)).isoformat(),
        "used": False
    })
    return token


async def send_activation_for_user(user_id: str):
    """Generate activation token and send email to user."""
    user = await db.users.find_one({"id": user_id})
    if not user:
        print(f"[ACTIVATION EMAIL] User {user_id} not found")
        return False
    
    token = await generate_activation_token(user_id)
    frontend_url = os.environ.get('FRONTEND_URL', os.environ.get('APP_URL', ''))
    activation_url = f"{frontend_url}/api/auth/activate?token={token}"
    
    user_password = user.get("temp_password", "")
    
    try:
        send_activation_email(user["email"], user.get("full_name", ""), activation_url, user_password)
        print(f"[ACTIVATION EMAIL] Sent successfully to {user['email']}")
        if user_password:
            await db.users.update_one({"id": user_id}, {"$unset": {"temp_password": ""}})
        return True
    except Exception as e:
        print(f"[ACTIVATION EMAIL] Error sending to {user.get('email', 'unknown')}: {e}")
        import traceback
        traceback.print_exc()
        return False

# Wire up activation callback for MercadoPago routes
set_activation_callback(send_activation_for_user)


@app.get("/api/auth/activate")
async def activate_account(token: str):
    """Activate user account via email token."""
    from fastapi.responses import HTMLResponse
    
    record = await db.activation_tokens.find_one({"token": token, "used": False})
    if not record:
        html = """
        <html><body style="font-family:Arial;max-width:500px;margin:80px auto;text-align:center;">
        <h2 style="color:red;">Enlace invalido</h2>
        <p>Este enlace de activacion ya fue usado o no es valido.</p>
        <p style="margin-top:20px;"><a href="/" style="color:#000;text-decoration:underline;">Ir al inicio</a></p>
        </body></html>
        """
        return HTMLResponse(content=html, status_code=400)
    
    expires_at = datetime.fromisoformat(record["expires_at"])
    if datetime.now(timezone.utc) > expires_at:
        html = """
        <html><body style="font-family:Arial;max-width:500px;margin:80px auto;text-align:center;">
        <h2 style="color:red;">Enlace expirado</h2>
        <p>Este enlace de activacion ha expirado. Contacta al administrador.</p>
        <p style="margin-top:20px;"><a href="/" style="color:#000;text-decoration:underline;">Ir al inicio</a></p>
        </body></html>
        """
        return HTMLResponse(content=html, status_code=400)
    
    user_id = record["user_id"]
    
    # Activate the user
    subscription_expires = datetime.now(timezone.utc) + timedelta(days=180)
    await db.users.update_one(
        {"id": user_id},
        {
            "$set": {
                "is_approved": True,
                "payment_status": "completed",
                "subscription_expires": subscription_expires.isoformat(),
                "activated_at": datetime.now(timezone.utc).isoformat(),
                "account_type": "paid"
            }
        }
    )
    
    # Mark token as used
    await db.activation_tokens.update_one(
        {"token": token},
        {"$set": {"used": True}}
    )
    
    frontend_url = os.environ.get('FRONTEND_URL', os.environ.get('APP_URL', ''))
    html = f"""
    <html><body style="font-family:Arial;max-width:500px;margin:80px auto;text-align:center;">
    <div style="background:#f0fdf4;border-radius:16px;padding:32px;border:1px solid #bbf7d0;">
        <h2 style="color:#166534;margin-bottom:16px;">Cuenta Activada</h2>
        <p style="color:#333;font-size:16px;">Tu cuenta ha sido activada exitosamente. Ya puedes iniciar sesion.</p>
    </div>
    <p style="margin-top:30px;">
        <a href="{frontend_url}/login" style="background:#000;color:#fff;padding:14px 40px;border-radius:50px;text-decoration:none;font-size:16px;font-weight:bold;display:inline-block;">
            Iniciar Sesion
        </a>
    </p>
    </body></html>
    """
    return HTMLResponse(content=html)


@api_router.post("/auth/resend-activation/{user_id}")
async def resend_activation(user_id: str):
    """Resend activation email for a user."""
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    
    if user.get("account_type") == "paid":
        raise HTTPException(status_code=400, detail="La cuenta ya esta activada")
    
    success = await send_activation_for_user(user_id)
    if success:
        return {"message": "Correo de activacion reenviado exitosamente."}
    else:
        raise HTTPException(status_code=500, detail="Error al enviar el correo de activacion.")


# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# =============================================================================
# APP SETTINGS (Dynamic pricing)
# =============================================================================

@app.get("/api/settings/price")
async def get_subscription_price():
    """Public endpoint to get current subscription price."""
    settings = await db.app_settings.find_one({"key": "subscription_price"}, {"_id": 0})
    if settings:
        return {"price": settings["value"], "currency": "MXN", "duration_months": 6}
    return {"price": 1500, "currency": "MXN", "duration_months": 6}

@app.get("/api/admin/settings/price")
async def admin_get_price(admin_user: User = Depends(get_admin_user)):
    """Admin endpoint to get current price."""
    settings = await db.app_settings.find_one({"key": "subscription_price"}, {"_id": 0})
    price = settings["value"] if settings else 1500
    return {"price": price, "currency": "MXN", "duration_months": 6}

class PriceUpdateRequest(BaseModel):
    price: float

@app.post("/api/admin/settings/price")
async def admin_set_price(req: PriceUpdateRequest, admin_user: User = Depends(get_admin_user)):
    """Admin endpoint to update subscription price."""
    if req.price <= 0:
        raise HTTPException(status_code=400, detail="El precio debe ser mayor a 0")
    await db.app_settings.update_one(
        {"key": "subscription_price"},
        {"$set": {"key": "subscription_price", "value": req.price, "updated_at": datetime.now(timezone.utc).isoformat(), "updated_by": admin_user.id}},
        upsert=True
    )
    return {"message": "Precio actualizado", "price": req.price}


# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_all():
    """Cleanup on shutdown"""
    scheduler.shutdown(wait=False)
    client.close()
    logger.info("🛑 Application shutdown complete")